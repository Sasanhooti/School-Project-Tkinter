from tkinter import *
from tkinter.constants import *
import sys ; sys.setrecursionlimit(sys.getrecursionlimit() * 5) 
import sqlite3
import calendar 
from PIL import ImageTk, Image
from deep_translator import GoogleTranslator
import tkinter.messagebox as messagebox
import wolframalpha
import os  
import json as js
import re
from  tkinter import ttk
i=0
#################################################################################

from tkinter import *
from tkinter import ttk
from PIL import Image, ImageTk
import customtkinter
import threading
conn = sqlite3.connect("user1.db")
c = conn.cursor()
c.execute("CREATE TABLE IF NOT EXISTS user_info_maneger(name TEXT, last_name TEXT, email TEXT,password TEXT,  time TEXT, number TEXT,born TEXT, name_school TEXT)")
c.execute("CREATE TABLE IF NOT EXISTS user_info_techer(name TEXT, last_name TEXT, email TEXT,password TEXT,  time TEXT, number TEXT,born TEXT, name_maneger TEXT, last_name_maneger TEXT)")
c.execute("CREATE TABLE IF NOT EXISTS user_profile_maneger(name TEXT, last_name TEXT, email TEXT,  time TEXT, number TEXT,born TEXT, name_school TEXT)")
c.execute("CREATE TABLE IF NOT EXISTS user_profile_techer(name TEXT, last_name TEXT, email TEXT ,  time TEXT, number TEXT,born TEXT, name_maneger TEXT, last_name_maneger TEXT)")
c.execute("CREATE TABLE IF NOT EXISTS Languge(user TEXT)")
def load():
    root = Tk()
    root.geometry('530x600')
    root.configure(bg="#fff")
    root.title("در حال بالا آمدن")
    root.iconbitmap(bitmap = 'Icons\\icon.ico')
    welcom_label = Label(text="خوش آمدید به روبوسان",bg="#fff",font=("B Homa",15), fg="#57a1f8")
    welcom_label.place(x=190,y=25)
    import random
    names = [ 2, 4, 5, 8, 9, 10,3]
    selected_names = random.sample(names, 1)
    image = Image.open(f'lomge/{selected_names}.png')
    
    
    img = ImageTk.PhotoImage(image)
    Label(root,image=img,bg="#fff").place(x=70,y=65)
    program_label = Label(root,bg="#fff", text=".... سازی آماده درحال", font=("B Homa" , 13), fg="#57a1f8")
    program_label.place(x=190, y=450)
    
    program = ttk.Style()
    program.theme_use("clam")
    program.configure("red.Horizontal.TProgressbar",background="#108cff")
    
    program= ttk.Progressbar(root, orient=HORIZONTAL, length=400, mode="determinate", style="red.Horizontal.TProgressbar")
    program.place(x=60, y=500) 
    
    def top():
        root.withdraw()
        root.destroy()
        techer()
    def load():
        global i
        if i<=10:
            txt=((str(10*i))+ "%"+".... درحال آماده سازی")
            program_label.config(text=txt)
            program_label.after(600, load)
            program['value'] = 10*i
            i +=1
        else:
            top()
    load()
    root.resizable(False,False)
    root.mainloop()
######ساخت صفحه ورود معلم
def techer():
    
    def log():
            
            email = user.get()
            password = code.get()
            conn = sqlite3.connect("user1.db")
            c = conn.cursor()
            c.execute("SELECT * FROM user_info_techer WHERE email = ? AND password = ?",(email,password))
            results = c.fetchall()
            
                
            
            if results:
                for i in results:
                    userNamelog = str(i[0])
                    userLastNamelog = str(i[1])
                    userEmaillog = str(i[2])
                    usertime = str(i[4])
                    usernumber = str(i[5])
                    userborn = str(i[6])
                    name_maneger = str(i[7])
                    last_name_maneger = str(i[8])
                    
                                    
                c.execute("INSERT INTO user_profile_techer VALUES (:nameProfile, :lastnameProfile, :emailProfile, :timeProfile, :numberProfile, :bornProfile, :name_manegerProfile, :last_name_manegerProfile)",
                        {
                            "nameProfile":userNamelog,
                            "lastnameProfile":userLastNamelog,
                            "emailProfile":userEmaillog,
                            "timeProfile":usertime,
                            "numberProfile":usernumber,
                            "bornProfile":userborn,
                            "name_manegerProfile": name_maneger,
                            "last_name_manegerProfile": last_name_maneger
                        })
                conn.commit()
                conn.close()
                messagebox.showinfo("Log in","خوش آمدید {} {}".format(i[0], i[1]))
                root.destroy()
                StartClass()

            else:
                messagebox.showerror("Warning", "Fill out every single entery")
    root =Tk() 
    root.title("ورود معلم")
    root.geometry('925x500+300+200')
    root.iconbitmap(bitmap = 'Icons\\icon.ico')
    root.configure(bg="#fff")
    root.resizable(False,False)
    import random
    names = [ 2, 4, 5,  8, 9, 10, 3]
    selected_names = random.sample(names, 1)
    image = Image.open(f'lomge/{selected_names}.png')
    img = ImageTk.PhotoImage(image)
    Label(root,image=img,bg='white').place(x=50,y=50)
    
    frame=Frame(root,width=350,height=350,bg="white")
    frame.place(x=480,y=70)
    heading=Label(frame,text="صفحه ورود",fg="#57a1f8",bg="white",border=0,font=('B Kamran Outline',43))
    heading.place(x=90,y=-44)
    #######################
    def on_enter(e):
        if (user.get())=="نام کاربری":
            user.delete(0, 'end')
    def on_leave(e):
        if len(user.get())==0:
            Label(frame,text="نام کاربری را وارد کن",fg="red",bg="white",border=0,font=("B Homa",13)).place(x=25, y=117)
            user.insert(0, "نام کاربری")
        else:
            Label(frame,text="تایید شد",justify=RIGHT,width=10,fg="green",bg="white",border=0,font=("B Homa",13),padx=50).place(x=-22, y=117)
    user = Entry(frame,width=40,fg='black',border=0,bg='white',justify=RIGHT,font=("B Homa",13))
    user.place(x=30,y=80)
    user.insert(0,"نام کاربری")
    user.bind("<FocusIn>",on_enter)
    user.bind("<FocusOut>",on_leave)
    user.bind()
    Frame(frame,width=295,height=2,bg='black').place(x=25,y=107)
    #######################
    def on_enter(e):
        if (code.get())=="رمز ورود":
            code.delete(0, 'end')
    def on_leave(e):
    
        if len(code.get())==0:
           Label(frame,text="رمز را وارد کن",fg="red",bg="white",border=0,font=("B Homa",13)).place(x=25, y=185)
           code.insert(0, "رمز ورود")
        elif len(code.get())<8:
            Label(frame,text="8 حداقل کارکتر تعداد",fg="red",bg="white",border=0,font=("B Homa",13)).place(x=25, y=185)
        else:
            Label(frame,text="تایید شد",justify=RIGHT,width=10,fg="green",bg="white",border=0,font=("B Homa",13),padx=50).place(x=-22, y=183)
    def manege():
        root.destroy()
        maneger()
    code = Entry(frame,width=40,fg='black',border=0,bg='white',justify=RIGHT,font=("B Homa",13))
    code.place(x=30,y=150)
    code.insert(0,"رمز ورود")
    code.bind("<FocusIn>",on_enter)
    code.bind("<FocusOut>",on_leave)
    Frame(frame,width=295,height=2,bg='black').place(x=25,y=177)
    ########################

    Button(frame,width=39,pady=7,text="ورود",bg='#57a1f8',fg='white',border=0,command=log).place(x=35,y=255)
    label = Label(frame,text="مدیر هستی ؟",fg='black',bg='white',font=("B Homa",9))
    label.place(x=200,y=290)
    sing = Button(frame,width=6,text="مدیر",border=0,bg='white',cursor='hand2',fg='#57a1f8',command=manege)
    sing.place(x=155,y=290)
    ########
    def helpme():
        root1=Tk()
        root1.title("Help") 
        root1.iconbitmap(bitmap = 'Icons\\icon.ico')
        root1.geometry("350x110")
        Label(root1, text=("""سلام ...
            شما برای ورود باید در کادر اول نام کاربری   
            در کادر دوم پسوردی که مدیر دبیرستانی که
            انجا تدریس میکنید برایتان ساخته وارد شوید .
            اگر نمیدانید از مدیرتان درخواست کنید که بهتان بدهد""")).pack()

        root1.mainloop() 
    label = Label(frame,text="نیاز به کمک داری ؟",fg='black',bg='white',font=("B Homa",9))
    label.place(x=200,y=310)
    sing2 = Button(frame,width=6,text="کمک",border=0,bg='white',cursor='hand2',fg='#57a1f8' ,command=helpme)
    sing2.place(x=155,y=310)
    
    root.mainloop()
################################################################################
def maneger():
    def log():
            
            email = user.get()
            password = code.get()
            conn = sqlite3.connect("user1.db")
            c = conn.cursor()
            c.execute("SELECT * FROM user_info_maneger WHERE email = ? AND password = ?",(email,password))
            results = c.fetchall()
            
                
            
            if results:
                for i in results:
                    userNamelog = str(i[0])
                    userLastNamelog = str(i[1])
                    userEmaillog = str(i[2])
                    usertime = str(i[4])
                    usernumber = str(i[5])
                    userborn = str(i[6])
                    username_school = str(i[7])
                    
                                    
                c.execute("INSERT INTO user_profile_maneger VALUES (:nameProfile, :lastnameProfile, :emailProfile, :timeProfile, :numberProfile, :bornProfile, :name_schoolProfile)",
                        {
                            "nameProfile":userNamelog,
                            "lastnameProfile":userLastNamelog,
                            "emailProfile":userEmaillog,
                            "timeProfile":usertime,
                            "numberProfile":usernumber,
                            "bornProfile":userborn ,
                            "name_schoolProfile": username_school,
                        })
                conn.commit()
                conn.close()
                messagebox.showinfo("Log in","خوش آمدید {} {}".format(i[0], i[1]))
                root.destroy()
                managehome()

            else:
                messagebox.showerror("Warning", "Fill out every single entery")
    root =Tk() 
    root.title("ورود مدیر")
    root.geometry('925x500+300+200')
    root.iconbitmap(bitmap = 'Icons\\icon.ico')
    root.configure(bg="#fff")
    root.resizable(False,False)
    import random
    names = [ 2, 4, 5,  8, 9, 10, 3]
    selected_names = random.sample(names, 1)
    image = Image.open(f'lomge/{selected_names}.png')
    img = ImageTk.PhotoImage(image)
    Label(root,image=img,bg='white').place(x=50,y=50)
    
    frame=Frame(root,width=350,height=350,bg="white")
    frame.place(x=480,y=70)
    heading=Label(frame,text="صفحه ورود مدیر",fg="#57a1f8",bg="white",border=0,font=('B Kamran Outline',40))
    heading.place(x=40,y=-40)
    #######################
    def on_enter(e):
        if (user.get())=="نام کاربری":
            user.delete(0, 'end')
    def on_leave(e):
        if len(user.get())==0:
            Label(frame,text="نام کاربری را وارد کن",fg="red",bg="white",border=0,font=("B Homa",13)).place(x=25, y=117)
            user.insert(0, "نام کاربری")
        else:
            Label(frame,text="تایید شد",justify=RIGHT,width=10,fg="green",bg="white",border=0,font=("B Homa",13),padx=50).place(x=-22, y=117)
    user = Entry(frame,width=40,fg='black',border=0,bg='white',justify=RIGHT,font=("B Homa",13))
    user.place(x=30,y=80)
    user.insert(0,"نام کاربری")
    user.bind("<FocusIn>",on_enter)
    user.bind("<FocusOut>",on_leave)
    user.bind()
    Frame(frame,width=295,height=2,bg='black').place(x=25,y=107)
    #######################
    def on_enter(e):
        if (code.get())=="رمز ورود":
            code.delete(0, 'end')
    def on_leave(e):
    
        if len(code.get())==0:
           Label(frame,text="رمز را وارد کن",fg="red",bg="white",border=0,font=("B Homa",13)).place(x=25, y=185)
           code.insert(0, "رمز ورود")
        elif len(code.get())<8:
            Label(frame,text="8 حداقل کارکتر تعداد",fg="red",bg="white",border=0,font=("B Homa",13)).place(x=25, y=185)
        else:
            Label(frame,text="تایید شد",justify=RIGHT,width=10,fg="green",bg="white",border=0,font=("B Homa",13),padx=50).place(x=-22, y=183)
    def register():
        root.destroy()
        register12()
        
    code = Entry(frame,width=40,fg='black',border=0,bg='white',justify=RIGHT,font=("B Homa",13))
    code.place(x=30,y=150)
    code.insert(0,"رمز ورود")
    code.bind("<FocusIn>",on_enter)
    code.bind("<FocusOut>",on_leave)
    Frame(frame,width=295,height=2,bg='black').place(x=25,y=177)
    ########################
    
    import customtkinter
    Button(frame,width=39,pady=7,text='ورود',bg='#57a1f8',fg='white',border=0,command=log).place(x=35,y=255)
    label = Label(frame,text="حساب کاربری نداری ؟",fg='black',bg='white',font=("B Homa",9))
    label.place(x=200,y=290)
    sing = Button(frame,width=6,text='ثبت نام',border=0,bg='white',cursor='hand2',fg='#57a1f8',command=register)
    sing.place(x=155,y=290)
    ########
    def helpme():
        root.destroy()
        techer()
    label = Label(frame,text="معلم هستی ؟",fg='black',bg='white',font=("B Homa",9))
    label.place(x=200,y=310)
    sing2 = Button(frame,width=6,text="معلم",border=0,bg='white',cursor='hand2',fg='#57a1f8' ,command=helpme)
    sing2.place(x=155,y=310)
    
    root.mainloop()
#################################################################################
def register12():
    def register():
            import datetime
            name = nameE.get()
            last_name = last_nameE.get()
            email = user.get()
            password = code.get()
            number= numberE.get()
            age =ageE.get()
            name_school=name_schoolE.get()
            age1 = int(age)
            now = datetime.datetime.now()
            year_born =  str(now.year - age1)
            now = datetime.datetime.today()
            mm = str(now.month)
            dd = str(now.day)
            yyyy = str(now.year)
            hour = str(now.hour)
            mi = str(now.minute)
            ss = str(now.second)
            x=str(mm + "/" + dd + "/" + yyyy + " " + hour + ":" + mi + ":" + ss)
            
                
                
                
            while True:
                        
                        conn = sqlite3.connect("user1.db")
                        c = conn.cursor()
                        c.execute("INSERT INTO user_info_maneger VALUES (:User4name, :User4lastname, :User4email, :User4password, :User4time, :User4number, :User4born, :User4name_school)",
                            {
                                "User4name": name,
                                "User4lastname": last_name,
                                "User4email": email,
                                "User4password": password,
                                "User4time" : x,
                                "User4number":number,
                                "User4born":year_born,
                                "User4name_school":name_school,
                        
                            })
            
                        conn.commit()
                        conn.close()
                        root.destroy()
                        messagebox.showinfo("sing in","شما ثبت نام شدید")
                        maneger()
    
    root =Tk()
    root.title("ثبت نام مدیر")
    root.geometry('350x670')
    root.iconbitmap(bitmap = 'Icons\\icon.ico')
    root.configure(bg="#fff")
    def log():
        root.destroy()
        maneger()
    root.resizable(False,False)
    def sing():
        pass
    
    def helpme():
        pass
    
    frame=root
    heading=Label(frame,text="ثبت نام کردن",fg="#57a1f8",bg="white",border=0,font=("B Kamran Outline",30))
    heading.place(x=70,y=5)
    ###############3
    def on_enter(e):
        if nameE.get() == "نام":
            nameE.delete(0, 'end')
    def on_leave(e):
        if len(nameE.get())==0:
            Label(frame,text="نام را وارد کن",fg="red",bg="white",border=0,font=("B Homa",13)).place(x=25, y=110)
            nameE.insert(0, 'نام')
        else:
            Label(frame,text="تایید شد",justify=RIGHT,width=10,fg="green",bg="white",border=0,font=("B Homa",13),padx=50).place(x=-22, y=110)
    nameE = Entry(frame,width=40,fg='black',border=0,bg='white',justify=RIGHT,font=("B Homa",13))
    nameE.place(x=30,y=80)
    nameE.insert(0,'نام')
    nameE.bind("<FocusIn>",on_enter)
    nameE.bind("<FocusOut>",on_leave)
    nameE.bind()
    Frame(frame,width=295,height=2,bg='black').place(x=25,y=107)
    ###################
    def on_enter(e):
        if last_nameE.get()=="نام خانوادگی":
            last_nameE.delete(0, 'end')
    def on_leave(e):
        if len(last_nameE.get())==0:
            Label(frame,text="نام خانوادگی را وارد کن",fg="red",bg="white",border=0,font=("B Homa",13)).place(x=25, y=180)
            last_nameE.insert(0, "نام خانوادگی")
        else:
            Label(text="تایید شد",width=10,fg="green",bg="white",border=0,font=("B Homa",13),padx=50).place(x=-22, y=180)
    last_nameE = Entry(frame,width=40,fg='black',border=0,bg='white',justify=RIGHT,font=("B Homa",13))
    last_nameE.place(x=30,y=150)
    last_nameE.insert(0,"نام خانوادگی")
    last_nameE.bind("<FocusIn>",on_enter)
    last_nameE.bind("<FocusOut>",on_leave)
    last_nameE.bind()
    Frame(frame,width=295,height=2,bg='black').place(x=25,y=177)
    #######################
    def on_enter(e):
        if (user.get())=="نام کاربری":
            user.delete(0, 'end')
    def on_leave(e):
        if len(user.get())==0:
            Label(frame,text="نام کاربری را وارد کن",fg="red",bg="white",border=0,font=("B Homa",13)).place(x=25, y=257)
            user.insert(0, "نام کاربری")
        else: 
            Label(frame,text="تایید شد",justify=RIGHT,width=10,fg="green",bg="white",border=0,font=("B Homa",13),padx=50).place(x=-22, y=257)
    
    user = Entry(frame,width=40,fg='black',border=0,bg='white',justify=RIGHT,font=("B Homa",13))
    user.place(x=30,y=220)
    user.insert(0,"نام کاربری")
    user.bind("<FocusIn>",on_enter)
    user.bind("<FocusOut>",on_leave)
    user.bind()
    Frame(frame,width=295,height=2,bg='black').place(x=25,y=247)
    
    #######################
    def on_enter(e):
        if (code.get())=="رمز ورود":
            code.delete(0, 'end')
    def on_leave(e):
    
        if len(code.get())==0:
           Label(frame,text="رمز را وارد کن",fg="red",bg="white",border=0,font=("B Homa",13)).place(x=25, y=327)
           code.insert(0, "رمز ورود")
        elif len(code.get())<8:
            Label(frame,text="باید بیشتر از 8 کارکتر باشد",fg="red",bg="white",border=0,font=("B Homa",13)).place(x=25, y=327)
           
        else:
            Label(text="تایید شد",fg="green",bg="white",border=0,font=("B Homa",13),padx=50).place(x=-22, y=327)
    code = Entry(frame,width=40,fg='black',border=0,bg='white',justify=RIGHT,font=("B Homa",13))
    code.place(x=30,y=290)
    code.insert(0,"رمز ورود")
    code.bind("<FocusIn>",on_enter)
    code.bind("<FocusOut>",on_leave)
    Frame(frame,width=295,height=2,bg='black').place(x=25,y=317)
    ########################
    def on_enter(e):
        if (numberE.get())=="شماره تلفن":
            numberE.delete(0, 'end')
    def on_leave(e):
        if len(numberE.get())==0:
            Label(frame,text="شماره تلفن رو وارد کن",fg="red",bg="white",border=0,font=("B Homa",13)).place(x=25, y=397)
            numberE.insert(0, "شماره تلفن")
        else:
            Label(text="تایید شد",fg="green",bg="white",border=0,font=("B Homa",13),padx=50).place(x=-22, y=397)
    numberE = Entry(frame,width=40,fg='black',border=0,bg='white',justify=RIGHT,font=("B Homa",13))
    numberE.place(x=30,y=360)
    numberE.insert(0,"شماره تلفن")
    numberE.bind("<FocusIn>",on_enter)
    numberE.bind("<FocusOut>",on_leave)
    numberE.bind()
    Frame(frame,width=295,height=2,bg='black').place(x=25,y=387)
    ########################
    def on_enter(e):
        if (ageE.get())=="چند سالته":
            ageE.delete(0, 'end')
    def on_leave(e):
        if len(ageE.get())==0:
            Label(frame,text="کن وارد را سنت",fg="red",bg="white",border=0,font=("B Homa",13)).place(x=25, y=467)
            ageE.insert(0, "چند سالته")
        else:
            Label(text="تایید شد",fg="green",bg="white",border=0,font=("B Homa",13),padx=50).place(x=-22, y=467)
    ageE = Entry(frame,width=40,fg='black',border=0,bg='white',justify=RIGHT,font=("B Homa",13))
    ageE.place(x=30,y=430)
    ageE.insert(0,"چند سالته")
    ageE.bind("<FocusIn>",on_enter)
    ageE.bind("<FocusOut>",on_leave)
    ageE.bind()
    Frame(frame,width=295,height=2,bg='black').place(x=25,y=457)
    ########################
    def on_enter(e):
        if (name_schoolE.get())=="مدرسه نام":
            name_schoolE.delete(0, 'end')
    def on_leave(e):
        if len(name_schoolE.get())==0:
            Label(frame,text="؟کن وارد را مدرسه نام",fg="red",bg="white",border=0,font=("B Homa",13)).place(x=25, y=537)
            name_schoolE.insert(0,"مدرسه نام")
        else:
            Label(text="تایید شد",fg="green",bg="white",border=0,font=("B Homa",13),padx=50).place(x=-22, y=537)
    name_schoolE = Entry(frame,width=40,fg='black',border=0,bg='white',justify=RIGHT,font=("B Homa",13))
    name_schoolE.place(x=30,y=500)
    name_schoolE.insert(0,"مدرسه نام")
    name_schoolE.bind("<FocusIn>",on_enter)
    name_schoolE.bind("<FocusOut>",on_leave)
    name_schoolE.bind()
    Frame(frame,width=295,height=2,bg='black').place(x=25,y=527)
    ########################
    Button(frame,width=39,pady=7,text='ثبت نام کردن',bg='#57a1f8',fg='white',border=0,command=register).place(x=35,y=570)
    label = Label(frame,text="آیا اکانت داری ؟",fg='black',bg='white',font=("B Homa",9))
    label.place(x=200,y=610)
    sing2 = Button(frame,width=6,text='ورود',border=0,bg='white',cursor='hand2',fg='#57a1f8' ,command=log)
    sing2.place(x=155,y=610)
    ########
    label = Label(frame,text="نیاز به کمک داری ؟",fg='black',bg='white',font=("B Homa",9))
    label.place(x=200,y=630)
    sing2 = Button(frame,width=6,text="کمک",border=0,bg='white',cursor='hand2',fg='#57a1f8' ,command=helpme)
    sing2.place(x=155,y=630)
    
    root.mainloop()
#################################################################################
def managehome():
    import requests
    import datetime
    import pytz
    import tkinter as tk
    import time
    import math
    import tkinter
    import tkinter.messagebox
    import customtkinter
    import sqlite3
    conn = sqlite3.connect('user1.db')
    c = conn.cursor()               
    profilename1 = "200"

    
    roote = Tk()
    roote.title("Robosan")
    roote.iconbitmap(bitmap = 'Icons\\icon.ico')
    roote.geometry(f"1390x560")
    fg_cccc = " "
    def dark():
        global fg_cccc
        if switch.get() == "on":
            customtkinter.set_appearance_mode("Dark")
            fg_cccc="dark"
        else:
            customtkinter.set_appearance_mode("Light")
            fg_cccc="light"
    root1 = customtkinter.CTkFrame(roote,fg_color=("#FFFFFF","#2E2E2E"), width=140, corner_radius=0)
    root1.grid(row=0, column=0, rowspan=4, sticky="nsew")
    
    #################"Chat"
    import datetime
        
    conn = sqlite3.connect('user1.db')
    c = conn.cursor()
        
    c.execute("SELECT * FROM user_profile_maneger")
    results = c.fetchall()
    now = datetime.datetime.now()
    now = datetime.datetime.today()
    mm = str(now.month)
    dd = str(now.day)
    yyyy = str(now.year)
    hour = str(now.hour)
    mi = str(now.minute)
    ss = str(now.second)
    x=str(mm + "/" + dd + "/" + yyyy + " " + hour + ":" + mi + ":" + ss )
    if results:
            for i in results:
                profilename = i[0]
                profikelastname = i[1]
                profileemail = i[2]
                profiletime = i[3]
    
    # Create the table for storing contact information
    conn.execute('CREATE TABLE IF NOT EXISTS '+profileemail+profikelastname+profikelastname+' (user text);')
    conn.commit()
    print(("open :"+(str(x))))
    c.execute('INSERT INTO '+profileemail+profikelastname+profikelastname+' VALUES (:user)',{
        "user":("open (برنامه) :"+(str(x))),
        })
    conn.commit()
    # configure grid layout (4x4)
    
    def robo2(tip: str):
        if profilename1 == "200":
         if tip == "ما درباره":
            ("ما درباره شد باز"+(str(x)))
            Introduction()
         elif tip == "دفترچه":
            ("دفترچه شد باز"+(str(x)))
            note()
         elif tip == "حساب ماشین":
            ("حساب ماشین شد باز"+(str(x)))
            cul()
         elif tip == "trans":
            ("open trans :"+(str(x)))
            tans1()
         elif tip == "search":
            ("open search :"+(str(x)))
            search()
         elif tip == "BMI":
            ("شد باز BMI :"+(str(x)))
            BMI()
         elif tip == "گوگل در جستجو":
            ("گوگل در جستجو شد باز"+(str(x)))
            google()
         elif  tip == "هوا و آب":
            ("هوا و آب شد باز"+(str(x)))
            weather()
         elif  tip == "کد مورس":
            ("مورس کد شد باز"+(str(x)))
            mors()
         elif  tip == "پلیر موزیک":
            ("پلیر موزیک شد باز"+(str(x)))
            mus()
         elif  tip  == "شمار ثانیه":
            ("شمار ثانیه شد باز"+(str(x)))
            corno1()
         elif  tip == "دوز بازی":
            ("دوز بازی شد باز"+(str(x)))
            Tic1()
         elif  tip == "متن ادیت":
            ("متن ادیت شد باز"+(str(x)))
            edit_text1()
         elif  tip == "انس گرم پند محاسبه":
            ("انس گرم پند محاسبه شد باز"+(str(x)))
            Grame()
         elif  tip == "تقویم":
            ("تقویم شد باز"+(str(x)))
            calendare1()
         elif  tip == "کد ار یو کی":
            ("کد ار یو کی شد باز"+(str(x)))
            Qr1()
         elif  tip == "تلفن دفترچه":
            ("تلفن دفتر شد باز"+(str(x)))
            phone1()
         elif  tip == "من روزانه برنامه":
            ("من روزانه برنامه شد باز"+(str(x)))
            todo1()
         elif  tip == "بات چت":
            ("بات چت شد باز"+(str(x)))
            chatbot1()
        else:
         if tip == "Introduction":
            ("open Introduction :"+(str(x)))
            Introduction()
         elif tip == "note book":
            ("open note book :"+(str(x)))
            note()
         elif tip == "calculator":
            ("open calculator :"+(str(x)))
            cul()
         elif tip == "trans":
            ("open trans :"+(str(x)))
            tans1()
         elif tip == "search":
            ("open search :"+(str(x)))
            search()
         elif tip == "BMI":
            ("open BMI :"+(str(x)))
            BMI()
         elif tip == "search google":
            ("open search google :"+(str(x)))
            google()
         elif  tip == "weather":
            ("open weather :"+(str(x)))
            weather()
         elif  tip == "mourse code":
            ("open mourse code :"+(str(x)))
            mors()
         elif  tip == "music player":
            ("open music player :"+(str(x)))
            mus()
         elif  tip  == "Countdown":
            ("open Countdown :"+(str(x)))
            corno1()
         elif  tip == "Tic Tac":
            ("open Tic Tac :"+(str(x)))
            Tic1()
         elif  tip == "Edit Text":
            ("open Edit Text :"+(str(x)))
            edit_text1()
         elif  tip == "Gram Pound Ounce":
            ("open Gram Pound Ounce :"+(str(x)))
            Grame()
         elif  tip == "calendar":
            ("open calendar :"+(str(x)))
            calendare1()
         elif  tip == "QrCode":
            ("open QrCode :"+(str(x)))
            Qr1()
         elif  tip == "Phone Book":
            ("open Phone Book :"+(str(x)))
            phone1()
         elif  tip == "todo":
            ("open todo :"+(str(x)))
            todo1()
         elif  tip == "Chat Bot":
            ("open Chat Bot :"+(str(x)))
    def menu():
        pass
    def tans1():
        c.execute('INSERT INTO '+profileemail+profikelastname+profikelastname+' VALUES (:user)',{
        "user":("open trans :"+(str(x))),
        })
        conn.commit()
        trans1()
    def openfile():
        pass
    def savefile():
        pass
    def Introduction():
        c.execute('INSERT INTO '+profileemail+profikelastname+profikelastname+' VALUES (:user)',{
        "user":("open (ما درباره) :"+(str(x))),
        })
        conn.commit()    
        Introduction1()
    def pack():
        pass
    
    def note():
        c.execute('INSERT INTO '+profileemail+profikelastname+profikelastname+' VALUES (:user)',{
        "user":("open (دفترچه) :"+(str(x))),
        })
        conn.commit()
        note1()
        
    def corno1():
        c.execute('INSERT INTO '+profileemail+profikelastname+profikelastname+' VALUES (:user)',{
        "user":("open (تایمر) :"+(str(x))),
        })
        conn.commit()
        roote.destroy()
        corno()
        robo()
        
    def profile():
        c.execute('INSERT INTO '+profileemail+profikelastname+profikelastname+' VALUES (:user)',{
        "user":("open Profile :"+(str(x))),
        })
        conn.commit()
        Profile1()
    def cul():
        c.execute('INSERT INTO '+profileemail+profikelastname+profikelastname+' VALUES (:user)',{
        "user":("open Calculator :"+(str(x))),
        })
        conn.commit()
        
        cul1()
    def search():
        c.execute('INSERT INTO '+profileemail+profikelastname+profikelastname+' VALUES (:user)',{
        "user":("open Search :"+(str(x))),
        })
        conn.commit()
        
        search1()
        os.system("python rock.py")
    def BMI():
        c.execute('INSERT INTO '+profileemail+profikelastname+profikelastname+' VALUES (:user)',{
        "user":("open BMI :"+(str(x))),
        })
        conn.commit()
        roote.destroy()
        BMI1()
         
    def Grame():
        c.execute('INSERT INTO '+profileemail+profikelastname+profikelastname+' VALUES (:user)',{
        "user":("open (جرمی محاسبگر) :"+(str(x))),
        })
        conn.commit()
        Gram()
        
    def Qr1():
        c.execute('INSERT INTO '+profileemail+profikelastname+profikelastname+' VALUES (:user)',{
        "user":("open QrCode :"+(str(x))),
        })
        conn.commit()
        os.system("python Qr.py")
        
    def weather():
        c.execute('INSERT INTO '+profileemail+profikelastname+profikelastname+' VALUES (:user)',{
        "user":("open Weather :"+(str(x))),
        })
        conn.commit()
        
        weather1()
    def mors():
        c.execute('INSERT INTO '+profileemail+profikelastname+profikelastname+' VALUES (:user)',{
        "user":("open (کد مورس) :"+(str(x))),
        })
        conn.commit()
        
        mors1()
    def mus():
        c.execute('INSERT INTO '+profileemail+profikelastname+profikelastname+' VALUES (:user)',{
        "user":("open (پلیر موزیک) :"+(str(x))),
        })
        conn.commit()
                           
        mus1()
    def backe():
        conn.commit()
        roote.destroy()
        load()
    def google():
        c.execute('INSERT INTO '+profileemail+profikelastname+profikelastname+' VALUES (:user)',{
        "user":("open google :"+(str(x))),
        })
        conn.commit()
        os.system("python google.py")
    def help1():
        c.execute('INSERT INTO '+profileemail+profikelastname+profikelastname+' VALUES (:user)',{
        "user":("open Help :"+(str(x))),
        })
        conn.commit()
        help()
    def about():
        
        os.system("python about.py")   
    def cmd():
        
        os.system("python robosan.py")    
    def timeis():
        pass
    def langu():
        try:
           c.execute('INSERT INTO '+profileemail+profikelastname+profikelastname+' VALUES (:user)',{
            "user":("open Edit languge :"+(str(x))),
            })
           conn.commit()   
           GoogleTranslator(target="fa").translate("Please register or Login to your account")
           roote.destroy()
            
        except:
            messagebox.showerror("Error","No Network")
    def Tic1():
        c.execute('INSERT INTO '+profileemail+profikelastname+profikelastname+' VALUES (:user)',{
        "user":("open (دوز بازی) :"+(str(x))),
        })
        conn.commit()
        Tic()
    def edit_text1():
        c.execute('INSERT INTO '+profileemail+profikelastname+profikelastname+' VALUES (:user)',{
        "user":("open (متن ادیتور) :"+(str(x))),
        })
        conn.commit()
        edit_text()
    def calendare1():
        c.execute('INSERT INTO '+profileemail+profikelastname+profikelastname+' VALUES (:user)',{
        "user":("open Calendare :"+(str(x))),
        })
        conn.commit()
        calendare1()
    def phone1():
        c.execute('INSERT INTO '+profileemail+profikelastname+profikelastname+' VALUES (:user)',{
        "user":("open (تلفن دفتر) :"+(str(x))),
        })
        conn.commit()
        phone()
    def todo1():
        c.execute('INSERT INTO '+profileemail+profikelastname+profikelastname+' VALUES (:user)',{
        "user":("open (کارهام دفتر) :"+(str(x))),
        })
        conn.commit()
        todo()
    def chatbot1():
        c.execute('INSERT INTO '+profileemail+profikelastname+profikelastname+' VALUES (:user)',{
        "user":("open Chat Bot :"+(str(x))),
        })
        conn.commit()
        chatbot()
    def map():
        c.execute('INSERT INTO '+profileemail+profikelastname+profikelastname+' VALUES (:user)',{
        "user":("open World Map :"+(str(x))),
        })
        conn.commit()
        os.system("python map.py")
        
    def open_input_dialog_event():
        try:
            from deep_translator import GoogleTranslator
            textbox.delete('0.0',END)
            dialog = combobox_1.get()
            
             
            def get_country_name(language_name):
                    from googletrans import LANGUAGES
                    for code, name in LANGUAGES.items():
                        if name == language_name:
                            return code
            language_name = dialog
            country_name = get_country_name(language_name)
            output =GoogleTranslator(target=country_name).translate(entry1.get())
            textbox.insert("0.0", "Transletor:\n" +str(output)+"\n")
        except:
            messagebox.showerror("Warning", "not internet")  
    
            
    def open_input_dialog_event1():   
       try:

        import openai
        openai.api_key = "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"
        # set up the GPT model
        model_engine = "text-davinci-003"
        prompt = "Hello! I'm a chatbot. How can I help you?"
        textbox.delete('0.0',END)
        if combobox_12.get() == "Wolfram":
            import wolframalpha
            client=wolframalpha.Client('26AW5W-23YY3TTWHG')
            res = client.query(entry2.get())
            answer = next(res.results).text
            textbox.insert("0.0", "Wolfram:\n" +str(answer))
        elif combobox_12.get() == "Wikipedia":
            import wikipedia
            Iran = wikipedia.summary(entry2.get())
            textbox.insert("0.0", "Wikipedia:\n" +str(Iran))
       except:
           messagebox.showerror("we have problem","We have a problem pls ckek leter") 
    
    def change_appearance_mode_event(new_appearance_mode: str):
        customtkinter.set_appearance_mode(new_appearance_mode)
    def change_scaling_event(new_scaling: str):
        new_scaling_float = int(new_scaling.replace("%", "")) / 100
        customtkinter.set_widget_scaling(new_scaling_float)
    def sidebar_button_event():
        backe()
    
    def Reset():
        roote.destroy()
        robo()
    def history():
        textbox.delete('0.0',END)
        import sqlite3
        conn = sqlite3.connect('user1.db')
        c = conn.cursor()
            
        c.execute("SELECT * FROM user_profile_maneger")
        results = c.fetchall()
        textbox.delete('0.0',END)
        textbox.insert("0.0", "History:\n")
        if results:
                for i in results:
                    profilename = i[0]
                    profikelastname = i[1]
                    profileemail = i[2]
                    profiletime = i[3]
        for row in c.execute('SELECT * FROM '+profileemail+profikelastname+profikelastname+''):
            textbox.configure(state='normal')
            textbox.insert('end', "User: "+row[0] + "\n")
            textbox.configure(state='disabled')
            
    #creat bg
    home_frame = customtkinter.CTkFrame(root1, corner_radius=0, fg_color="transparent")
    home_frame.grid(row=0, column=2,sticky="nsew")

    second_frame = customtkinter.CTkFrame(root1, corner_radius=0, width=1070,height=560, fg_color="transparent")
    third_frame = customtkinter.CTkFrame(root1, corner_radius=0, width=1250,height=560, fg_color="transparent")

    root = home_frame

    def select_frame_by_name(name):
        # set button color for selected button
        home_button.configure(fg_color=("gray75", "gray25") if name == "home" else "transparent")
        frame_2_button.configure(fg_color=("gray75", "gray25") if name == "chat" else "transparent")
        teacher_button.configure(fg_color=("gray75", "gray25") if name == "teacher" else "transparent")

        # show selected frame
        if name == "home":
            home_frame.grid(row=0, column=2, sticky="nsew")
        else:
            home_frame.grid_forget()
        if name == "chat":
            second_frame.grid(row=0, column=2, sticky="nsew")
        else:
            second_frame.grid_forget()
        if name == "teacher":
            third_frame.grid(row=0, column=1, sticky="nsew")
        else:
            third_frame.grid_forget()
    def home_button_event():
        select_frame_by_name("home")

    def chat_button_event():
        select_frame_by_name("chat")
    
    def teacher_button_event():
        select_frame_by_name("teacher")


    navigation_frame = customtkinter.CTkFrame(root1,width=140, corner_radius=0)
    navigation_frame.grid(row=0, column=0,rowspan=4, sticky="nsew")
    navigation_frame.grid_rowconfigure(4, weight=1)
    from PIL import Image
    home_image = customtkinter.CTkImage(light_image=Image.open(os.path.join("test_images\\home_dark.png")),
                                                 dark_image=Image.open(os.path.join("test_images\\home_light.png")), size=(20, 20))
    chat_image = customtkinter.CTkImage(light_image=Image.open(os.path.join("test_images\\chat_dark.png")),
                                                 dark_image=Image.open(os.path.join("test_images\\chat_light.png")), size=(20, 20))
    add_user_image = customtkinter.CTkImage(light_image=Image.open(os.path.join("test_images\\add_user_dark.png")),
                                                     dark_image=Image.open(os.path.join("test_images\\add_user_light.png")), size=(20, 20))
    
    home_button = customtkinter.CTkButton(navigation_frame, corner_radius=0, height=40, border_spacing=10, text="Home",
                                                   fg_color="transparent", text_color=("gray10", "gray90"), hover_color=("gray70", "gray30"),
                                                   image=home_image, anchor="w", command=home_button_event)
    home_button.grid(row=1, column=0, sticky="ew")

    frame_2_button = customtkinter.CTkButton(navigation_frame, corner_radius=0, height=40, border_spacing=10, text="Chat",
                                                      fg_color="transparent", text_color=("gray10", "gray90"), hover_color=("gray70", "gray30"),
                                                      image=chat_image, anchor="w", command=chat_button_event)
    frame_2_button.grid(row=2, column=0, sticky="ew")
    teacher_button = customtkinter.CTkButton(navigation_frame, corner_radius=0, height=40, border_spacing=10, text="List Teacher",
                                                      fg_color="transparent", text_color=("gray10", "gray90"), hover_color=("gray70", "gray30"),
                                                      image=add_user_image, anchor="w", command=teacher_button_event)
    teacher_button.grid(row=3, column=0, sticky="ew")
    ######
    
    width = 1099
    height = 600
    # create third fram
    
    
    
   

    def show():
        roote.destroy()
        showsleep()
        roote()
    # create sidebar frame with widgets
    




    ######################################
    sidebar_frame = customtkinter.CTkFrame(root1, width=140,fg_color=("#FFFFFF","#2E2E2E"), corner_radius=0)
    root1.grid(row=0, column=0, rowspan=4, sticky="nsew")
    sidebar_frame.grid(row=0, column=1, rowspan=4, sticky="nsew")
    sidebar_frame.grid_rowconfigure(7, weight=1)
    if profilename1 == "200":
        logo_label = customtkinter.CTkLabel(sidebar_frame, text="روبوسان", font=customtkinter.CTkFont(size=20, weight="bold",family="B Roya"))
    else:
        logo_label = customtkinter.CTkLabel(sidebar_frame, text="RoBoSaN", font=customtkinter.CTkFont(size=20, weight="bold"))
    logo_label.grid(row=0, column=0, padx=20, pady=(20, 10))
    if profilename1 == "200":
        tipe = customtkinter.CTkOptionMenu(sidebar_frame, values=["ما درباره",  "BMI", "کد مورس", "پلیر موزیک", "شمار ثانیه", "دوز بازی", "متن ادیت", "انس گرم پند محاسبه", "کد ار یو کی", "تلفن دفترچه", "من روزانه برنامه"],font=customtkinter.CTkFont("B Homa",15),
    
                                                           command=robo2) 
    tipe.grid(row=3, column=0, padx=20, pady=(10))
    if profilename1 == "200":
        tipe.set("منو")    
    else:
        tipe.set("Menu")
    #################
    def news():
     try:
        import requests
        from bs4 import BeautifulSoup
        from PIL import Image
        from io import BytesIO
        import os

        #    ایجاد پوشه‌ی "news" اگر وجود ندارد
        if not os.path.exists('news'):
            os.makedirs('news')

        # آدرس وب‌سایتی که می‌خواهید از آن تصاویر را دانلود کنید
        url = 'https://zamandaily.ir/'

        # دریافت محتوای صفحه وب
        response = requests.get(url)
        soup = BeautifulSoup(response.text, 'html.parser')

        # پیدا کردن تمام تگ‌های img در صفحه وب
        images = soup.find_all('img')

        # تعیین محدوده‌ی دانلود تصاویر
        start_index = 5 # شروع دانلود از تصویر پنجم
        end_index = 10 # پایان دانلود در تصویر دهم

        # دانلود و ذخیره‌ی تصاویر در پوشه‌ی "news"
        for index, img in enumerate(images[start_index:end_index], start=start_index):
            img_url = img['src']
            img_data = requests.get(img_url).content
            img_name = f'news/image_{index}.jpg'
    
            with open(img_name, 'wb') as file:
                file.write(img_data)
                print(f'تصویر {img_name} دانلود و در پوشه‌ی "news" ذخیره شد.')

            # نمایش تصویر
            image = Image.open(BytesIO(img_data))
            image.show()
        messagebox.showinfo("stop","صفحات روزنامه تمام شد")
        print('دانلود و نمایش تصاویر به پایان رسید.')
     except:
        messagebox.showerror("not internet","اینترنت قطع است لطف اینترنت رو چک کنید")
         
    def pdf():

        image = Image.open("pdf\\Capture.png")
        image.show()
        time.sleep(5)
        image = Image.open("pdf\\Capture1.png")
        image.show()
        time.sleep(5)
        image = Image.open("pdf\\Capture2.png")
        image.show()
        time.sleep(5)
        image = Image.open("pdf\\Capture3.png")
        image.show()
        time.sleep(5)
        image = Image.open("pdf\\Capture4.png")
        image.show()
        time.sleep(5)


    
    
    sidebar_button_5 = customtkinter.CTkButton(sidebar_frame,text="اخبار فوری",font=customtkinter.CTkFont("B Homa",15),command=news)
    sidebar_button_5.grid(row=4, column=0, padx=20, pady=10)

    sidebar_button_5 = customtkinter.CTkButton(sidebar_frame,text="توضیحات",font=customtkinter.CTkFont("B Homa",15),command=pdf)
    sidebar_button_5.grid(row=5, column=0, padx=20, pady=10)
    if profilename1 == "200":  
        sidebar_button_6 = customtkinter.CTkButton(sidebar_frame,text="خروج از حساب کاربری",font=customtkinter.CTkFont("B Homa",15), command=backe)
        sidebar_button_6.grid(row=6, column=0, padx=10, pady=0)
    else:
        sidebar_button_3 = customtkinter.CTkButton(sidebar_frame,text="Log Out", command=backe)
        sidebar_button_3.grid(row=3, column=0, padx=10, pady=0)
        sidebar_button_10 = customtkinter.CTkButton(sidebar_frame,text="Show Sleep", command=show)
        sidebar_button_10.grid(row=4, column=0, padx=10, pady=0)
    
    
    
    if profilename1 == "200":
        sidebar_button_3 = customtkinter.CTkButton(sidebar_frame,text="خروج",font=customtkinter.CTkFont("B Homa",15), command=backe)
    else:
        sidebar_button_3 = customtkinter.CTkButton(sidebar_frame,text="Log Out", command=backe)
    if profilename1 == "200":
        switch = customtkinter.CTkSwitch(master=sidebar_frame, text="تاریک",command=dark,font=customtkinter.CTkFont("B Homa",15), onvalue="on", offvalue="off")
    else:
        switch = customtkinter.CTkSwitch(master=sidebar_frame, text="Sleep",command=dark, onvalue="on", offvalue="off")
    switch.grid(row=7, column=0, padx=1, pady=1)
    if profilename1 == "200":
        scaling_label = customtkinter.CTkLabel(sidebar_frame, text="اندازه متن :",font=customtkinter.CTkFont("B Homa",15), anchor="w")
    else:
        scaling_label = customtkinter.CTkLabel(sidebar_frame, text="UI Scaling:", anchor="w")
    scaling_label.grid(row=8, column=0, padx=20, pady=(1))
    scaling_optionemenu = customtkinter.CTkOptionMenu(sidebar_frame, values=["80%", "90%", "100%", "110%", "120%"],
                                                           command=change_scaling_event)
    scaling_optionemenu.grid(row=8, column=0, padx=20, pady=(10,20))
    # create main entry and button
    from tkintermapview import TkinterMapView
    # create textbox
    textbox = customtkinter.CTkTextbox(master=root,width=250)
    textbox.grid(row=0, column=2, padx=(20, 0), pady=(20, 0), sticky="nsew")

    
    
    def search_event(event=None):
            map_widget.set_address(entry32.get())
    mape = customtkinter.CTkFrame(root,corner_radius=0)
    mape.grid(row=1, column=2, padx=(20, 0), pady=(20, 0), sticky="nsew")
    map_widget = TkinterMapView(mape, corner_radius=0,width=353,height=220)
    map_widget.grid(row=1, column=0, sticky="nswe", padx=20, pady=(10, 10))
    if profilename1 == "200":
        entry32 = customtkinter.CTkEntry(mape, placeholder_text="..... نام مکان مورد نظرت را وارد کن",justify=RIGHT,font=customtkinter.CTkFont("B Homa",13))
    else:
        entry32 = customtkinter.CTkEntry(mape, placeholder_text="Enter Contry or ...")
    entry32.grid(row=0, column=0,  padx=1, pady=1, sticky="nsew")
    entry32.bind("<Return>", search_event)
    
    
    
    # create tabview
    tabview = customtkinter.CTkTabview(root, width=320)
    tabview.grid(row=0, column=3, padx=(20, 0), pady=(20, 0), sticky="nsew")
    if profilename1 == "200":
        tabview.add("مدیر امکانات")
        tabview.tab("مدیر امکانات").grid_columnconfigure(0, weight=1)
        tabview.add("ترجمگر")
        tabview.add("هوا و آب")
        tabview.add("بات چت")
        tabview.add("حساب ماشین")
        tabview.tab("ترجمگر").grid_columnconfigure(0, weight=1) # configure grid of individual tabs
        tabview.tab("هوا و آب").grid_columnconfigure(0, weight=1)
        tabview.tab("بات چت").grid_columnconfigure(0, weight=1)
        tabview.tab("حساب ماشین").grid_columnconfigure(0, weight=1)
    else:
        tabview.add("Transletor")
        tabview.add("Weather")
        tabview.add("ChatBot")
        tabview.add("calculet")
        tabview.tab("Transletor").grid_columnconfigure(0, weight=1) # configure grid of individual tabs
        tabview.tab("Weather").grid_columnconfigure(0, weight=1)
        tabview.tab("ChatBot").grid_columnconfigure(0, weight=1)
        tabview.tab("calculet").grid_columnconfigure(0, weight=1)
    def addteacher():
        roote.destroy()
        addteacher1()
        managehome()
    def addteacher1():
        def register():
            import datetime
            name = nameE.get()
            last_name = last_nameE.get()
            email = user.get()
            password = code.get()
            number= numberE.get()
            age =ageE.get()
            now = datetime.datetime.today()
            mm = str(now.month)
            dd = str(now.day)
            yyyy = str(now.year)
            hour = str(now.hour)
            mi = str(now.minute)
            ss = str(now.second)
            x=str(mm + "/" + dd + "/" + yyyy + " " + hour + ":" + mi + ":" + ss)
            conn = sqlite3.connect("user1.db")
            c = conn.cursor()
    
            c.execute("SELECT * FROM user_profile_maneger")
            results = c.fetchall()
        
            if results:
                for i in results:
                    profile_name = i[0]
                    profile_last_name = i[1]

            # نام پوشه‌ای که می‌خواهید ایجاد کنید
            
                
                
            while True:
                        
                        conn = sqlite3.connect("user1.db")
                        c = conn.cursor()
                        c.execute("INSERT INTO user_info_techer VALUES (:User4name, :User4lastname, :User4email, :User4password, :User4time, :User4number, :User4born, :User4name_maneger, :User4last_name_maneger)",
                            {
                                "User4name": name,
                                "User4lastname": last_name,
                                "User4email": email,
                                "User4password": password,
                                "User4time" : x,
                                "User4number":number,
                                "User4born":age,
                                "User4name_maneger":profile_name,
                                "User4last_name_maneger":profile_last_name,
                        
                            })
            
                        conn.commit()
                        desktop_path = os.path.join(os.path.join(os.environ['USERPROFILE']), "Desktop")
                        folder_name = os.path.join(desktop_path,"معلم_"+name+"_"+last_name)

                        # ایجاد پوشه در مسیر فعلی
                        if not os.path.exists(folder_name):
                            os.makedirs(folder_name)
                            print(f'پوشه‌ای با نام "{folder_name}" ایجاد شد.')
                        else:
                            pass
                        root.destroy()
                        messagebox.showinfo("sing in","آقا/خانم "+name+" "+last_name+" ثبت شد")
                        
                        
    
        root =Tk()
        root.title("ثبت نام")
        root.geometry('350x670')
        root.iconbitmap(bitmap = 'Icons\\icon.ico')
        root.configure(bg="#fff")
        root.resizable(False,False)
    
        frame=root
        heading=Label(frame,text="معلم کردن نام ثبت",fg="#57a1f8",bg="white",border=0,font=("B Homa",23))
        heading.place(x=120,y=30)
        ###############3
        def on_enter(e):
            if nameE.get() == "نام":
                nameE.delete(0, 'end')
        def on_leave(e):
            if len(nameE.get())==0:
                Label(frame,text="نام را وارد کن",fg="red",bg="white",border=0,font=("B Homa",13)).place(x=25, y=110)
                nameE.insert(0, 'نام')
            else:
                Label(frame,text="تایید شد",justify=RIGHT,width=10,fg="green",bg="white",border=0,font=("B Homa",13),padx=50).place(x=-22, y=110)
        nameE = Entry(frame,width=40,fg='black',border=0,bg='white',justify=RIGHT,font=("B Homa",13))
        nameE.place(x=30,y=80)
        nameE.insert(0,'نام')
        nameE.bind("<FocusIn>",on_enter)
        nameE.bind("<FocusOut>",on_leave)
        nameE.bind()
        Frame(frame,width=295,height=2,bg='black').place(x=25,y=107)
        ###################
        def on_enter(e):
            if last_nameE.get()=="نام خانوادگی":
                last_nameE.delete(0, 'end')
        def on_leave(e):
            if len(last_nameE.get())==0:
                Label(frame,text="نام خانوادگی را وارد کن",fg="red",bg="white",border=0,font=("B Homa",13)).place(x=25, y=180)
                last_nameE.insert(0, "نام خانوادگی")
            else:
                Label(text="تایید شد",width=10,fg="green",bg="white",border=0,font=("B Homa",13),padx=50).place(x=-22, y=180)
        last_nameE = Entry(frame,width=40,fg='black',border=0,bg='white',justify=RIGHT,font=("B Homa",13))
        last_nameE.place(x=30,y=150)
        last_nameE.insert(0,"نام خانوادگی")
        last_nameE.bind("<FocusIn>",on_enter)
        last_nameE.bind("<FocusOut>",on_leave)
        last_nameE.bind()
        Frame(frame,width=295,height=2,bg='black').place(x=25,y=177)
        #######################
        def on_enter(e):
            if (user.get())=="نام کاربری":
                user.delete(0, 'end')
        def on_leave(e):
            if len(user.get())==0:
                Label(frame,text="نام کاربری را وارد کن",fg="red",bg="white",border=0,font=("B Homa",13)).place(x=25, y=257)
                user.insert(0, "نام کاربری")
            else: 
                Label(frame,text="تایید شد",justify=RIGHT,width=10,fg="green",bg="white",border=0,font=("B Homa",13),padx=50).place(x=-22, y=257)
    
        user = Entry(frame,width=40,fg='black',border=0,bg='white',justify=RIGHT,font=("B Homa",13))
        user.place(x=30,y=220)
        user.insert(0,"نام کاربری")
        user.bind("<FocusIn>",on_enter)
        user.bind("<FocusOut>",on_leave)
        user.bind()
        Frame(frame,width=295,height=2,bg='black').place(x=25,y=247)
    
        #######################
        def on_enter(e):
            if (code.get())=="رمز ورود":
                code.delete(0, 'end')
        def on_leave(e):
    
            if len(code.get())==0:
                Label(frame,text="رمز را وارد کن",fg="red",bg="white",border=0,font=("B Homa",13)).place(x=25, y=327)
                code.insert(0, "رمز ورود")
            elif len(code.get())<8:
                Label(frame,text="رمز باید بیشتر از 8 کارکتر باشد",fg="red",bg="white",border=0,font=("B Homa",13)).place(x=25, y=327)
           
            else:
                Label(text="تایید شد",fg="green",width=10,bg="white",border=0,font=("B Homa",13),padx=50).place(x=-22, y=327)
        code = Entry(frame,width=40,fg='black',border=0,bg='white',justify=RIGHT,font=("B Homa",13))
        code.place(x=30,y=290)
        code.insert(0,"رمز ورود")
        code.bind("<FocusIn>",on_enter)
        code.bind("<FocusOut>",on_leave)
        Frame(frame,width=295,height=2,bg='black').place(x=25,y=317)
        ########################
        def on_enter(e):
            if (numberE.get())=="شماره تلفن":
                numberE.delete(0, 'end')
        def on_leave(e):
            if len(numberE.get())==0:
                Label(frame,text="شماره تلفن رو وارد کن",fg="red",bg="white",border=0,font=("B Homa",13)).place(x=25, y=397)
                numberE.insert(0, "شماره تلفن")
            else:
                Label(text="تایید شد",fg="green",width=10,bg="white",border=0,font=("B Homa",13),padx=50).place(x=-22, y=397)
        
        numberE = Entry(frame,width=40,fg='black',border=0,bg='white',justify=RIGHT,font=("B Homa",13))

        numberE.place(x=30,y=360)
        numberE.insert(0,"شماره تلفن")
        numberE.bind("<FocusIn>",on_enter)
        numberE.bind("<FocusOut>",on_leave)
        numberE.bind()
        Frame(frame,width=295,height=2,bg='black').place(x=25,y=387)
        ########################
        def on_enter(e):
            if (ageE.get())=="درس آموزشی":
                ageE.delete(0, 'end')
        def on_leave(e):
            if len(ageE.get())==0:
                Label(frame,text="درس آموزشی معلم را وارد کن",fg="red",bg="white",border=0,font=("B Homa",13)).place(x=25, y=467)
                ageE.insert(0, "درس آموزشی")
            else:
                Label(text="تایید شد",width=10,fg="green",bg="white",border=0,font=("B Homa",13),padx=50).place(x=-22, y=467)
        ageE = Entry(frame,width=40,fg='black',border=0,bg='white',justify=RIGHT,font=("B Homa",13))
        ageE.place(x=30,y=430)
        ageE.insert(0,"درس آموزشی")
        ageE.bind("<FocusIn>",on_enter)
        ageE.bind("<FocusOut>",on_leave)
        ageE.bind()
        Frame(frame,width=295,height=2,bg='black').place(x=25,y=457)
        ########################

        ########################
        Button(frame,width=39,pady=7,text="ثپت نام معلم جدید",bg='#57a1f8',fg='white',border=0,command=register).place(x=35,y=570)
    
        ########
    
    
        root.mainloop()
    add_user_image = customtkinter.CTkImage(light_image=Image.open(os.path.join("test_images\\add_user_light.png")), size=(20, 20))
    string_input = customtkinter.CTkButton(tabview.tab("مدیر امکانات"), text="اضافه کردن معلم",image=add_user_image,font=customtkinter.CTkFont("B Homa",13), command=addteacher)
    string_input.grid(row=1, column=0, padx=20, pady=(10, 10))
    string_input1 = customtkinter.CTkButton(tabview.tab("مدیر امکانات"), text="اضافه کردن دانش آموز",image=add_user_image,font=customtkinter.CTkFont("B Homa",13), command=open_input_dialog_event)
    string_input1.grid(row=2, column=0, padx=20, pady=(10, 10))
    string_input2 = customtkinter.CTkButton(tabview.tab("مدیر امکانات"), text="لیست معلمان",font=customtkinter.CTkFont("B Homa",13), command=open_input_dialog_event)
    string_input2.grid(row=3, column=0, padx=20, pady=(10, 10))
    string_input3 = customtkinter.CTkButton(tabview.tab("مدیر امکانات"), text="لیست دانش آموزان",font=customtkinter.CTkFont("B Homa",13), command=open_input_dialog_event)
    string_input3.grid(row=4, column=0, padx=20, pady=(10, 10))
    try:
     try:
        api_key="1ceaad7d14b2cb4f9403429d6e09a147"
        base_url="https://api.openweathermap.org/data/2.5/weather?q="
        url=base_url + "Tehran" + "&appid=" + api_key
        #o=(url)
    
        result = requests.get(url)
        weather = result.json()
     except requests.exceptions.ConnectionError:
        messagebox.showwarning('خطا',"در فرایند وصل شدن دچار مشکل شدیم از وصل بودن اینترنت خود مطمعن شوید و اگر آفلاین باشیم نمیتوانید از برنامه های آنلاین استفاده کنید")
     except:
        messagebox.showerror('Error',"Some Errored Occured\nTry again Later!")
    # print(weather)
     if weather['cod']=='404' and weather['message']=='city not found':
        messagebox.showerror("Error","Entered City Not Found")
     elif weather['cod']=='400' and weather['message']=='Nothing to geocode':
        messagebox.showinfo("Warning",'Enter The city name')
     else:
        # getting time according to timezone
        lon=weather['coord']['lon']  # longitutde
        lat=weather['coord']['lat']  # latitude
        type=weather['weather'][0]['main']
        
        # sets the temperature and degree label
        temp=int(weather['main']['temp']-273)
        if type=="Clear":
            img="clear.png"
        elif type=="Clouds":
            img='clouds.png'
        elif type=="Rain":
            img='rain.png'
        elif type=='Haze':
            img='haze.png'
        else:
            img='main.png'
     if profilename1 == "200":
      from PIL import Image,ImageTk
      feel=customtkinter.CTkLabel(tabview.tab("هوا و آب"),text=f"Feels Like {int(weather['main']['feels_like']-273)}° | {weather['weather'][0]['main']}", anchor="w")
      feel.place(x=100,y=55)
      temperature=customtkinter.CTkLabel(tabview.tab("هوا و آب"),text=int(weather['main']['temp']-273), anchor="w",font=customtkinter.CTkFont(size=23))
      temperature.place(x=94,y=25)
      degree=customtkinter.CTkLabel(tabview.tab("هوا و آب"),text="°C",font=customtkinter.CTkFont(size=15), anchor="w")
      degree.place(x=120,y=22)
      location=customtkinter.CTkLabel(tabview.tab("هوا و آب"),text=weather['name'], anchor="w")
      location.place(x=100,y=75)
      img3=Image.open(f"Icons/{img}")
      resizeimg3=img3.resize((80,80))
      finalimg3=ImageTk.PhotoImage(resizeimg3)
      icons = customtkinter.CTkLabel(tabview.tab("هوا و آب"),image=finalimg3,text="", anchor="w")
      icons.place(x=10,y=20)
     else:
      
      feel=customtkinter.CTkLabel(tabview.tab("Weather"),text=f"Feels Like {int(weather['main']['feels_like']-273)}° | {weather['weather'][0]['main']}", anchor="w")
      feel.place(x=100,y=55)
      temperature=customtkinter.CTkLabel(tabview.tab("Weather"),text=int(weather['main']['temp']-273), anchor="w",font=customtkinter.CTkFont(size=23))
      temperature.place(x=94,y=25)
      degree=customtkinter.CTkLabel(tabview.tab("Weather"),text="°C",font=customtkinter.CTkFont(size=15), anchor="w")
      degree.place(x=120,y=22)
      location=customtkinter.CTkLabel(tabview.tab("Weather"),text=weather['name'], anchor="w")
      location.place(x=100,y=75)
      img3=Image.open(f"Icons/{img}")
      resizeimg3=img3.resize((80,80))
      finalimg3=ImageTk.PhotoImage(resizeimg3)
      icons = customtkinter.CTkLabel(tabview.tab("Weather"),image=finalimg3,text="", anchor="w")
      icons.place(x=10,y=20)
    except:
        print("problem")
   

    if profilename1 == "200":
        entry1 = customtkinter.CTkEntry(tabview.tab("ترجمگر"), placeholder_text=".... متن",justify=RIGHT,font=customtkinter.CTkFont("B Homa",13))
        entry1.grid(row=0, column=0,  padx=20, pady=(1, 1), sticky="nsew")
    else:
        entry1 = customtkinter.CTkEntry(tabview.tab("Transletor"), placeholder_text="Text ...")
        entry1.grid(row=0, column=0,  padx=20, pady=(1, 1), sticky="nsew")
    entry1.bind("<Return>", open_input_dialog_event)
    from deep_translator import GoogleTranslator
    from googletrans import LANGUAGES
    values = list(LANGUAGES.values())
    if profilename1 == "200":
        combobox_1 = customtkinter.CTkComboBox(tabview.tab("ترجمگر"), values=values)
    else:
        combobox_1 = customtkinter.CTkComboBox(tabview.tab("Transletor"), values=values)
    combobox_1.grid(row=1, column=0, padx=20, pady=(10, 10))
    search_keyword = tk.StringVar()
    def createButton():
                            conn = sqlite3.connect('user1.db')
                            c = conn.cursor()
                            c.execute('INSERT INTO '+profileemail+profikelastname+profikelastname+' VALUES (:user)',{
                                      "user":(" open calculetor :"+(str(x))),
                                      })
                    
                            '''
                            DOCSTRING: Method that creates the buttons
                            INPUT: nothing
                            OUTPUT: creates a button
                            '''
                            
                            #We first create each button one by one with the value we want
                            #Using addButton() method which is described below
                            b0 = addButton(0)
                            b1 = addButton(1)
                            b2 = addButton(2)
                            b3 = addButton(3)
                            b4 = addButton(4)
                            b5 = addButton(5)
                            b6 = addButton(6)
                            b7 = addButton(7)
                            b8 = addButton(8)
                            b9 =  addButton(9)
                            b_add = addButton('+')
                            b_sub = addButton('-')
                            b_mult = addButton("×")
                            b_div = addButton('÷')
                            b_clear = addButton('c')
                            b_equal = addButton('=')
                            b_tavan = addButton("^")
                            b_baghi = addButton("mod")
                            b_edame = addButton("div")
                            b_sqrt = addButton("r")
                            import math
                            
    
                            
                            
                    
                            #Arrange the buttons into lists which represent calculator rows
                            row1=[b7,b8,b9,b_add]
                            row2=[b4,b5,b6,b_sub]
                            row3=[b1,b2,b3,b_mult]
                            row4=[b_clear,b0,b_equal,b_div]
                            row5=[b_tavan,b_baghi,b_edame,b_sqrt]
                    
                            #Assign each button to a particular location on the GUI
                            r=1
                            for row in [row1, row2, row3, row4, row5]:
                                c=0
                                for buttn in row:
                                    buttn.grid(row=r, column=c, columnspan=1)
                                    c+=1
                                r+=1
                    
                    
                    
                    
    def addButton(value):
                    
                                '''
                                DOCSTRING: Method to process the creation of a button and make it clickable
                                INPUT: value of the button (1,2,3,4,5,6,7,8,9,0,+,-,*,/,c,=,**)
                                OUTPUT: returns a designed button object
                                '''
                                return Button(master, text=value, width=9, command = lambda: clickButton(str(value)))
                        
                    
                    
                    
    def clickButton( value):
                            
                            '''
                            DOCSTRING: Method to program the actions that will happen in the calculator after a click of each button
                            INPUT: value of the button (1,2,3,4,5,6,7,8,9,0,+,-,*,/,c,=)
                            OUTPUT: what action will be performed when a particular button is clicked
                            '''
                            
                            #Get the equation that's entered by the user
                            current_equation=str(equation.get())
                            
                            #If user clicked "c", then clear the screen
                            if value == 'c':
                                equation.delete(-1, END)
                            
                            #If user clicked "=", then compute the answer and display it
                            elif value == '=':
                                answer = str(eval(current_equation))
                                equation.delete(-1, END)
                                equation.insert(0, answer)
                                c.execute('INSERT INTO '+profileemail+profikelastname+profikelastname+' VALUES (:user)',{
                    "user":(" calculetor:"+equation.get()+"="+answer+" :"+(str(x))),
                    })
                            elif value == "back":
                                root.destroy()
                            elif value == "^":
                                equation.insert((clickButton(str("**"))))
                            elif value == "÷":
                                equation.insert((clickButton(str("/"))))
                            elif value == "×":
                                equation.insert((clickButton(str("*"))))
                            elif value == "mod":
                                equation.insert((clickButton(str("%"))))
                            elif value == "div":
                                equation.insert((clickButton(str("//"))))
                            
                            
                                
                            
                            #If user clicked any other button, then add it to the equation line
                            else:
                                equation.delete(0, END)
                                equation.insert(-1, current_equation+value)
                            if value == "r":
                                import math
                                mn10 =math.sqrt(int(eval(current_equation)))
                                answer = str(mn10)
                                equation.delete(-1, END)
                                equation.insert(0, answer)
    if profilename1 == "200":
        master = tabview.tab("حساب ماشین")
    else:
        master = tabview.tab("calculet")               
    #Add a name to our application
    
    #Create a line where we display the equation
    equation=Entry(master, width=36, borderwidth=5)
    #Assign a position for the equation line in the grey application window
    equation.grid(row=0, column=0, columnspan=4, padx=10, pady=10)
    #Execute the .creteButton() method
    createButton()
    def filter_values(*args):
        keyword = search_keyword.get().lower()
        filtered_values = [value for value in values if keyword in value.lower()]
        combobox_1['values'] = filtered_values

    search_keyword.trace('w', filter_values)
    if profilename1 == "200":
        string_input_button = customtkinter.CTkButton(tabview.tab("ترجمگر"), text="ترجمه",font=customtkinter.CTkFont("B Homa",13), command=open_input_dialog_event)
    else:
        string_input_button = customtkinter.CTkButton(tabview.tab("Transletor"), text="Transletor", command=open_input_dialog_event)
    string_input_button.grid(row=2, column=0, padx=20, pady=(10, 10))
    
    if profilename1 == "200":
        combobox_12 = customtkinter.CTkComboBox(tabview.tab("بات چت"),
                                                        values=[ "Wolfram", "Wikipedia"])
    else:
        combobox_12 = customtkinter.CTkComboBox(tabview.tab("ChatBot"),
                                                        values=["Chatgpt", "Wolfram", "Wikipedia"])
    combobox_12.grid(row=1, column=0, padx=20, pady=(10, 10))
    if profilename1 == "200":
        entry2 = customtkinter.CTkEntry(tabview.tab("بات چت"), placeholder_text=".... متن",justify=RIGHT,font=customtkinter.CTkFont("B Homa",13))
    else:
        entry2 = customtkinter.CTkEntry(tabview.tab("ChatBot"), placeholder_text="Text ...")
    entry2.grid(row=0, column=0,  padx=20, pady=(1, 1), sticky="nsew")
    entry2.bind("<Return>", open_input_dialog_event1)
    if profilename1 == "200":
        string_input_button1 = customtkinter.CTkButton(tabview.tab("بات چت"), text="جستجو",font=customtkinter.CTkFont("B Homa",13),
                                                       command=open_input_dialog_event1)
    else:
        string_input_button1 = customtkinter.CTkButton(tabview.tab("ChatBot"), text="Search",
                                                       command=open_input_dialog_event1)
    string_input_button1.grid(row=2, column=0, padx=20, pady=(10,10))
    
    
    tabview1 = customtkinter.CTkScrollableFrame(root,label_text='ها اپلیکیشن',corner_radius=0,width=261)
    tabview1.grid(row=0, column=4, padx=20, pady=(10,10), sticky="nsew")
    
    

    # لیست تصاویر برای گالری
    image_paths = ["lomge/[11].png", "lomge/[12].png", "lomge/[13].png", "lomge/[14].png"]

    # زمان تغییر تصاویر بین صفحات (به میلی ثانیه)
    interval = 2000

    # ایجاد پنجره اصلی Tkinter
    
    # ایجاد گالری
    
    # Add Button and Label
    if profilename1 == "200":
        weather11=tabview.tab("هوا و آب")
    else:
        weather11=tabview.tab("Weather")
    
    
 
    
    
    
    # Create the notebook to hold the notes gallery
    
    # create slider and progressbar frame
    
    
    # create scrollable frame
    if profilename1 == "200":
        scrollable_frame = customtkinter.CTkScrollableFrame(root, label_text="کاربر مشخصات",corner_radius=0)
    else:
        scrollable_frame = customtkinter.CTkScrollableFrame(root, label_text="Profile",corner_radius=0)
    scrollable_frame.grid(row=1, column=3, padx=20, pady=(10, 10), sticky="nsew")
    scrollable_frame.grid_columnconfigure(0, weight=1)
    scrollable_frame_switches = []
    import sqlite3
    conn = sqlite3.connect("user1.db")
    c = conn.cursor()
    
    c.execute("SELECT * FROM user_info")
    results = c.fetchall()
        
    if results:
            for i in results:
                profilename = i[0]
                profikelastname = i[1]
                profileemail = i[2]
                profiletime = i[3]
                profilenumber = i[4]
                profileborn = i[5]
                profilemr = i[6]
                profileschool = i[7]
    if profilename1 == "200":
        name = customtkinter.CTkLabel(master=scrollable_frame, text="نام :"+profilename,font=customtkinter.CTkFont("B Homa",13))
    else:
        name = customtkinter.CTkLabel(master=scrollable_frame, text="Name: "+profilemr+"."+profilename,font=customtkinter.CTkFont("B Homa",13))
    name.grid(row=0, column=0,  padx=10, pady=5, sticky="")
    if profilename1 == "200":
        last_name = customtkinter.CTkLabel(master=scrollable_frame, text="نام خانوادگی :"+profikelastname,font=customtkinter.CTkFont("B Homa",13))
    else:
        last_name = customtkinter.CTkLabel(master=scrollable_frame, text="Last Name: "+profikelastname,font=customtkinter.CTkFont("B Homa",13))
    last_name.grid(row=1, column=0,  padx=10, pady=5, sticky="")
    if profilename1 == "200":
        email = customtkinter.CTkLabel(master=scrollable_frame, text="نام کاربری :"+profileemail,font=customtkinter.CTkFont("B Homa",13))
    else:
        email = customtkinter.CTkLabel(master=scrollable_frame, text="Email: "+profileemail)
    email.grid(row=2, column=0,  padx=10, pady=5, sticky="")
    if profilename1 == "200":
        number = customtkinter.CTkLabel(master=scrollable_frame, text="شماره تلفن :"+profilenumber,font=customtkinter.CTkFont("B Homa",13))
    else:
        number = customtkinter.CTkLabel(master=scrollable_frame, text="Phone number: "+profilenumber)
    number.grid(row=3, column=0,  padx=10, pady=5, sticky="")
    if profilename1 == "200":
        timee = customtkinter.CTkLabel(master=scrollable_frame, text="شماره تلفن :"+profileborn,font=customtkinter.CTkFont("B Homa",13))
    else:
        timee = customtkinter.CTkLabel(master=scrollable_frame, text="Time: "+profiletime)
    timee.grid(row=4, column=0,  padx=10, pady=5, sticky="")
    if profilename1 == "200":
        age = customtkinter.CTkLabel(master=scrollable_frame, text="نام مدرسه :"+profileschool,font=customtkinter.CTkFont("B Homa",13))
    else:
        age = customtkinter.CTkLabel(master=scrollable_frame, text="Born: "+profileborn)
    age.grid(row=5, column=0,  padx=10, pady=5, sticky="")
    
    
    
    # create checkbox and switch frame
    if profilename1 == "200":
        checkbox_slider_frame = customtkinter.CTkScrollableFrame(root, label_text="برنامه",corner_radius=0)
    else:
        checkbox_slider_frame = customtkinter.CTkScrollableFrame(root, label_text="App",corner_radius=0)
    checkbox_slider_frame.grid(row=1, column=4, padx=20, pady=(10, 10), sticky="nsew")
    if profilename1 == "200":
        sidebar_button_33 = customtkinter.CTkButton(checkbox_slider_frame, text="نقشه جهان",command=map,font=customtkinter.CTkFont("B Homa",13))
    else:
        sidebar_button_33 = customtkinter.CTkButton(checkbox_slider_frame, text="World Map",command=map,font=customtkinter.CTkFont("B Homa",13))
    sidebar_button_33.grid(row=1, column=0, padx=20, pady=(10, 10))
    if profilename1 == "200":
        sidebar_button_333 = customtkinter.CTkButton(checkbox_slider_frame, text="تاریخ عمل کردم",command=history,font=customtkinter.CTkFont("B Homa",13))
    else:
        sidebar_button_333 = customtkinter.CTkButton(checkbox_slider_frame, text="History",command=history)
    sidebar_button_333.grid(row=2, column=0, padx=20, pady=(10, 10))
    if profilename1 == "200":
        sidebar_button_333 = customtkinter.CTkButton(checkbox_slider_frame, text="بروز شدن برنامه",command=Reset,font=customtkinter.CTkFont("B Homa",13))
    else:
        sidebar_button_333 = customtkinter.CTkButton(checkbox_slider_frame, text="Reset",command=Reset)
    sidebar_button_333.grid(row=4, column=0, padx=20, pady=(10, 10))
    
    def change_map( new_map: str):
        if new_map == "OpenStreetMap":
            map_widget.set_tile_server("https://a.tile.openstreetmap.org/{z}/{x}/{y}.png")
        elif new_map == "Google normal":
            map_widget.set_tile_server("https://mt0.google.com/vt/lyrs=m&hl=en&x={x}&y={y}&z={z}&s=Ga", max_zoom=22)
        elif new_map == "Google satellite":
            map_widget.set_tile_server("https://mt0.google.com/vt/lyrs=s&hl=en&x={x}&y={y}&z={z}&s=Ga", max_zoom=22)
    map_option_menu = customtkinter.CTkOptionMenu(checkbox_slider_frame, values=["OpenStreetMap", "Google normal", "Google satellite"],
                                                                   command=change_map)
    map_option_menu.grid(row=3, column=0, padx=20, pady=10)
    
    
    
    
    
    
    
    
    
    # set default values
    sidebar_button_3.configure(state="disabled", text="Disabled CTkButton")
    
    
    
    
    scaling_optionemenu.set("100%")
    combobox_1.set("Laguege")
    import random
    text="سلام به نام خدا من ساسان هستم"
    from time import sleep
    words = text.split()
    reversed_text = ' '.join(words[::-1])
    texts="""⠀⠀⠀⠀⠀⠀⠀⠀⣀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⠀⢹⣿⣭⣑⠒⠦⠤⢀⣀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⢀⣀⡸⠿⠿⣿⣿⣷⣦⣤⣄⣈⡉⠛⠲⢤⣀⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠲⢾⣿⣿⣶⣶⣿⣿⣿⣿⣿⣿⣿⣿⣿⣷⣶⣦⠈⢷⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠈⠙⣿⣿⣿⣿⣿⣿⣿⣿⣿⠛⠿⣿⠋⢿⣇⡈⡇⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⢀⣾⣿⣿⣿⣿⣿⣿⣿⢹⡟⠀⠀⠀⠀⠘⣿⣿⢹⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⢀⣾⣿⣿⣿⣿⣿⣿⣿⣿⠈⠧⣀⠀⠀⡠⠂⠃⣿⡞⡄⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⢀⣾⣿⡟⢿⣿⣿⣿⡿⠘⠿⠃⠀⠀⠈⠙⠛⢿⣷⡃⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⢸⣿⢿⣧⠸⣿⣿⣿⡇⠀⠀⠀⠀⠀⠀⠀⠀⢹⡇⠛⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⠀⢸⣿⣷⣿⣿⣿⡇⠀⠀⠀⠀⠀⠀⠀⢀⣼⡇⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⠀⠘⠟⠁⢿⣿⣿⣧⠄⠀⡀⢀⡰⠖⠒⢿⣿⡇⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⣀⡀⠀⠀⠀⠀⢸⣁⡽⢿⣀⣐⠡⠂⠱⡄⠀⠀⠙⠇⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⣮⣏⡅⠀⠀⢠⡇⠀⠀⠀⠟⠁⠀⠀⣰⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠏⠶⡄⠀⢀⣀⡟⣦⣀⢸⣀⣀⣠⣾⣯⠕⠀⠀⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⣧⢲⠇⢸⡇⣚⣿⣿⣿⣯⣿⣿⣶⣶⡾⡀⠀⠀⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⡶⠺⡃⠀⠱⣻⣿⣿⣿⣿⣿⣿⣿⣿⣿⢹⠀⠀⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⣶⠭⠅⠀⠻⢻⢼⣿⣿⣿⣿⣿⣿⣿⣿⡘⠀⠀⠀⠀⠀⠀⠀⠀⠀
        ⣏⣚⠁⠀⠀⣸⣼⣿⣿⠟⠉⠻⣿⣿⡿⡆⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⣯⢭⡇⠀⠀⢸⣻⣿⠁⠀⠀⠀⠈⢻⣿⠇SASUKE⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠁⠀⠀⠀⠀⠁⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
"""
    texts1="""
⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠤⣄⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⡆⠀⠀⠀⠀
⠀⠀⠀⠀⠀⢀⣤⣴⣶⣶⣶⣶⣶⣦⣤⡄⠊⠀⠀⠀⠀⠀
⠀⠦⣤⣶⣾⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣷⣄⠀⠀⠀⠀
⠀⠀⣼⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣷⡀⠀⠀
⠀⣸⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣧⠀⠀
⣠⣿⣿⣿⡿⢻⣿⣿⣿⣿⣿⣿⢿⣿⣿⣿⣿⣿⣿⣿⡧⠀
⠉⢹⣿⣿⢓⣺⡿⠟⠛⠛⢻⣿⣼⣿⣿⣿⣿⣿⣿⣿⣇⠀
⠀⢸⣿⣿⠉⠀⠀⠀⠀⠀⠀⠀⠀⢀⣿⡿⠿⢿⣿⣿⣿⠀
⡀⠈⣿⣿⡀⠀⠀⠀⠄⠀⠀⠀⠀⢸⣿⣗⠏⢺⣿⣿⣿⡇
⠀⢀⡈⣿⣷⣤⣄⣀⠀⠀⠀⢀⣤⣿⣿⣶⣾⣿⣿⣿⣿⣧
⠀⠘⠉⢹⣿⣿⣿⣿⣿⡇⠀⢸⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿
⠀⠀⠀⣾⣿⣿⠟⡿⠟⡁⠄⠚⠉⠀⠘⢿⣿⣿⣿⣿⣿⡇
⠀⠀⠀⣿⠟⠁⠈⠉⠁⠀⠀⠀⠀⠀⠀⠀⠙⡿⠿⠏⠿⠃
⠀⠀⠀⡜⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠘⡄⠀⠀⠀"""
    texts2="""
⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀  ⣤⣶⣶⣶⣶⣶⣦⣄⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⢰⣿⣿⣿⣿⣿⣿⣿⣿⣿⣷⡄⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⣠⢿⣿⣿⡿⣿⣿⣿⣿⣿⣿⣿⣿⣀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⣰⣿⣿⣿⣿⡇⣿⣷⣿⣿⣿⣿⣿⣿⣯⡄⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⡰⣿⣿⣿⣇⣿⣀⠸⡟⢹⣿⣿⣿⣿⣿⣿⣷⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⢀⢡⣿⣿⣿⡇⠝⠋⠀⠀⠀⢿⢿⣿⣿⣿⣿⣿⡇⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠸⢸⠸⣿⣿⣇⠀⠀⠀⠀⠀⠀⠊⣽⣿⣿⣿⠁⣷⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⢿⣿⣿⣷⣄⠀⠀⠀⢠⣴⣿⣿⣿⠋⣠⡏⡄⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠐⠾⣿⣟⡻⠉⠀⠀⠀⠈⢿⠋⣿⡿⠚⠋⠁⡁⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⢀⣴⣶⣾⣿⣿⡄⠀⣳⡶⡦⡀⣿⣿⣷⣶⣤⡾⠁⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⢸⣿⣿⣿⣿⣿⣿⡆⠀⡇⡿⠉⣺⣿⣿⣿⣿⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⢸⣿⣿⣿⣿⣿⣯⠽⢲⠇⠣⠐⠚⢻⣿⣿⣿⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⣿⣿⣿⣿⣿⡐⣾⡏⣷⠀⠀⣼⣷⡧⣿⣿⣦⣄⡀⠀⠀⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⣻⣿⣿⣿⣿⣿⣮⠳⣿⣇⢈⣿⠟⣬⣿⣿⣿⣿⣿⡦⢄⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⠀⠀⢀⢄⣾⣿⣿⣿⣿⣿⣿⣿⣷⣜⢿⣼⢏⣾⣿⣿⣿⢻⣿⣝⣿⣦⡑⢄⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⣠⣶⣷⣿⣿⠃⠘⣿⣿⣿⣿⣿⣿⣿⡷⣥⣿⣿⣿⣿⣿⠀⠹⣿⣿⣿⣷⡀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⣇⣤⣾⣿⣿⡿⠻⡏⠀⠀⠸⣿⣿⣿⣿⣿⣿⣮⣾⣿⣿⣿⣿⡇⠀⠀⠙⣿⣿⡿⡇⠀⠀⠀⠀⠀
⠀⠀⢀⡴⣫⣿⣿⣿⠋⠀⠀⡇⠀⠀⢰⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⡇⠀⠀⠀⢘⣿⣿⣟⢦⡸⠀⠀⠀
⠀⡰⠋⣴⣿⣟⣿⠃⠀⠀⠀⠈⠀⠀⣸⣿⣿⣿⣿⣿⣿⣇⣽⣿⣿⣿⣿⣇⠀⠀⠀⠁⠸⣿⢻⣦⠉⢆⠀⠀
⢠⠇⡔⣿⠏⠏⠙⠆⠀⠀⠀⠀⢀⣜⣛⡻⢿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⡿⡀⠀⠀⠀⠀⡇⡇⠹⣷⡈⡄⠀
⠀⡸⣴⡏⠀⠀⠀⠀⠀⠀⠀⢀⣾⣿⣿⣿⣿⣿⣿⣿⣿⣿⣻⣿⣿⣿⣿⣿⣿⡄⠀⠀⠀⡇⡇⠀⢻⡿⡇⠀
⠀⣿⣿⣆⠀⠀⠀⠀⠀⠀⢀⣼⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⡀⠀⣰⠿⠤⠒⡛⢹⣿⠄
⠀⣿⣷⡆⠁⠀⠀⠀⠀⢠⣿⣿⠟⠻⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⡟⠻⢷⡀⠀⠀⠀⠀⠀⣸⣿⠀
⠀⠈⠿⢿⣄⠀⠀⠀⢞⠌⡹⠁⠀⠀⢻⡇⠹⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⠁⢳⠀⠀⠁⠀⠀⠀⠀⢠⣿⡏⠀
⠀⠀⠀⠈⠂⠀⠀⠀⠈⣿⠁⠀⠀⠀⡇⠁⠀⠘⢿⣿⣿⠿⠟⠋⠛⠛⠛⠀⢸⠀⠀⡀⠂⠀⠀⠐⠛⠉⠀⠀
⠀⠀⠀⠀⠀⠀⠀⠀⠀⠐⠕⣠⡄⣰⡇⠀⠀⠀⢸⣧⠀⠀⠀⠀⠀⠀⠀⢀⣸⠠⡪⠊⠀⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠈⢫⣽⡋⠭⠶⠮⢽⣿⣆⠀⠀⠀⠀⢠⣿⣓⣽⠁⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⣿⣿⣿⣿⣿⣿⣿⣿⢹⣶⣦⣾⣿⣿⣿⡏⠀⠀⠀⠀⠀⠀⠀⠀⠀"""
    textbox.insert("0.0", texts1)
    
   
    
    
    conn.commit()
    roote.mainloop()
#################################################################################
def StartClass():
    import os
    import customtkinter as ctk
    from tkinter import filedialog
    import sqlite3
    import customtkinter
    # تابع برای دریافت لیست فایل‌های اکسل در فولدر x
    roote = Tk()
    roote.geometry("250x300")
    roote.title("شروع برنامه")
    roote.iconbitmap(bitmap = 'Icons\icon.ico')
    def get_excel_files(folder_path):
        
         return [f for f in os.listdir(folder_path) if f.endswith('.xlsx')]
        
     # تابع برای نوشتن نام فایل انتخاب شده در فایل متنی
    conn = sqlite3.connect('user1.db')
    c = conn.cursor()               
    profilename = "200"
    c.execute("SELECT * FROM Languge")
    results = c.fetchall()
    if results:
        for i in results:
            profilename1 = i[0]
        conn.commit()
        #########################
    def on_ok_click():
            print(listbox.get())
            email = listbox.get()
            while True:
                        
                        conn = sqlite3.connect("user1.db")
                        c = conn.cursor()
                        c.execute("INSERT INTO Languge VALUES (:usertext)",
                            {
                                "usertext": email,
                                
                        
                            })
            
                        conn.commit()
                        roote.destroy()
                        robo()
            
    # ایجاد پنجره اصلی
    

    root = customtkinter.CTkFrame(roote,fg_color=("white"),width=1000,height=1000, corner_radius=0)
    root.grid(row=0, column=0, rowspan=4, sticky="nsew")
    # انتخاب فولدر x


    # دریافت لیست فایل‌های اکسل
    conn = sqlite3.connect("user1.db")
    c = conn.cursor()
    c.execute("SELECT * FROM user_profile_techer")
    results = c.fetchall()
        
    if results:
                for i in results:
                    name = i[0]
                    last_name = i[1]
    folder_name = "معلم_"+name+"_"+last_name
    desktop_path = os.path.join(os.path.join(os.environ['USERPROFILE']), "Desktop")
    exfilename=desktop_path+"\\"+folder_name
    excel_files = (get_excel_files(exfilename))
    excel_files = [file.replace(".xlsx","")for file in excel_files]
    excel_files_error = ["ندارد وجود کلاسی هنوز"]
    print(excel_files)
    if excel_files == [] :
        listbox = ctk.CTkComboBox(root, values=excel_files_error)
    else:
        listbox = ctk.CTkComboBox(root, values=excel_files)
    # ایجاد لیست باکس برای نمایش فایل‌های اکسل

    for file in excel_files:
        file1 = file
        valu = file
    
    listbox.place(x=60,y=100)

    # تابع برای دکمه OK

    heading=Label(root,text="""کلاسی که امروز تدریس       
            میکنی رو انتخاب کن """,bg="white",font=('B Kamran Outline',14))
    heading.place(x=40,y=10)
    # ایجاد دکمه OK
    ok_button = ctk.CTkButton(root, text="OK", command=on_ok_click)
    ok_button.place(x=60,y=140)

    # اجرای حلقه اصلی
    roote.mainloop()
#################################################################################
def robo():
    import requests
    import datetime
    import pytz
    import tkinter as tk
    import time
    import math
    import tkinter
    import tkinter.messagebox
    import customtkinter
    import sqlite3
    conn = sqlite3.connect('user1.db')
    c = conn.cursor()               
    profilename = "200"

    c.execute("SELECT * FROM Languge")
    results = c.fetchall()
    if results:
        for i in results:
            classpresen = i[0]
    conn.commit()
    profilename1 = "200"
    
    roote = Tk()
    roote.title("Robosan")
    roote.iconbitmap(bitmap = 'Icons\\icon.ico')
    roote.geometry(f"1390x560")
    fg_cccc = " "
    def dark():
        global fg_cccc
        if switch.get() == "on":
            customtkinter.set_appearance_mode("Dark")
            fg_cccc="dark"
        else:
            customtkinter.set_appearance_mode("Light")
            fg_cccc="light"
    root1 = customtkinter.CTkFrame(roote,fg_color=("#FFFFFF","#2E2E2E"), width=140, corner_radius=0)
    root1.grid(row=0, column=0, rowspan=4, sticky="nsew")
    
    #################
    import datetime
        
    conn = sqlite3.connect('user1.db')
    c = conn.cursor()
        
    c.execute("SELECT * FROM user_profile_techer")
    results = c.fetchall()
    now = datetime.datetime.now()
    now = datetime.datetime.today()
    mm = str(now.month)
    dd = str(now.day)
    yyyy = str(now.year)
    hour = str(now.hour)
    mi = str(now.minute)
    ss = str(now.second)
    x=str(mm + "/" + dd + "/" + yyyy + " " + hour + ":" + mi + ":" + ss )
    if results:
            for i in results:
                profilename = i[0]
                profikelastname = i[1]
                profileemail = i[2]
                profiletime = i[3]
    
    # Create the table for storing contact information
    conn.execute('CREATE TABLE IF NOT EXISTS '+profileemail+profikelastname+profikelastname+' (user text);')
    conn.commit()
    print(("open :"+(str(x))))
    c.execute('INSERT INTO '+profileemail+profikelastname+profikelastname+' VALUES (:user)',{
        "user":("open (برنامه) :"+(str(x))),
        })
    conn.commit()
    # configure grid layout (4x4)
    
    def robo2(tip: str):
        if profilename1 == "200":
         if tip == "ما درباره":
            ("ما درباره شد باز"+(str(x)))
            Introduction()
         elif tip == "دفترچه":
            ("دفترچه شد باز"+(str(x)))
            note()
         elif tip == "حساب ماشین":
            ("حساب ماشین شد باز"+(str(x)))
            cul()
         elif tip == "trans":
            ("open trans :"+(str(x)))
            tans1()
         elif tip == "search":
            ("open search :"+(str(x)))
            search()
         elif tip == "BMI":
            ("شد باز BMI :"+(str(x)))
            BMI()
         elif tip == "گوگل در جستجو":
            ("گوگل در جستجو شد باز"+(str(x)))
            google()
         elif  tip == "هوا و آب":
            ("هوا و آب شد باز"+(str(x)))
            weather()
         elif  tip == "کد مورس":
            ("مورس کد شد باز"+(str(x)))
            mors()
         elif  tip == "پلیر موزیک":
            ("پلیر موزیک شد باز"+(str(x)))
            mus()
         elif  tip  == "شمار ثانیه":
            ("شمار ثانیه شد باز"+(str(x)))
            corno1()
         elif  tip == "دوز بازی":
            ("دوز بازی شد باز"+(str(x)))
            Tic1()
         elif  tip == "متن ادیت":
            ("متن ادیت شد باز"+(str(x)))
            edit_text1()
         elif  tip == "انس گرم پند محاسبه":
            ("انس گرم پند محاسبه شد باز"+(str(x)))
            Grame()
         elif  tip == "تقویم":
            ("تقویم شد باز"+(str(x)))
            calendare1()
         elif  tip == "کد ار یو کی":
            ("کد ار یو کی شد باز"+(str(x)))
            Qr1()
         elif  tip == "تلفن دفترچه":
            ("تلفن دفتر شد باز"+(str(x)))
            phone1()
         elif  tip == "من روزانه برنامه":
            ("من روزانه برنامه شد باز"+(str(x)))
            todo1()
         elif  tip == "بات چت":
            ("بات چت شد باز"+(str(x)))
            chatbot1()
        else:
         if tip == "Introduction":
            ("open Introduction :"+(str(x)))
            Introduction()
         elif tip == "note book":
            ("open note book :"+(str(x)))
            note()
         elif tip == "calculator":
            ("open calculator :"+(str(x)))
            cul()
         elif tip == "trans":
            ("open trans :"+(str(x)))
            tans1()
         elif tip == "search":
            ("open search :"+(str(x)))
            search()
         elif tip == "BMI":
            ("open BMI :"+(str(x)))
            BMI()
         elif tip == "search google":
            ("open search google :"+(str(x)))
            google()
         elif  tip == "weather":
            ("open weather :"+(str(x)))
            weather()
         elif  tip == "mourse code":
            ("open mourse code :"+(str(x)))
            mors()
         elif  tip == "music player":
            ("open music player :"+(str(x)))
            mus()
         elif  tip  == "Countdown":
            ("open Countdown :"+(str(x)))
            corno1()
         elif  tip == "Tic Tac":
            ("open Tic Tac :"+(str(x)))
            Tic1()
         elif  tip == "Edit Text":
            ("open Edit Text :"+(str(x)))
            edit_text1()
         elif  tip == "Gram Pound Ounce":
            ("open Gram Pound Ounce :"+(str(x)))
            Grame()
         elif  tip == "calendar":
            ("open calendar :"+(str(x)))
            calendare1()
         elif  tip == "QrCode":
            ("open QrCode :"+(str(x)))
            Qr1()
         elif  tip == "Phone Book":
            ("open Phone Book :"+(str(x)))
            phone1()
         elif  tip == "todo":
            ("open todo :"+(str(x)))
            todo1()
         elif  tip == "Chat Bot":
            ("open Chat Bot :"+(str(x)))
    def menu():
        pass
    def tans1():
        c.execute('INSERT INTO '+profileemail+profikelastname+profikelastname+' VALUES (:user)',{
        "user":("open trans :"+(str(x))),
        })
        conn.commit()
        trans1()
    def openfile():
        pass
    def savefile():
        pass
    def Introduction():
        c.execute('INSERT INTO '+profileemail+profikelastname+profikelastname+' VALUES (:user)',{
        "user":("open (ما درباره) :"+(str(x))),
        })
        conn.commit()    
        Introduction1()
    def pack():
        pass
    
    def note():
        c.execute('INSERT INTO '+profileemail+profikelastname+profikelastname+' VALUES (:user)',{
        "user":("open (دفترچه) :"+(str(x))),
        })
        conn.commit()
        note1()
        
    def corno1():
        c.execute('INSERT INTO '+profileemail+profikelastname+profikelastname+' VALUES (:user)',{
        "user":("open (تایمر) :"+(str(x))),
        })
        conn.commit()
        roote.destroy()
        corno()
        robo()
        
    def profile():
        c.execute('INSERT INTO '+profileemail+profikelastname+profikelastname+' VALUES (:user)',{
        "user":("open Profile :"+(str(x))),
        })
        conn.commit()
        Profile1()
    def cul():
        c.execute('INSERT INTO '+profileemail+profikelastname+profikelastname+' VALUES (:user)',{
        "user":("open Calculator :"+(str(x))),
        })
        conn.commit()
        
        cul1()
    def search():
        c.execute('INSERT INTO '+profileemail+profikelastname+profikelastname+' VALUES (:user)',{
        "user":("open Search :"+(str(x))),
        })
        conn.commit()
        
        search1()
        os.system("python rock.py")
    def BMI():
        c.execute('INSERT INTO '+profileemail+profikelastname+profikelastname+' VALUES (:user)',{
        "user":("open BMI :"+(str(x))),
        })
        conn.commit()
        roote.destroy()
        BMI1()
         
    def Grame():
        c.execute('INSERT INTO '+profileemail+profikelastname+profikelastname+' VALUES (:user)',{
        "user":("open (جرمی محاسبگر) :"+(str(x))),
        })
        conn.commit()
        Gram()
        
    def Qr1():
        c.execute('INSERT INTO '+profileemail+profikelastname+profikelastname+' VALUES (:user)',{
        "user":("open QrCode :"+(str(x))),
        })
        conn.commit()
        os.system("python Qr.py")
        
    def weather():
        c.execute('INSERT INTO '+profileemail+profikelastname+profikelastname+' VALUES (:user)',{
        "user":("open Weather :"+(str(x))),
        })
        conn.commit()
        
        weather1()
    def mors():
        c.execute('INSERT INTO '+profileemail+profikelastname+profikelastname+' VALUES (:user)',{
        "user":("open (کد مورس) :"+(str(x))),
        })
        conn.commit()
        
        mors1()
    def mus():
        c.execute('INSERT INTO '+profileemail+profikelastname+profikelastname+' VALUES (:user)',{
        "user":("open (پلیر موزیک) :"+(str(x))),
        })
        conn.commit()
                           
        mus1()
    def backe():
        conn.commit()
        roote.destroy()
        load()
    def google():
        c.execute('INSERT INTO '+profileemail+profikelastname+profikelastname+' VALUES (:user)',{
        "user":("open google :"+(str(x))),
        })
        conn.commit()
        os.system("python google.py")
    def help1():
        c.execute('INSERT INTO '+profileemail+profikelastname+profikelastname+' VALUES (:user)',{
        "user":("open Help :"+(str(x))),
        })
        conn.commit()
        help()
    def about():
        
        os.system("python about.py")   
    def cmd():
        
        os.system("python robosan.py")    
    def timeis():
        pass
    def langu():
        try:
           c.execute('INSERT INTO '+profileemail+profikelastname+profikelastname+' VALUES (:user)',{
            "user":("open Edit languge :"+(str(x))),
            })
           conn.commit()   
           GoogleTranslator(target="fa").translate("Please register or Login to your account")
           roote.destroy()
            
        except:
            messagebox.showerror("Error","No Network")
    def Tic1():
        c.execute('INSERT INTO '+profileemail+profikelastname+profikelastname+' VALUES (:user)',{
        "user":("open (دوز بازی) :"+(str(x))),
        })
        conn.commit()
        Tic()
    def edit_text1():
        c.execute('INSERT INTO '+profileemail+profikelastname+profikelastname+' VALUES (:user)',{
        "user":("open (متن ادیتور) :"+(str(x))),
        })
        conn.commit()
        edit_text()
    def calendare1():
        c.execute('INSERT INTO '+profileemail+profikelastname+profikelastname+' VALUES (:user)',{
        "user":("open Calendare :"+(str(x))),
        })
        conn.commit()
        calendare1()
    def phone1():
        c.execute('INSERT INTO '+profileemail+profikelastname+profikelastname+' VALUES (:user)',{
        "user":("open (تلفن دفتر) :"+(str(x))),
        })
        conn.commit()
        phone()
    def todo1():
        c.execute('INSERT INTO '+profileemail+profikelastname+profikelastname+' VALUES (:user)',{
        "user":("open (کارهام دفتر) :"+(str(x))),
        })
        conn.commit()
        todo()
    def chatbot1():
        c.execute('INSERT INTO '+profileemail+profikelastname+profikelastname+' VALUES (:user)',{
        "user":("open Chat Bot :"+(str(x))),
        })
        conn.commit()
        chatbot()
    def map():
        c.execute('INSERT INTO '+profileemail+profikelastname+profikelastname+' VALUES (:user)',{
        "user":("open World Map :"+(str(x))),
        })
        conn.commit()
        os.system("python map.py")
        
    def open_input_dialog_event():
        try:
            from deep_translator import GoogleTranslator
            textbox.delete('0.0',END)
            dialog = combobox_1.get()
            
             
            def get_country_name(language_name):
                    from googletrans import LANGUAGES
                    for code, name in LANGUAGES.items():
                        if name == language_name:
                            return code
            language_name = dialog
            country_name = get_country_name(language_name)
            output =GoogleTranslator(target=country_name).translate(entry1.get())
            textbox.insert("0.0", "Transletor:\n" +str(output)+"\n")
        except:
            messagebox.showerror("Warning", "not internet")  
    
            
    def open_input_dialog_event1():   
       try:

        import openai
        openai.api_key = "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"
        # set up the GPT model
        model_engine = "text-davinci-003"
        prompt = "Hello! I'm a chatbot. How can I help you?"
        textbox.delete('0.0',END)
        if combobox_12.get() == "Wolfram":
            import wolframalpha
            client=wolframalpha.Client('26AW5W-23YY3TTWHG')
            res = client.query(entry2.get())
            answer = next(res.results).text
            textbox.insert("0.0", "Wolfram:\n" +str(answer))
        elif combobox_12.get() == "Wikipedia":
            import wikipedia
            Iran = wikipedia.summary(entry2.get())
            textbox.insert("0.0", "Wikipedia:\n" +str(Iran))
       except:
           messagebox.showerror("we have problem","We have a problem pls ckek leter") 
    
    def change_appearance_mode_event(new_appearance_mode: str):
        customtkinter.set_appearance_mode(new_appearance_mode)
    def change_scaling_event(new_scaling: str):
        new_scaling_float = int(new_scaling.replace("%", "")) / 100
        customtkinter.set_widget_scaling(new_scaling_float)
    def sidebar_button_event():
        backe()
    
    def Reset():
        roote.destroy()
        robo()
    def history():
        textbox.delete('0.0',END)
        import sqlite3
        conn = sqlite3.connect('user1.db')
        c = conn.cursor()
            
        c.execute("SELECT * FROM user_profile_techer")
        results = c.fetchall()
        textbox.delete('0.0',END)
        textbox.insert("0.0", "History:\n")
        if results:
                for i in results:
                    profilename = i[0]
                    profikelastname = i[1]
                    profileemail = i[2]
                    profiletime = i[3]
        for row in c.execute('SELECT * FROM '+profileemail+profikelastname+profikelastname+''):
            textbox.configure(state='normal')
            textbox.insert('end', "User: "+row[0] + "\n")
            textbox.configure(state='disabled')
            
    #creat bg
    home_frame = customtkinter.CTkFrame(root1, corner_radius=0, fg_color="transparent")
    home_frame.grid(row=0, column=2,sticky="nsew")

    second_frame = customtkinter.CTkFrame(root1, corner_radius=0, width=1070,height=564 , fg_color="transparent")
    conn = sqlite3.connect("user1.db")
    cursor = conn.cursor()
    c = conn.cursor()
    
    c.execute("SELECT * FROM user_profile_techer")
    results = c.fetchall()
    if results:
        for i in results:
            name_maneger = str(i[6])
            last_name_maneger = str(i[7])



    manager_frame = customtkinter.CTkFrame(root1, corner_radius=0, width=1070,height=564, fg_color="transparent")
    root = home_frame

    def select_frame_by_name(name):
        # set button color for selected button
        home_button.configure(fg_color=("gray75", "gray25") if name == "home" else "transparent")
        frame_2_button.configure(fg_color=("gray75", "gray25") if name == "chat" else "transparent")
        frame_3_button.configure(fg_color=("gray75", "gray25") if name == profikelastname+" "+profilename  else "transparent")

        # show selected frame
        if name == "home":
            home_frame.grid(row=0, column=2, sticky="nsew")
        else:
            home_frame.grid_forget()
        if name == "chat":
            second_frame.grid(row=0, column=2, sticky="nsew")

        else:
            second_frame.grid_forget()
        if name == "maneger":
            manager_frame.grid(row=0, column=3, sticky="nsew")

        else:
            manager_frame.grid_forget()

    def home_button_event():
        select_frame_by_name("home")

    def chat_button_event():
        select_frame_by_name("chat")

    def maneger_button_event():
        select_frame_by_name("maneger")
    
    


    navigation_frame = customtkinter.CTkFrame(root1,width=140, corner_radius=0)
    navigation_frame.grid(row=0, column=0,rowspan=4, sticky="nsew")
    navigation_frame.grid_rowconfigure(4, weight=1)
    from PIL import Image
    home_image = customtkinter.CTkImage(light_image=Image.open(os.path.join("test_images\\home_dark.png")),
                                                 dark_image=Image.open(os.path.join("test_images\\home_light.png")), size=(20, 20))
    chat_image = customtkinter.CTkImage(light_image=Image.open(os.path.join("test_images\\chat_dark.png")),
                                                 dark_image=Image.open(os.path.join("test_images\\chat_light.png")), size=(20, 20))
    add_user_image = customtkinter.CTkImage(light_image=Image.open(os.path.join("test_images\\add_user_dark.png")),
                                                     dark_image=Image.open(os.path.join("test_images\\add_user_light.png")), size=(20, 20))
    
    home_button = customtkinter.CTkButton(navigation_frame, corner_radius=0, height=40, border_spacing=10, text="Home",
                                                   fg_color="transparent", text_color=("gray10", "gray90"), hover_color=("gray70", "gray30"),
                                                   image=home_image, anchor="w", command=home_button_event)
    home_button.grid(row=1, column=0, sticky="ew")

    frame_2_button = customtkinter.CTkButton(navigation_frame, corner_radius=0, height=40, border_spacing=10, text="Chat",
                                                      fg_color="transparent", text_color=("gray10", "gray90"), hover_color=("gray70", "gray30"),
                                                      image=chat_image, anchor="w", command=chat_button_event)
    frame_2_button.grid(row=2, column=0, sticky="ew")

    frame_3_button = customtkinter.CTkButton(second_frame, corner_radius=0, height=40, border_spacing=10, text=last_name_maneger+" "+name_maneger,
                                                      fg_color="transparent", text_color=("gray10", "gray90"), hover_color=("gray70", "gray30"),
                                                      image=chat_image, anchor="w", command=maneger_button_event)
    frame_3_button.grid(row=1, column=0, sticky="ew")
    ######
    
    width = 1099
    height = 600
    # create third fram
    
    
    
   

    def show():
        roote.destroy()
        showsleep()
        roote()
    # create sidebar frame with widgets
    def sidebar():
        c.execute('INSERT INTO '+profileemail+profikelastname+profikelastname+' VALUES (:user)',{
        "user":("open (غیاب و حضور) :"+(str(x))),
        })
        conn.commit()
        roote.destroy()
        presant()
    
    def add():
        import sqlite3
        conn = sqlite3.connect("user1.db")
        c = conn.cursor()
        c.execute('INSERT INTO '+profileemail+profikelastname+profikelastname+' VALUES (:user)',{
        "user":("open کلاس کردن اضاقه :"+(str(x))),
        })
        conn.commit()
        import os
        from tkinter import Tk
        from tkinter.filedialog import askopenfilename

        Tk().withdraw() # ما نمی‌خواهیم یک پنجره کامل GUI داشته باشیم
        filename = askopenfilename() # نمایش دیالوگ انتخاب فایل و بازگشت مسیر فایل انتخاب شده
        import sqlite3
        conn = sqlite3.connect("user1.db")
        c = conn.cursor()
        c.execute("SELECT * FROM user_profile_techer")
        results = c.fetchall()   
        if results:
            for i in results:
                name = i[0]
                last_name = i[1]
        desktop_path = os.path.join(os.path.join(os.environ['USERPROFILE']), "Desktop")
        folder_name = os.path.join(desktop_path,"معلم_"+name+"_"+last_name)
        if filename:
            destination_folder = folder_name
            if not os.path.exists(destination_folder):
                os.makedirs(destination_folder) # ایجاد پوشه اگر وجود ندارد
            destination_path = os.path.join(destination_folder, os.path.basename(filename))
            os.rename(filename, destination_path) # انتقال فایل
            print(f'فایل به مسیر {destination_path} منتقل شد.')
        else:
            print('هیچ فایلی انتخاب نشده است.')




    ######################################
    sidebar_frame = customtkinter.CTkFrame(root1, width=140,fg_color=("#FFFFFF","#2E2E2E"), corner_radius=0)
    root1.grid(row=0, column=0, rowspan=4, sticky="nsew")
    sidebar_frame.grid(row=0, column=1, rowspan=4, sticky="nsew")
    sidebar_frame.grid_rowconfigure(7, weight=1)
    if profilename1 == "200":
        logo_label = customtkinter.CTkLabel(sidebar_frame, text="روبوسان", font=customtkinter.CTkFont(size=20, weight="bold",family="B Roya"))
    else:
        logo_label = customtkinter.CTkLabel(sidebar_frame, text="RoBoSaN", font=customtkinter.CTkFont(size=20, weight="bold"))
    logo_label.grid(row=0, column=0, padx=20, pady=(20, 10))
    sidebar_button_3 = customtkinter.CTkButton(sidebar_frame,text="حضور و غیاب",font=customtkinter.CTkFont("B Homa",15),command=sidebar)
    sidebar_button_3.grid(row=1, column=0, padx=20, pady=10)

    sidebar_button_4 = customtkinter.CTkButton(sidebar_frame,text="افزودن کلاس",font=customtkinter.CTkFont("B Homa",15),command=add)
    sidebar_button_4.grid(row=2, column=0, padx=20, pady=10)
    if profilename1 == "200":
        tipe = customtkinter.CTkOptionMenu(sidebar_frame, values=["ما درباره",  "BMI", "کد مورس", "پلیر موزیک", "شمار ثانیه", "دوز بازی", "متن ادیت", "انس گرم پند محاسبه", "کد ار یو کی", "تلفن دفترچه", "من روزانه برنامه"],font=customtkinter.CTkFont("B Homa",15),
    
                                                           command=robo2) 
    tipe.grid(row=3, column=0, padx=20, pady=(10))
    if profilename1 == "200":
        tipe.set("منو")    
    else:
        tipe.set("Menu")
    #################
    def news():
     try:
        import requests
        from bs4 import BeautifulSoup
        from PIL import Image
        from io import BytesIO
        import os

        #    ایجاد پوشه‌ی "news" اگر وجود ندارد
        if not os.path.exists('news'):
            os.makedirs('news')

        # آدرس وب‌سایتی که می‌خواهید از آن تصاویر را دانلود کنید
        url = 'https://zamandaily.ir/'

        # دریافت محتوای صفحه وب
        response = requests.get(url)
        soup = BeautifulSoup(response.text, 'html.parser')

        # پیدا کردن تمام تگ‌های img در صفحه وب
        images = soup.find_all('img')

        # تعیین محدوده‌ی دانلود تصاویر
        start_index = 5 # شروع دانلود از تصویر پنجم
        end_index = 10 # پایان دانلود در تصویر دهم

        # دانلود و ذخیره‌ی تصاویر در پوشه‌ی "news"
        for index, img in enumerate(images[start_index:end_index], start=start_index):
            img_url = img['src']
            img_data = requests.get(img_url).content
            img_name = f'news/image_{index}.jpg'
    
            with open(img_name, 'wb') as file:
                file.write(img_data)
                print(f'تصویر {img_name} دانلود و در پوشه‌ی "news" ذخیره شد.')

            # نمایش تصویر
            image = Image.open(BytesIO(img_data))
            image.show()
        messagebox.showinfo("stop","صفحات روزنامه تمام شد")
        print('دانلود و نمایش تصاویر به پایان رسید.')
     except:
        messagebox.showerror("not internet","اینترنت قطع است لطف اینترنت رو چک کنید")
         
    def pdf():

        image = Image.open("pdf\\Capture.png")
        image.show()
        time.sleep(5)
        image = Image.open("pdf\\Capture1.png")
        image.show()
        time.sleep(5)
        image = Image.open("pdf\\Capture2.png")
        image.show()
        time.sleep(5)
        image = Image.open("pdf\\Capture3.png")
        image.show()
        time.sleep(5)
        image = Image.open("pdf\\Capture4.png")
        image.show()
        time.sleep(5)


    
    
    sidebar_button_5 = customtkinter.CTkButton(sidebar_frame,text="اخبار فوری",font=customtkinter.CTkFont("B Homa",15),command=news)
    sidebar_button_5.grid(row=4, column=0, padx=20, pady=10)

    sidebar_button_5 = customtkinter.CTkButton(sidebar_frame,text="توضیحات",font=customtkinter.CTkFont("B Homa",15),command=pdf)
    sidebar_button_5.grid(row=5, column=0, padx=20, pady=10)
    if profilename1 == "200":  
        sidebar_button_6 = customtkinter.CTkButton(sidebar_frame,text="خروج از حساب کاربری",font=customtkinter.CTkFont("B Homa",15), command=backe)
        sidebar_button_6.grid(row=6, column=0, padx=10, pady=0)
    else:
        sidebar_button_3 = customtkinter.CTkButton(sidebar_frame,text="Log Out", command=backe)
        sidebar_button_3.grid(row=3, column=0, padx=10, pady=0)
        sidebar_button_10 = customtkinter.CTkButton(sidebar_frame,text="Show Sleep", command=show)
        sidebar_button_10.grid(row=4, column=0, padx=10, pady=0)
    
    
    def dark():
        if switch.get() == "on":
            customtkinter.set_appearance_mode("Dark")
        else:
            customtkinter.set_appearance_mode("Light")
    if profilename1 == "200":
        sidebar_button_3 = customtkinter.CTkButton(sidebar_frame,text="خروج",font=customtkinter.CTkFont("B Homa",15), command=backe)
    else:
        sidebar_button_3 = customtkinter.CTkButton(sidebar_frame,text="Log Out", command=backe)
    if profilename1 == "200":
        switch = customtkinter.CTkSwitch(master=sidebar_frame, text="تاریک",command=dark,font=customtkinter.CTkFont("B Homa",15), onvalue="on", offvalue="off")
    else:
        switch = customtkinter.CTkSwitch(master=sidebar_frame, text="Sleep",command=dark, onvalue="on", offvalue="off")
    switch.grid(row=7, column=0, padx=1, pady=1)
    if profilename1 == "200":
        scaling_label = customtkinter.CTkLabel(sidebar_frame, text="اندازه متن :",font=customtkinter.CTkFont("B Homa",15), anchor="w")
    else:
        scaling_label = customtkinter.CTkLabel(sidebar_frame, text="UI Scaling:", anchor="w")
    scaling_label.grid(row=8, column=0, padx=20, pady=(1))
    scaling_optionemenu = customtkinter.CTkOptionMenu(sidebar_frame, values=["80%", "90%", "100%", "110%", "120%"],
                                                           command=change_scaling_event)
    scaling_optionemenu.grid(row=8, column=0, padx=20, pady=(10,20))
    # create main entry and button
    from tkintermapview import TkinterMapView
    # create textbox
    textbox = customtkinter.CTkTextbox(master=root,width=250)
    textbox.grid(row=0, column=2, padx=(20, 0), pady=(20, 0), sticky="nsew")

    
    
    def search_event(event=None):
            map_widget.set_address(entry32.get())
    mape = customtkinter.CTkFrame(root,corner_radius=0)
    mape.grid(row=1, column=2, padx=(20, 0), pady=(20, 0), sticky="nsew")
    map_widget = TkinterMapView(mape, corner_radius=0,width=353,height=220)
    map_widget.grid(row=1, column=0, sticky="nswe", padx=20, pady=(10, 10))
    if profilename1 == "200":
        entry32 = customtkinter.CTkEntry(mape, placeholder_text="..... نام مکان مورد نظرت را وارد کن",justify=RIGHT,font=customtkinter.CTkFont("B Homa",13))
    else:
        entry32 = customtkinter.CTkEntry(mape, placeholder_text="Enter Contry or ...")
    entry32.grid(row=0, column=0,  padx=1, pady=1, sticky="nsew")
    entry32.bind("<Return>", search_event)
    
    
    
    # create tabview
    tabview = customtkinter.CTkTabview(root, width=320)
    tabview.grid(row=0, column=3, padx=(20, 0), pady=(20, 0), sticky="nsew")
    if profilename1 == "200":
        tabview.add("ترجمگر")
        tabview.add("هوا و آب")
        tabview.add("بات چت")
        tabview.add("حساب ماشین")
        tabview.tab("ترجمگر").grid_columnconfigure(0, weight=1) # configure grid of individual tabs
        tabview.tab("هوا و آب").grid_columnconfigure(0, weight=1)
        tabview.tab("بات چت").grid_columnconfigure(0, weight=1)
        tabview.tab("حساب ماشین").grid_columnconfigure(0, weight=1)
    else:
        tabview.add("Transletor")
        tabview.add("Weather")
        tabview.add("ChatBot")
        tabview.add("calculet")
        tabview.tab("Transletor").grid_columnconfigure(0, weight=1) # configure grid of individual tabs
        tabview.tab("Weather").grid_columnconfigure(0, weight=1)
        tabview.tab("ChatBot").grid_columnconfigure(0, weight=1)
        tabview.tab("calculet").grid_columnconfigure(0, weight=1)
    
    try:
     try:
        api_key="1ceaad7d14b2cb4f9403429d6e09a147"
        base_url="https://api.openweathermap.org/data/2.5/weather?q="
        url=base_url + "Tehran" + "&appid=" + api_key
        #o=(url)
    
        result = requests.get(url)
        weather = result.json()
     except requests.exceptions.ConnectionError:
        messagebox.showwarning('خطا',"در فرایند وصل شدن دچار مشکل شدیم از وصل بودن اینترنت خود مطمعن شوید و اگر آفلاین باشیم نمیتوانید از برنامه های آنلاین استفاده کنید")
     except:
        messagebox.showerror('Error',"Some Errored Occured\nTry again Later!")
    # print(weather)
     if weather['cod']=='404' and weather['message']=='city not found':
        messagebox.showerror("Error","Entered City Not Found")
     elif weather['cod']=='400' and weather['message']=='Nothing to geocode':
        messagebox.showinfo("Warning",'Enter The city name')
     else:
        # getting time according to timezone
        lon=weather['coord']['lon']  # longitutde
        lat=weather['coord']['lat']  # latitude
        type=weather['weather'][0]['main']
        
        # sets the temperature and degree label
        temp=int(weather['main']['temp']-273)
        if type=="Clear":
            img="clear.png"
        elif type=="Clouds":
            img='clouds.png'
        elif type=="Rain":
            img='rain.png'
        elif type=='Haze':
            img='haze.png'
        else:
            img='main.png'
     if profilename1 == "200":
      from PIL import Image,ImageTk
      feel=customtkinter.CTkLabel(tabview.tab("هوا و آب"),text=f"Feels Like {int(weather['main']['feels_like']-273)}° | {weather['weather'][0]['main']}", anchor="w")
      feel.place(x=100,y=55)
      temperature=customtkinter.CTkLabel(tabview.tab("هوا و آب"),text=int(weather['main']['temp']-273), anchor="w",font=customtkinter.CTkFont(size=23))
      temperature.place(x=94,y=25)
      degree=customtkinter.CTkLabel(tabview.tab("هوا و آب"),text="°C",font=customtkinter.CTkFont(size=15), anchor="w")
      degree.place(x=120,y=22)
      location=customtkinter.CTkLabel(tabview.tab("هوا و آب"),text=weather['name'], anchor="w")
      location.place(x=100,y=75)
      img3=Image.open(f"Icons/{img}")
      resizeimg3=img3.resize((80,80))
      finalimg3=ImageTk.PhotoImage(resizeimg3)
      icons = customtkinter.CTkLabel(tabview.tab("هوا و آب"),image=finalimg3,text="", anchor="w")
      icons.place(x=10,y=20)
     else:
      
      feel=customtkinter.CTkLabel(tabview.tab("Weather"),text=f"Feels Like {int(weather['main']['feels_like']-273)}° | {weather['weather'][0]['main']}", anchor="w")
      feel.place(x=100,y=55)
      temperature=customtkinter.CTkLabel(tabview.tab("Weather"),text=int(weather['main']['temp']-273), anchor="w",font=customtkinter.CTkFont(size=23))
      temperature.place(x=94,y=25)
      degree=customtkinter.CTkLabel(tabview.tab("Weather"),text="°C",font=customtkinter.CTkFont(size=15), anchor="w")
      degree.place(x=120,y=22)
      location=customtkinter.CTkLabel(tabview.tab("Weather"),text=weather['name'], anchor="w")
      location.place(x=100,y=75)
      img3=Image.open(f"Icons/{img}")
      resizeimg3=img3.resize((80,80))
      finalimg3=ImageTk.PhotoImage(resizeimg3)
      icons = customtkinter.CTkLabel(tabview.tab("Weather"),image=finalimg3,text="", anchor="w")
      icons.place(x=10,y=20)
    except:
        print("problem")
    #######################
    
        #######################

    if profilename1 == "200":
        entry1 = customtkinter.CTkEntry(tabview.tab("ترجمگر"), placeholder_text=".... متن",justify=RIGHT,font=customtkinter.CTkFont("B Homa",13))
        entry1.grid(row=0, column=0,  padx=20, pady=(1, 1), sticky="nsew")
    else:
        entry1 = customtkinter.CTkEntry(tabview.tab("Transletor"), placeholder_text="Text ...")
        entry1.grid(row=0, column=0,  padx=20, pady=(1, 1), sticky="nsew")
    entry1.bind("<Return>", open_input_dialog_event)
    from deep_translator import GoogleTranslator
    from googletrans import LANGUAGES
    values = list(LANGUAGES.values())
    if profilename1 == "200":
        combobox_1 = customtkinter.CTkComboBox(tabview.tab("ترجمگر"), values=values)
    else:
        combobox_1 = customtkinter.CTkComboBox(tabview.tab("Transletor"), values=values)
    combobox_1.grid(row=1, column=0, padx=20, pady=(10, 10))
    search_keyword = tk.StringVar()
    def createButton():
                            conn = sqlite3.connect('user1.db')
                            c = conn.cursor()
                            c.execute('INSERT INTO '+profileemail+profikelastname+profikelastname+' VALUES (:user)',{
                                      "user":(" open calculetor :"+(str(x))),
                                      })
                    
                            '''
                            DOCSTRING: Method that creates the buttons
                            INPUT: nothing
                            OUTPUT: creates a button
                            '''
                            
                            #We first create each button one by one with the value we want
                            #Using addButton() method which is described below
                            b0 = addButton(0)
                            b1 = addButton(1)
                            b2 = addButton(2)
                            b3 = addButton(3)
                            b4 = addButton(4)
                            b5 = addButton(5)
                            b6 = addButton(6)
                            b7 = addButton(7)
                            b8 = addButton(8)
                            b9 =  addButton(9)
                            b_add = addButton('+')
                            b_sub = addButton('-')
                            b_mult = addButton("×")
                            b_div = addButton('÷')
                            b_clear = addButton('c')
                            b_equal = addButton('=')
                            b_tavan = addButton("^")
                            b_baghi = addButton("mod")
                            b_edame = addButton("div")
                            b_sqrt = addButton("r")
                            import math
                            
    
                            
                            
                    
                            #Arrange the buttons into lists which represent calculator rows
                            row1=[b7,b8,b9,b_add]
                            row2=[b4,b5,b6,b_sub]
                            row3=[b1,b2,b3,b_mult]
                            row4=[b_clear,b0,b_equal,b_div]
                            row5=[b_tavan,b_baghi,b_edame,b_sqrt]
                    
                            #Assign each button to a particular location on the GUI
                            r=1
                            for row in [row1, row2, row3, row4, row5]:
                                c=0
                                for buttn in row:
                                    buttn.grid(row=r, column=c, columnspan=1)
                                    c+=1
                                r+=1
                    
                    
                    
                    
    def addButton(value):
                    
                                '''
                                DOCSTRING: Method to process the creation of a button and make it clickable
                                INPUT: value of the button (1,2,3,4,5,6,7,8,9,0,+,-,*,/,c,=,**)
                                OUTPUT: returns a designed button object
                                '''
                                return Button(master, text=value, width=9, command = lambda: clickButton(str(value)))
                        
                    
                    
                    
    def clickButton( value):
                            
                            '''
                            DOCSTRING: Method to program the actions that will happen in the calculator after a click of each button
                            INPUT: value of the button (1,2,3,4,5,6,7,8,9,0,+,-,*,/,c,=)
                            OUTPUT: what action will be performed when a particular button is clicked
                            '''
                            
                            #Get the equation that's entered by the user
                            current_equation=str(equation.get())
                            
                            #If user clicked "c", then clear the screen
                            if value == 'c':
                                equation.delete(-1, END)
                            
                            #If user clicked "=", then compute the answer and display it
                            elif value == '=':
                                answer = str(eval(current_equation))
                                equation.delete(-1, END)
                                equation.insert(0, answer)
                                c.execute('INSERT INTO '+profileemail+profikelastname+profikelastname+' VALUES (:user)',{
                    "user":(" calculetor:"+equation.get()+"="+answer+" :"+(str(x))),
                    })
                            elif value == "back":
                                root.destroy()
                            elif value == "^":
                                equation.insert((clickButton(str("**"))))
                            elif value == "÷":
                                equation.insert((clickButton(str("/"))))
                            elif value == "×":
                                equation.insert((clickButton(str("*"))))
                            elif value == "mod":
                                equation.insert((clickButton(str("%"))))
                            elif value == "div":
                                equation.insert((clickButton(str("//"))))
                            
                            
                                
                            
                            #If user clicked any other button, then add it to the equation line
                            else:
                                equation.delete(0, END)
                                equation.insert(-1, current_equation+value)
                            if value == "r":
                                import math
                                mn10 =math.sqrt(int(eval(current_equation)))
                                answer = str(mn10)
                                equation.delete(-1, END)
                                equation.insert(0, answer)
    if profilename1 == "200":
        master = tabview.tab("حساب ماشین")
    else:
        master = tabview.tab("calculet")               
    #Add a name to our application
    
    #Create a line where we display the equation
    equation=Entry(master, width=36, borderwidth=5)
    #Assign a position for the equation line in the grey application window
    equation.grid(row=0, column=0, columnspan=4, padx=10, pady=10)
    #Execute the .creteButton() method
    createButton()
    def filter_values(*args):
        keyword = search_keyword.get().lower()
        filtered_values = [value for value in values if keyword in value.lower()]
        combobox_1['values'] = filtered_values

    search_keyword.trace('w', filter_values)
    if profilename1 == "200":
        string_input_button = customtkinter.CTkButton(tabview.tab("ترجمگر"), text="ترجمه",font=customtkinter.CTkFont("B Homa",13), command=open_input_dialog_event)
    else:
        string_input_button = customtkinter.CTkButton(tabview.tab("Transletor"), text="Transletor", command=open_input_dialog_event)
    string_input_button.grid(row=2, column=0, padx=20, pady=(10, 10))
    
    if profilename1 == "200":
        combobox_12 = customtkinter.CTkComboBox(tabview.tab("بات چت"),
                                                        values=[ "Wolfram", "Wikipedia"])
    else:
        combobox_12 = customtkinter.CTkComboBox(tabview.tab("ChatBot"),
                                                        values=["Chatgpt", "Wolfram", "Wikipedia"])
    combobox_12.grid(row=1, column=0, padx=20, pady=(10, 10))
    if profilename1 == "200":
        entry2 = customtkinter.CTkEntry(tabview.tab("بات چت"), placeholder_text=".... متن",justify=RIGHT,font=customtkinter.CTkFont("B Homa",13))
    else:
        entry2 = customtkinter.CTkEntry(tabview.tab("ChatBot"), placeholder_text="Text ...")
    entry2.grid(row=0, column=0,  padx=20, pady=(1, 1), sticky="nsew")
    entry2.bind("<Return>", open_input_dialog_event1)
    if profilename1 == "200":
        string_input_button1 = customtkinter.CTkButton(tabview.tab("بات چت"), text="جستجو",font=customtkinter.CTkFont("B Homa",13),
                                                       command=open_input_dialog_event1)
    else:
        string_input_button1 = customtkinter.CTkButton(tabview.tab("ChatBot"), text="Search",
                                                       command=open_input_dialog_event1)
    string_input_button1.grid(row=2, column=0, padx=20, pady=(10,10))
    
    
    tabview1 = customtkinter.CTkScrollableFrame(root,label_text="غابین اسامی",width=275)
    tabview1.grid(row=0, column=4, padx=(20, 0), pady=(20, 0), sticky="nsew")
    c.execute("SELECT * FROM user_profile_techer")
    results = c.fetchall()
    import pandas as pd
    if results:
                for i in results:
                    name = i[0]
                    last_name = i[1]
    folder_name = "معلم_"+name+"_"+last_name
    desktop_path = os.path.join(os.path.join(os.environ['USERPROFILE']), "Desktop")
    c.execute("SELECT * FROM Languge")
    results = c.fetchall()
    if results:
        for i in results:
            eeeeee = i[0]
        conn.commit()
    file_path=desktop_path+"\\"+folder_name+"\\"+eeeeee+".xlsx"
    df = pd.read_excel(file_path)

    # تابع برای پیدا کردن وضعیت حضور و غیاب بر اساس تاریخ
    def get_attendance_status(date):
        if date in df.columns:
            attendance = df[['نام', 'نام خانوادگی',date]]
            present_students = attendance[attendance[date] == 'حاضر']
            absent_students = attendance[attendance[date] == 'غایب']
            print (present_students[['نام', 'نام خانوادگی']].to_string(index=False))
            if absent_students[['نام', 'نام خانوادگی']].to_string(index=False) in """Empty DataFrame
Columns: [نام, نام خانوادگی]
Index: []""":
                customtkinter.CTkLabel(tabview1,text="غایب نداریم",justify=RIGHT,font=customtkinter.CTkFont("B Homa",20)).pack()
            else:
                customtkinter.CTkLabel(tabview1,text=absent_students[['نام', 'نام خانوادگی']].to_string(index=False),justify=RIGHT,font=customtkinter.CTkFont("B Homa",20)).pack()
    from persiantools.jdatetime import JalaliDate
    from datetime import datetime
    today=datetime.today()
    today_date = str(JalaliDate(today))
    date =  today_date
     # تاریخ مورد نظر خود را وارد کنید
    get_attendance_status(date)
    
    

    # لیست تصاویر برای گالری
    image_paths = ["lomge/[11].png", "lomge/[12].png", "lomge/[13].png", "lomge/[14].png"]

    # زمان تغییر تصاویر بین صفحات (به میلی ثانیه)
    interval = 2000

    # ایجاد پنجره اصلی Tkinter
    
    # ایجاد گالری
    
    # Add Button and Label
    if profilename1 == "200":
        weather11=tabview.tab("هوا و آب")
    else:
        weather11=tabview.tab("Weather")
    
    
 
    
    
    
    # Create the notebook to hold the notes gallery
    
    # create slider and progressbar frame
    
    
    # create scrollable frame
    if profilename1 == "200":
        scrollable_frame = customtkinter.CTkScrollableFrame(root, label_text="کاربر مشخصات",corner_radius=0)
    else:
        scrollable_frame = customtkinter.CTkScrollableFrame(root, label_text="Profile",corner_radius=0)
    scrollable_frame.grid(row=1, column=3, padx=20, pady=(10, 10), sticky="nsew")
    scrollable_frame.grid_columnconfigure(0, weight=1)
    scrollable_frame_switches = []
    import sqlite3
    conn = sqlite3.connect("user1.db")
    c = conn.cursor()
    
    c.execute("SELECT * FROM user_profile_techer")
    results = c.fetchall()
        
    if results:
            for i in results:
                profilename = i[0]
                profikelastname = i[1]
                profileemail = i[2]
                profiletime = i[3]
                profilenumber = i[4]
                profileborn = i[5]
                profilemr = i[6]
    if profilename1 == "200":
        name = customtkinter.CTkLabel(master=scrollable_frame, text="نام :"+profilename,font=customtkinter.CTkFont("B Homa",13))
    else:
        name = customtkinter.CTkLabel(master=scrollable_frame, text="Name: "+profilemr+"."+profilename,font=customtkinter.CTkFont("B Homa",13))
    name.grid(row=0, column=0,  padx=10, pady=5, sticky="")
    if profilename1 == "200":
        last_name = customtkinter.CTkLabel(master=scrollable_frame, text="نام خانوادگی :"+profikelastname,font=customtkinter.CTkFont("B Homa",13))
    else:
        last_name = customtkinter.CTkLabel(master=scrollable_frame, text="Last Name: "+profikelastname,font=customtkinter.CTkFont("B Homa",13))
    last_name.grid(row=1, column=0,  padx=10, pady=5, sticky="")
    if profilename1 == "200":
        email = customtkinter.CTkLabel(master=scrollable_frame, text="نام کاربری :"+profileemail,font=customtkinter.CTkFont("B Homa",13))
    else:
        email = customtkinter.CTkLabel(master=scrollable_frame, text="Email: "+profileemail)
    email.grid(row=2, column=0,  padx=10, pady=5, sticky="")
    if profilename1 == "200":
        number = customtkinter.CTkLabel(master=scrollable_frame, text="شماره تلفن :"+profilenumber,font=customtkinter.CTkFont("B Homa",13))
    else:
        number = customtkinter.CTkLabel(master=scrollable_frame, text="Phone number: "+profilenumber)
    number.grid(row=3, column=0,  padx=10, pady=5, sticky="")
    if profilename1 == "200":
        timee = customtkinter.CTkLabel(master=scrollable_frame, text="تاریخ ثبت نام :"+profiletime,font=customtkinter.CTkFont("B Homa",13))
    else:
        timee = customtkinter.CTkLabel(master=scrollable_frame, text="Time: "+profiletime)
    timee.grid(row=4, column=0,  padx=10, pady=5, sticky="")
    if profilename1 == "200":
        age = customtkinter.CTkLabel(master=scrollable_frame, text="درس آموزشی :"+profileborn,font=customtkinter.CTkFont("B Homa",13))
    else:
        age = customtkinter.CTkLabel(master=scrollable_frame, text="Born: "+profileborn)
    age.grid(row=5, column=0,  padx=10, pady=5, sticky="")
    
    
    
    # create checkbox and switch frame
    if profilename1 == "200":
        checkbox_slider_frame = customtkinter.CTkScrollableFrame(root, label_text="برنامه",corner_radius=0)
    else:
        checkbox_slider_frame = customtkinter.CTkScrollableFrame(root, label_text="App",corner_radius=0)
    checkbox_slider_frame.grid(row=1, column=4, padx=20, pady=(10, 10), sticky="nsew")
    if profilename1 == "200":
        sidebar_button_33 = customtkinter.CTkButton(checkbox_slider_frame, text="نقشه جهان",command=map,font=customtkinter.CTkFont("B Homa",13))
    else:
        sidebar_button_33 = customtkinter.CTkButton(checkbox_slider_frame, text="World Map",command=map,font=customtkinter.CTkFont("B Homa",13))
    sidebar_button_33.grid(row=1, column=0, padx=20, pady=(10, 10))
    if profilename1 == "200":
        sidebar_button_333 = customtkinter.CTkButton(checkbox_slider_frame, text="تاریخ عمل کردم",command=history,font=customtkinter.CTkFont("B Homa",13))
    else:
        sidebar_button_333 = customtkinter.CTkButton(checkbox_slider_frame, text="History",command=history)
    sidebar_button_333.grid(row=2, column=0, padx=20, pady=(10, 10))
    if profilename1 == "200":
        sidebar_button_333 = customtkinter.CTkButton(checkbox_slider_frame, text="بروز شدن برنامه",command=Reset,font=customtkinter.CTkFont("B Homa",13))
    else:
        sidebar_button_333 = customtkinter.CTkButton(checkbox_slider_frame, text="Reset",command=Reset)
    sidebar_button_333.grid(row=4, column=0, padx=20, pady=(10, 10))
    
    def change_map( new_map: str):
        if new_map == "OpenStreetMap":
            map_widget.set_tile_server("https://a.tile.openstreetmap.org/{z}/{x}/{y}.png")
        elif new_map == "Google normal":
            map_widget.set_tile_server("https://mt0.google.com/vt/lyrs=m&hl=en&x={x}&y={y}&z={z}&s=Ga", max_zoom=22)
        elif new_map == "Google satellite":
            map_widget.set_tile_server("https://mt0.google.com/vt/lyrs=s&hl=en&x={x}&y={y}&z={z}&s=Ga", max_zoom=22)
    map_option_menu = customtkinter.CTkOptionMenu(checkbox_slider_frame, values=["OpenStreetMap", "Google normal", "Google satellite"],
                                                                   command=change_map)
    map_option_menu.grid(row=3, column=0, padx=20, pady=10)
    
    
    
    
    
    
    
    
    
    # set default values
    sidebar_button_3.configure(state="disabled", text="Disabled CTkButton")
    
    
    
    
    scaling_optionemenu.set("100%")
    combobox_1.set("Laguege")
    import random
    text="سلام به نام خدا من ساسان هستم"
    from time import sleep
    words = text.split()
    reversed_text = ' '.join(words[::-1])
    texts="""⠀⠀⠀⠀⠀⠀⠀⠀⣀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⠀⢹⣿⣭⣑⠒⠦⠤⢀⣀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⢀⣀⡸⠿⠿⣿⣿⣷⣦⣤⣄⣈⡉⠛⠲⢤⣀⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠲⢾⣿⣿⣶⣶⣿⣿⣿⣿⣿⣿⣿⣿⣿⣷⣶⣦⠈⢷⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠈⠙⣿⣿⣿⣿⣿⣿⣿⣿⣿⠛⠿⣿⠋⢿⣇⡈⡇⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⢀⣾⣿⣿⣿⣿⣿⣿⣿⢹⡟⠀⠀⠀⠀⠘⣿⣿⢹⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⢀⣾⣿⣿⣿⣿⣿⣿⣿⣿⠈⠧⣀⠀⠀⡠⠂⠃⣿⡞⡄⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⢀⣾⣿⡟⢿⣿⣿⣿⡿⠘⠿⠃⠀⠀⠈⠙⠛⢿⣷⡃⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⢸⣿⢿⣧⠸⣿⣿⣿⡇⠀⠀⠀⠀⠀⠀⠀⠀⢹⡇⠛⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⠀⢸⣿⣷⣿⣿⣿⡇⠀⠀⠀⠀⠀⠀⠀⢀⣼⡇⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⠀⠘⠟⠁⢿⣿⣿⣧⠄⠀⡀⢀⡰⠖⠒⢿⣿⡇⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⣀⡀⠀⠀⠀⠀⢸⣁⡽⢿⣀⣐⠡⠂⠱⡄⠀⠀⠙⠇⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⣮⣏⡅⠀⠀⢠⡇⠀⠀⠀⠟⠁⠀⠀⣰⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠏⠶⡄⠀⢀⣀⡟⣦⣀⢸⣀⣀⣠⣾⣯⠕⠀⠀⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⣧⢲⠇⢸⡇⣚⣿⣿⣿⣯⣿⣿⣶⣶⡾⡀⠀⠀⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⡶⠺⡃⠀⠱⣻⣿⣿⣿⣿⣿⣿⣿⣿⣿⢹⠀⠀⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⣶⠭⠅⠀⠻⢻⢼⣿⣿⣿⣿⣿⣿⣿⣿⡘⠀⠀⠀⠀⠀⠀⠀⠀⠀
        ⣏⣚⠁⠀⠀⣸⣼⣿⣿⠟⠉⠻⣿⣿⡿⡆⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⣯⢭⡇⠀⠀⢸⣻⣿⠁⠀⠀⠀⠈⢻⣿⠇SASUKE⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠁⠀⠀⠀⠀⠁⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
"""
    texts1="""
⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠤⣄⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⡆⠀⠀⠀⠀
⠀⠀⠀⠀⠀⢀⣤⣴⣶⣶⣶⣶⣶⣦⣤⡄⠊⠀⠀⠀⠀⠀
⠀⠦⣤⣶⣾⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣷⣄⠀⠀⠀⠀
⠀⠀⣼⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣷⡀⠀⠀
⠀⣸⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣧⠀⠀
⣠⣿⣿⣿⡿⢻⣿⣿⣿⣿⣿⣿⢿⣿⣿⣿⣿⣿⣿⣿⡧⠀
⠉⢹⣿⣿⢓⣺⡿⠟⠛⠛⢻⣿⣼⣿⣿⣿⣿⣿⣿⣿⣇⠀
⠀⢸⣿⣿⠉⠀⠀⠀⠀⠀⠀⠀⠀⢀⣿⡿⠿⢿⣿⣿⣿⠀
⡀⠈⣿⣿⡀⠀⠀⠀⠄⠀⠀⠀⠀⢸⣿⣗⠏⢺⣿⣿⣿⡇
⠀⢀⡈⣿⣷⣤⣄⣀⠀⠀⠀⢀⣤⣿⣿⣶⣾⣿⣿⣿⣿⣧
⠀⠘⠉⢹⣿⣿⣿⣿⣿⡇⠀⢸⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿
⠀⠀⠀⣾⣿⣿⠟⡿⠟⡁⠄⠚⠉⠀⠘⢿⣿⣿⣿⣿⣿⡇
⠀⠀⠀⣿⠟⠁⠈⠉⠁⠀⠀⠀⠀⠀⠀⠀⠙⡿⠿⠏⠿⠃
⠀⠀⠀⡜⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠘⡄⠀⠀⠀"""
    texts2="""
⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀  ⣤⣶⣶⣶⣶⣶⣦⣄⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⢰⣿⣿⣿⣿⣿⣿⣿⣿⣿⣷⡄⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⣠⢿⣿⣿⡿⣿⣿⣿⣿⣿⣿⣿⣿⣀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⣰⣿⣿⣿⣿⡇⣿⣷⣿⣿⣿⣿⣿⣿⣯⡄⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⡰⣿⣿⣿⣇⣿⣀⠸⡟⢹⣿⣿⣿⣿⣿⣿⣷⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⢀⢡⣿⣿⣿⡇⠝⠋⠀⠀⠀⢿⢿⣿⣿⣿⣿⣿⡇⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠸⢸⠸⣿⣿⣇⠀⠀⠀⠀⠀⠀⠊⣽⣿⣿⣿⠁⣷⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⢿⣿⣿⣷⣄⠀⠀⠀⢠⣴⣿⣿⣿⠋⣠⡏⡄⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠐⠾⣿⣟⡻⠉⠀⠀⠀⠈⢿⠋⣿⡿⠚⠋⠁⡁⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⢀⣴⣶⣾⣿⣿⡄⠀⣳⡶⡦⡀⣿⣿⣷⣶⣤⡾⠁⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⢸⣿⣿⣿⣿⣿⣿⡆⠀⡇⡿⠉⣺⣿⣿⣿⣿⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⢸⣿⣿⣿⣿⣿⣯⠽⢲⠇⠣⠐⠚⢻⣿⣿⣿⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⣿⣿⣿⣿⣿⡐⣾⡏⣷⠀⠀⣼⣷⡧⣿⣿⣦⣄⡀⠀⠀⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⣻⣿⣿⣿⣿⣿⣮⠳⣿⣇⢈⣿⠟⣬⣿⣿⣿⣿⣿⡦⢄⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⠀⠀⢀⢄⣾⣿⣿⣿⣿⣿⣿⣿⣷⣜⢿⣼⢏⣾⣿⣿⣿⢻⣿⣝⣿⣦⡑⢄⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⣠⣶⣷⣿⣿⠃⠘⣿⣿⣿⣿⣿⣿⣿⡷⣥⣿⣿⣿⣿⣿⠀⠹⣿⣿⣿⣷⡀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⣇⣤⣾⣿⣿⡿⠻⡏⠀⠀⠸⣿⣿⣿⣿⣿⣿⣮⣾⣿⣿⣿⣿⡇⠀⠀⠙⣿⣿⡿⡇⠀⠀⠀⠀⠀
⠀⠀⢀⡴⣫⣿⣿⣿⠋⠀⠀⡇⠀⠀⢰⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⡇⠀⠀⠀⢘⣿⣿⣟⢦⡸⠀⠀⠀
⠀⡰⠋⣴⣿⣟⣿⠃⠀⠀⠀⠈⠀⠀⣸⣿⣿⣿⣿⣿⣿⣇⣽⣿⣿⣿⣿⣇⠀⠀⠀⠁⠸⣿⢻⣦⠉⢆⠀⠀
⢠⠇⡔⣿⠏⠏⠙⠆⠀⠀⠀⠀⢀⣜⣛⡻⢿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⡿⡀⠀⠀⠀⠀⡇⡇⠹⣷⡈⡄⠀
⠀⡸⣴⡏⠀⠀⠀⠀⠀⠀⠀⢀⣾⣿⣿⣿⣿⣿⣿⣿⣿⣿⣻⣿⣿⣿⣿⣿⣿⡄⠀⠀⠀⡇⡇⠀⢻⡿⡇⠀
⠀⣿⣿⣆⠀⠀⠀⠀⠀⠀⢀⣼⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⡀⠀⣰⠿⠤⠒⡛⢹⣿⠄
⠀⣿⣷⡆⠁⠀⠀⠀⠀⢠⣿⣿⠟⠻⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⡟⠻⢷⡀⠀⠀⠀⠀⠀⣸⣿⠀
⠀⠈⠿⢿⣄⠀⠀⠀⢞⠌⡹⠁⠀⠀⢻⡇⠹⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⠁⢳⠀⠀⠁⠀⠀⠀⠀⢠⣿⡏⠀
⠀⠀⠀⠈⠂⠀⠀⠀⠈⣿⠁⠀⠀⠀⡇⠁⠀⠘⢿⣿⣿⠿⠟⠋⠛⠛⠛⠀⢸⠀⠀⡀⠂⠀⠀⠐⠛⠉⠀⠀
⠀⠀⠀⠀⠀⠀⠀⠀⠀⠐⠕⣠⡄⣰⡇⠀⠀⠀⢸⣧⠀⠀⠀⠀⠀⠀⠀⢀⣸⠠⡪⠊⠀⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠈⢫⣽⡋⠭⠶⠮⢽⣿⣆⠀⠀⠀⠀⢠⣿⣓⣽⠁⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⣿⣿⣿⣿⣿⣿⣿⣿⢹⣶⣦⣾⣿⣿⣿⡏⠀⠀⠀⠀⠀⠀⠀⠀⠀"""
    textbox.insert("0.0", texts)
    
   
    
    
    conn.commit()
    roote.mainloop()
#################################################################################
def BMI1():
    from PIL import Image,ImageTk
    import tkinter as tk
    from tkinter import ttk
    import os
    import PIL.Image
    
    
    root =Tk()
    root.title("BMI")
    root.iconbitmap(bitmap = 'Icons\icon.ico')
    root.geometry('470x650+300+200')
    root.configure(bg="#f0f1f5")
    root.resizable(False,False)
    ###################
    #################
    frame=root
    
    ##################
    top=PhotoImage(file='Images/top.png')
    top_img=Label(frame,image=top,bg='#f0f1f5')
    top_img.place(x=-10,y=-10)
    ###############
    #######
    box=PhotoImage(file='Images/box.png')
    Label(frame,image=box).place(x=20,y=100)
    Label(frame,image=box).place(x=240,y=100)
    #####
    scale=PhotoImage(file='Images/scale.png')
    Label(frame,image=scale).place(x=20,y=310)
    #############
    def get_current_value():
        return '{: .2f}'.format(current_value.get())
    def slider_changed(event):
        
        Height.set(get_current_value())
        
        size=int(float(get_current_value()))
        img = (PIL.Image.open("Images\\man.png"))
        resize_img = img.resize((50,10+size))
        photo2 = ImageTk.PhotoImage(resize_img)
        secondimag.config(image=photo2)
        secondimag.place(y=550-size)
        secondimag.image = photo2
    #############
    secondimag=Label(frame)
    secondimag.place(x=100,y=530)
    ##############
    stayl = ttk.Style()
    ############
    stayl.configure("TScale",background="white")
    
    current_value = tk.DoubleVar()
    slider= ttk.Scale(frame,from_=0, to=220,orient='horizontal',style="TScale",
                      command=slider_changed,variable=current_value)
    slider.place(x=90,y=250)
    
    #####
    def get_current_value2():
        return '{: .2f}'.format(current_value2.get())
    def slider_changed2(event):
        Weight.set(get_current_value2())
    #############
    stayl2 = ttk.Style()
    ############
    stayl2.configure("TScale",background="white")
    
    current_value2 = tk.DoubleVar()
    slider2= ttk.Scale(frame,from_=0, to=200,orient='horizontal',style="TScale",
                      command=slider_changed2,variable=current_value2)
    slider2.place(x=300,y=250)
    #####
    def bmi():
        h=float(Height.get())
        w=float(Weight.get())
        
        m=h/100
        bmi=round(float(w/m**2),1)
        label1.config(text=bmi)
        if bmi <18.5 :
            mb="underweight"
        elif 18.5<bmi<24.9 :
            mb="best"
        elif 25<bmi<29.9:
            mb="overweight"
        elif 30<bmi<39.9:
            mb="overweight"
        elif 39.9<bmi:
            mb="not human"
        elif 9>bmi:
            mb="not human"
        label2.configure(text=mb)
        
    #####
    
    Height=StringVar()
    Weight = StringVar()
    height=Entry(frame,textvariable=Height,width=5,font='arial 50',bg='#fff',fg='#000',bd=0,justify=CENTER)
    height.place(x=35,y=160)
    height.insert(0, 00.00)
    ##############
    weight=Entry(frame,textvariable=Weight,width=5,font='arial 50',bg='#fff',fg='#000',bd=0,justify=CENTER)
    weight.place(x=255,y=160)
    weight.insert(0, 00.00)
    ##########
    Button(frame,text="View BMI",width=15,height=2,font="aria 10 bold",bg="#1f6e68",fg="white",command=bmi).place(x=200,y=340)
    
    label1=Label(frame,font="arial 40 bold")
    label1.place(x=360,y=320)
    
    label2=Label(frame,font="arial 20 bold")
    label2.place(x=200,y=430)
    def back():
        root.destroy()
        
    label = Label(frame,text="Do want to back?",fg='black',bg='#f0f1f5',font=("B Homa",9))
    label.place(x=200,y=400)
    sing = Button(frame,width=6,text='Back',border=0,bg='#f0f1f5',cursor='hand2',fg='#57a1f8',command=back)
    sing.place(x=290,y=400)
    
    
    
    root.mainloop()    
#################################################################################
def presant():
    from openpyxl import load_workbook
    from persiantools.jdatetime import JalaliDate
    from functools import partial
    from tkinter import ttk, filedialog
    import pandas as pd
    import tkinter as tk
    import customtkinter as ctk
    root = Tk()
    root.title('حضور و غیاب')
    root.geometry('400x200')
    conn = sqlite3.connect('user1.db')
    c = conn.cursor()               
    profilename = "200"
    c.execute("SELECT * FROM Languge")
    results = c.fetchall()
    if results:
        for i in results:
            profilename1 = i[0]
        conn.commit()
    print(profilename1)
    def record_attendance():
            import sqlite3
            conn = sqlite3.connect("user1.db")
            c = conn.cursor()
            classe=profilename1
            c.execute("SELECT * FROM user_profile_techer")
            results = c.fetchall()
        
            if results:
                for i in results:
                    name = i[0]
                    last_name = i[1]
            folder_name = "معلم_"+name+"_"+last_name
            desktop_path = os.path.join(os.path.join(os.environ['USERPROFILE']), "Desktop")
            exfilename=desktop_path+"\\"+folder_name+"\\"+classe+".xlsx"
            print(exfilename)
            file_path = exfilename

            # بارگذاری ورک‌بوک
            import openpyxl
            workbook = openpyxl.load_workbook(file_path)
            workbook1 =workbook.active
            # فرض می‌کنیم که آخرین شیت، شیت مورد نظر ما است
            sheet = workbook.worksheets[-1]
            print(workbook1.max_column)
            # دریافت تاریخ از سطر اول و ستون اول آخرین جدول
            last_table_first_row_date = sheet.cell(row=1,column=workbook1.max_column).value
            print(last_table_first_row_date)

            # دریافت تاریخ امروز با استفاده از کتابخانه jalalidate
            from datetime import datetime
            today=datetime.today()
            today_date = str(JalaliDate(today))
            print(today_date)
            
            # مقایسه تاریخ‌ها
            if str(last_table_first_row_date) == today_date:
                # ایجاد یک پنجره پیام # پنهان کردن پنجره اصلی
                xx=messagebox.askyesno("wite","شما قبلا حضور و غیاب کردین مایلید دوباره انجام دهید؟")
                if xx == False:
                    root.destroy()
                else:
                    sheet.delete_cols(workbook1.max_column)  
                    workbook.save(file_path)
                    try:     
                            # حذف جدول قبلی
                            sheet.delete_cols(workbook1.max_column)  
                            

                            # we will use these two files for reading excel file
                            # multiple methods will use it so I put them here
                            wb = load_workbook(filename = exfilename)
                            ws=wb.active
                            root.destroy()
                            import csv
                            import tkinter
                            class StudentProject:
                                exfilename=desktop_path+"\\"+folder_name+"\\"+classe+".xlsx"
                                # we will use these two files for reading excel file
                                # multiple methods will use it so I put them here
                                wb = load_workbook(filename = exfilename)
                                ws=wb.active
                                # ls will present the all students in excel file
                                ls=list()
                                # mylist is the list which we will save to new coulumn
                                mylist=list()
                                # this is a counter for our list 
                                i=0
                                # here are some variables for user interface
                                root=Tk()
                                root.iconbitmap('Icons\icon.ico')
                                root.title("presence and absence")
                                mylabel=ttk.Label()
                                # this is empty label which will be use in grid
                                blabel=ttk.Label()
                                def __init__(self):
                                    # for start lets get student names 
                                    self.ls=self.getstudentnames()
                                    # and persian current date
                                    self.mylist.append(self.getdate())
                                   
                                # this methid will get all student names from excel 
                                # this method returns a list named ls 
                                def getstudentnames(self):
                                    ls=list()
                                    rownum=self.ws.max_row
                                    for i in range(2,rownum+1):
                                      ls.append(self.ws.cell(i,2).value)
                                    return ls
                                # this method showes gui 
                                # I tried to write some code but needs improve     
                                def showgui(self):
                                            from datetime import datetime
                                            today=datetime.today()
                                            import tkinter.messagebox as messagebox
                                            window_height = 250
                                            window_width = 600
                                            screen_width = self.root.winfo_screenwidth()
                                            screen_height = self.root.winfo_screenheight()
                                            x_cordinate = int((screen_width/2) - (window_width/2))
                                            y_cordinate = int((screen_height/2) - (window_height/2))
                                            self.root.geometry("{}x{}+{}+{}".format(window_width, window_height, x_cordinate, y_cordinate))
                                            self.mylabel.config(text=self.ls[self.i],font=("B Yekan",25))
                                            self.blabel.config(text=" ",font=("B Yekan",40))
                                            self.blabel.grid(column=1, row=1)
                                            self.blabel.grid(column=1, row=2)
                                            self.blabel.grid(column=1, row=3)
                                            self.mylabel.grid(column=5, row=4)
                                            #self.s(self.root, text=self.ls[self.i],font=("B Yekan",25)).grid(column=2, row=3)
                                            ttk.Button(self.root, text="حاضر", command=partial(self.setmark,"حاضر")).grid(column=1, row=4)
                                            ttk.Button(self.root, text="غایب",command=partial(self.setmark,"غایب")).grid(column=2, row=4)
                                            ttk.Button(self.root, text="با تاخیر", command=partial(self.setmark,"با تاخیر")).grid(column=3, row=4)
                                            ttk.Button(self.root, text="اخراج از کلاس", command=partial(self.setmark,"اخراج از کلاس")).grid(column=4, row=4)
                                            self.root.mainloop()
                                # this method will add mark to mylist and call next student 
                                def setmark(self,mark):
                                                            self.mylist.append(mark)
                                                            self.nextstudent()
                                            
                                #this method will show next student
                                def nextstudent(self):
                                    
                                    if (self.i<len(self.ls)-1):
                                                            
                                                            self.i+=1
                                                            self.mylabel.config(text=self.ls[self.i],font=("B Yekan",25))
                                                            self.mylabel.grid(column=5, row=4)
                                    else:    
                            
                                        self.mylabel.config(text="حضور و غیاب تمام شد",font=("Calibri",25))
                                        self.mylabel.grid(column=5, row=4)
                                        messagebox.showinfo("حضور و غیاب","حضور و غیاب تمام شد")
                                        self.writetoexcel()
                                        self.root.destroy()
                                        self.root.mainloop()
                                        robo()
                                # this method writes mylist to excel file
                                def writetoexcel(self):
                                    clm=self.ws.max_column+1
                                    i=1
                                    for x in range(len(self.mylist)):
                                        
                                        t=self.ws.cell(row=i,column=clm)
                                        t.value=self.mylist[x]
                                        
                                        
                                        
                                        i=i+1
                                    print("sssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssss")
                                    self.wb.save(self.exfilename)
                                # this method gets systemdate and convert is to persian
                                def getdate(self):
                                    from datetime import datetime
                                    today=datetime.today()
                                    return str(JalaliDate(today))
                            # ----- class finished lets test
                            
                            
                            ob=StudentProject()
                            ob.showgui()
                    except:
                        messagebox.showerror("Error","کلاس وجود ندارد")
            else:
                
                        try:            
                            

                            # we will use these two files for reading excel file
                            # multiple methods will use it so I put them here
                            wb = load_workbook(filename = exfilename)
                            ws=wb.active
                            root.destroy()
                            import csv
                            import tkinter
                            class StudentProject:
                                exfilename=desktop_path+"\\"+folder_name+"\\"+classe+".xlsx"
                                # we will use these two files for reading excel file
                                # multiple methods will use it so I put them here
                                wb = load_workbook(filename = exfilename)
                                ws=wb.active
                                # ls will present the all students in excel file
                                ls=list()
                                # mylist is the list which we will save to new coulumn
                                mylist=list()
                                # this is a counter for our list 
                                i=0
                                # here are some variables for user interface
                                root=Tk()
                                root.iconbitmap('Icons\icon.ico')
                                root.title("presence and absence")
                                mylabel=ttk.Label()
                                # this is empty label which will be use in grid
                                blabel=ttk.Label()
                                def __init__(self):
                                    # for start lets get student names 
                                    self.ls=self.getstudentnames()
                                    # and persian current date
                                    self.mylist.append(self.getdate())
                                   
                                # this methid will get all student names from excel 
                                # this method returns a list named ls 
                                def getstudentnames(self):
                                    ls=list()
                                    rownum=self.ws.max_row
                                    for i in range(2,rownum+1):
                                      ls.append(self.ws.cell(i,2).value)
                                    return ls
                                # this method showes gui 
                                # I tried to write some code but needs improve     
                                def showgui(self):
                                            from datetime import datetime
                                            today=datetime.today()
                                            import tkinter.messagebox as messagebox
                                            window_height = 250
                                            window_width = 600
                                            screen_width = self.root.winfo_screenwidth()
                                            screen_height = self.root.winfo_screenheight()
                                            x_cordinate = int((screen_width/2) - (window_width/2))
                                            y_cordinate = int((screen_height/2) - (window_height/2))
                                            self.root.geometry("{}x{}+{}+{}".format(window_width, window_height, x_cordinate, y_cordinate))
                                            self.mylabel.config(text=self.ls[self.i],font=("Calibri",30))
                                            self.blabel.config(text=" ",font=("Calibri",40))
                                            self.blabel.grid(column=1, row=1)
                                            self.blabel.grid(column=1, row=2)
                                            self.blabel.grid(column=1, row=3)
                                            self.mylabel.grid(column=5, row=4)
                                            #self.s(self.root, text=self.ls[self.i],font=("B Yekan",25)).grid(column=2, row=3)
                                            ttk.Button(self.root, text="حاضر", command=partial(self.setmark,"حاضر")).grid(column=1, row=4)
                                            ttk.Button(self.root, text="غایب",command=partial(self.setmark,"غایب")).grid(column=2, row=4)
                                            ttk.Button(self.root, text="با تاخیر", command=partial(self.setmark,"با تاخیر")).grid(column=3, row=4)
                                            ttk.Button(self.root, text="اخراج از کلاس", command=partial(self.setmark,"اخراج از کلاس")).grid(column=4, row=4)
                                            self.root.mainloop()
                                # this method will add mark to mylist and call next student 
                                def setmark(self,mark):
                                                            self.mylist.append(mark)
                                                            self.nextstudent()
                                            
                                #this method will show next student
                                def nextstudent(self):
                                    
                                    if (self.i<len(self.ls)-1):
                                                            
                                                            self.i+=1
                                                            self.mylabel.config(text=" "+self.ls[self.i],font=("Calibri",30))
                                                            self.mylabel.grid(column=5, row=4)
                                    else:    
                            
                                        self.mylabel.config(text="حضور و غیاب تمام شد",font=("Calibri",30))
                                        self.mylabel.grid(column=5, row=4)
                                        import time
                                        messagebox.showinfo("حضور و غیاب","حضور و غیاب تمام شد")
                                        self.writetoexcel()
                                        self.root.destroy()
                                        self.root.mainloop()
                                        robo()
                                        
                                # this method writes mylist to excel file
                                def writetoexcel(self):
                                    clm=self.ws.max_column+1
                                    i=1
                                    for x in range(len(self.mylist)):
                                        
                                        t=self.ws.cell(row=i,column=clm)
                                        t.value=self.mylist[x]
                                        
                                        
                                        
                                        i=i+1
                                    
                                    self.wb.save(self.exfilename)
                                # this method gets systemdate and convert is to persian
                                def getdate(self):
                                    from datetime import datetime
                                    today=datetime.today()
                                    return str(JalaliDate(today))
                            # ----- class finished lets test
                            
                            
                            ob=StudentProject()
                            ob.showgui()
                            
                        except:
                                print()
                                messagebox.showerror("Error","کلاس وجود ندارد")

        # ایجاد متن
    record_attendance()


    # اجرای برنامه

    root.mainloop()
#################################################################################
def showsleep():    
    import sys
    if sys.version_info[0] == 2:
        import Tkinter as tk
    else:
        import tkinter as tk
    from time import sleep, strftime
    import datetime
    import feedparser
    import time
    from itertools import cycle
    dark = {1:16, 2:17, 3:18, 4:19, 5:19, 6:20, 7:20, 8:19, 9:18, 10:17, 11:16, 12:16}
    #month_number : sunrise_hour
    light = {1:8, 2:7, 3:6, 4:5, 5:4, 6:4, 7:4, 8:5, 9:6, 10:6, 11:7, 12:8}
    root = tk.Tk()
    current_theme = "not set"
    d = feedparser.parse('https://www.shahrekhabar.com/')
    post_list = cycle(d.entries)
    def exitt():
        sleep(3)
        root.destroy()
    def theme_updater(theme_bg, theme_fg):
        clock.config(bg=theme_bg, fg=theme_fg)
        extra_clock.config(bg=theme_bg, fg=theme_fg)
        label_rss.config(bg=theme_bg, fg=theme_fg)
        end.config(fg=theme_fg, bg=theme_bg)
    time1 = ''
    extra_time1 = ''
    clock = tk.Label(root, font=('calibri light', 150))
    clock.pack(fill=tk.BOTH, expand=1)
    extra_clock = tk.Label(root, font=('calibri light', 45))
    extra_clock.pack(fill=tk.BOTH, expand=1)
    label_rss = tk.Label(root, font=('calibri', 14))
    label_rss.pack(fill=tk.BOTH)
    end = tk.Button(root,width=6,text="Exit",border=0,cursor='hand2',fg='#57a1f8' ,command=lambda:exitt(), font=('bold', 18))
    end.pack(fill=tk.BOTH)
    def rssfeeds():
        post = (post_list)
        RSSFEED = post
        modTXT = 'Welcom. it is Sasan Hooti  '
        label_rss.config(text=modTXT)
        root.after(8000, rssfeeds)
    rssfeeds()
    def tick():
     global current_theme
     global time1
     time2 = strftime("%H:%M:%S")
     if time2 != time1:
      time1 = time2
      clock.config(text=time2)
     time_now = time.localtime()
     if time_now.tm_hour <= dark[time_now.tm_mon] or time_now.tm_hour < light[time_now.tm_mon]:
            if current_theme != "night":
                theme_updater("white", "black")
                current_theme = "night"
                print('Night time theme set.')
     else:
            if current_theme != "day":
                theme_updater("black", "white")
                current_theme = "day"
                print('Day time theme set.')
     clock.after(1000, tick)
    def ticki():
     global extra_time1
     extra_time2 = strftime("%A, %d %B %Y")
     if extra_time2 != extra_time1:
      extra_time1 = extra_time2
      extra_clock.config(text=extra_time2)
     extra_clock.after(1000, ticki)
    tick()
    ticki()
    w, h = root.winfo_screenwidth(), root.winfo_screenheight()
    root.overrideredirect(1)
    root.geometry("%dx%d+0+0" % (w, h))
    root.focus_set()
    root.mainloop()
#################################################################################
def mus1():
    import pygame
    import os
    import customtkinter
    class MusicPlayer:
        
        def __init__(self,root):
            self.root = root
            self.root.title("Music Player")
            self.root.iconbitmap(bitmap = 'Icons\\icon.ico')
            self.root.geometry("1000x200")
            pygame.init()
            pygame.mixer.init()
            self.track = StringVar()
            self.status = StringVar()
            trackframe = customtkinter.CTkScrollableFrame(self.root,width=600,height=100,label_text="Song Track")
            trackframe.place(x=0,y=0)
            songtrack = customtkinter.CTkLabel(trackframe,textvariable=self.track,width=20).grid(row=0,column=0,padx=10,pady=5)
            trackstatus = customtkinter.CTkLabel(trackframe,textvariable=self.status).grid(row=0,column=1,padx=10,pady=5)
            buttonframe = customtkinter.CTkScrollableFrame(self.root,width=600,height=100,label_text="Control Panel")
            buttonframe.place(x=0,y=100)
            playbtn = customtkinter.CTkButton(buttonframe,text="PLAY",command=self.playsong,width=6,font=("arial",16,"bold")).grid(row=0,column=0,padx=10,pady=5)
            playbtn = customtkinter.CTkButton(buttonframe,text="PAUSE",command=self.pausesong,width=8,font=("arial",16,"bold")).grid(row=0,column=1,padx=10,pady=5)
            playbtn = customtkinter.CTkButton(buttonframe,text="UNPAUSE",command=self.unpausesong,width=10,font=("arial",16,"bold")).grid(row=0,column=2,padx=10,pady=5)
            playbtn = customtkinter.CTkButton(buttonframe,text="STOP",command=self.stopsong,width=6,font=("arial",16,"bold")).grid(row=0,column=3,padx=10,pady=5)
            songsframe = customtkinter.CTkScrollableFrame(self.root,width=400,height=200,label_text="Song Playlist")
            songsframe.place(x=600,y=0)
            scroll_y = customtkinter.CTkScrollbar(songsframe)
            self.playlist = Listbox(songsframe,yscrollcommand=scroll_y.set,selectmode=SINGLE,font=("arial",12,"bold"))
            scroll_y.pack(side=RIGHT,fill=Y)
            scroll_y.configure(command=self.playlist.yview)
            self.playlist.pack(fill=BOTH)
            from tkinter.filedialog import askdirectory
            directory = askdirectory()
            os.chdir(directory)
            songtracks = os.listdir()
            try:
                for track in songtracks:
                    self.playlist.insert(END,track)
            except:
                import tkinter.messagebox as messagebox
                messagebox.showerror("Error","No File")
        def playsong(self):
            try:
                self.track.set(self.playlist.get(ACTIVE))
                self.status.set("-Playing")
                pygame.mixer.music.load(self.playlist.get(ACTIVE))
                pygame.mixer.music.play()
            except:
                import tkinter.messagebox as messagebox
                messagebox.showerror("Error","No File")
        def stopsong(self):
            self.status.set("-Stopped")
            pygame.mixer.music.stop()
    
        def pausesong(self):
            self.status.set("-Paused")
            pygame.mixer.music.pause()
    
        def unpausesong(self):
            self.status.set("-Playing")
            pygame.mixer.music.unpause()
    root = Tk()
    MusicPlayer(root)
    root.mainloop()
#################################################################################
def Tic():
   size_of_board = 600
   symbol_size = (size_of_board / 3 - size_of_board / 8) / 2
   symbol_thickness = 50
   symbol_X_color = '#EE4035'
   symbol_O_color = '#0492CF'
   Green_color = '#7BC043'
   import numpy as np
   
   class Tic_Tac_Toe():
       # ------------------------------------------------------------------
       # Initialization Functions:
       # ------------------------------------------------------------------
       def __init__(self):
           self.window = Tk()
           self.window.title('Tic-Tac-Toe')
           
           self.window.iconbitmap(bitmap = 'Icons\\icon.ico')
           self.canvas = Canvas(self.window, width=size_of_board, height=size_of_board)
           self.canvas.pack()
           # Input from user in form of clicks
           self.window.bind('<Button-1>', self.click)
   
           self.initialize_board()
           self.player_X_turns = True
           self.board_status = np.zeros(shape=(3, 3))
   
           self.player_X_starts = True
           self.reset_board = False
           self.gameover = False
           self.tie = False
           self.X_wins = False
           self.O_wins = False
   
           self.X_score = 0
           self.O_score = 0
           self.tie_score = 0
   
       def mainloop(self):
           self.window.mainloop()
   
       def initialize_board(self):
           for i in range(2):
               self.canvas.create_line((i + 1) * size_of_board / 3, 0, (i + 1) * size_of_board / 3, size_of_board)
   
           for i in range(2):
               self.canvas.create_line(0, (i + 1) * size_of_board / 3, size_of_board, (i + 1) * size_of_board / 3)
   
       def play_again(self):
           self.initialize_board()
           self.player_X_starts = not self.player_X_starts
           self.player_X_turns = self.player_X_starts
           self.board_status = np.zeros(shape=(3, 3))
   
       # ------------------------------------------------------------------
       # Drawing Functions:
       # The modules required to draw required game based object on canvas
       # ------------------------------------------------------------------
   
       def draw_O(self, logical_position):
           logical_position = np.array(logical_position)
           # logical_position = grid value on the board
           # grid_position = actual pixel values of the center of the grid
           grid_position = self.convert_logical_to_grid_position(logical_position)
           self.canvas.create_oval(grid_position[0] - symbol_size, grid_position[1] - symbol_size,
                                   grid_position[0] + symbol_size, grid_position[1] + symbol_size, width=symbol_thickness,
                                   outline=symbol_O_color)
   
       def draw_X(self, logical_position):
           grid_position = self.convert_logical_to_grid_position(logical_position)
           self.canvas.create_line(grid_position[0] - symbol_size, grid_position[1] - symbol_size,
                                   grid_position[0] + symbol_size, grid_position[1] + symbol_size, width=symbol_thickness,
                                   fill=symbol_X_color)
           self.canvas.create_line(grid_position[0] - symbol_size, grid_position[1] + symbol_size,
                                   grid_position[0] + symbol_size, grid_position[1] - symbol_size, width=symbol_thickness,
                                   fill=symbol_X_color)
   
       def display_gameover(self):
   
           if self.X_wins:
               self.X_score += 1
               text = 'Winner: Player 1 (X)'
               color = symbol_X_color
           elif self.O_wins:
               self.O_score += 1
               text = 'Winner: Player 2 (O)'
               color = symbol_O_color
           else:
               self.tie_score += 1
               text = 'Its a tie'
               color = 'gray'
   
           self.canvas.delete("all")
           self.canvas.create_text(size_of_board / 2, size_of_board / 3, font="cmr 60 bold", fill=color, text=text)
   
           score_text = 'Scores \n'
           self.canvas.create_text(size_of_board / 2, 5 * size_of_board / 8, font="cmr 40 bold", fill=Green_color,
                                   text=score_text)
   
           score_text = 'Player 1 (X) : ' + str(self.X_score) + '\n'
           score_text += 'Player 2 (O): ' + str(self.O_score) + '\n'
           score_text += 'Tie                    : ' + str(self.tie_score)
           self.canvas.create_text(size_of_board / 2, 3 * size_of_board / 4, font="cmr 30 bold", fill=Green_color,
                                   text=score_text)
           self.reset_board = True
   
           score_text = 'Click to play again \n'
           self.canvas.create_text(size_of_board / 2, 15 * size_of_board / 16, font="cmr 20 bold", fill="gray",
                                   text=score_text)
   
       # ------------------------------------------------------------------
       # Logical Functions:
       # The modules required to carry out game logic
       # ------------------------------------------------------------------
   
       def convert_logical_to_grid_position(self, logical_position):
           logical_position = np.array(logical_position, dtype=int)
           return (size_of_board / 3) * logical_position + size_of_board / 6
   
       def convert_grid_to_logical_position(self, grid_position):
           grid_position = np.array(grid_position)
           return np.array(grid_position // (size_of_board / 3), dtype=int)
   
       def is_grid_occupied(self, logical_position):
           if self.board_status[logical_position[0]][logical_position[1]] == 0:
               return False
           else:
               return True
   
       def is_winner(self, player):
   
           player = -1 if player == 'X' else 1
   
           # Three in a row
           for i in range(3):
               if self.board_status[i][0] == self.board_status[i][1] == self.board_status[i][2] == player:
                   return True
               if self.board_status[0][i] == self.board_status[1][i] == self.board_status[2][i] == player:
                   return True
   
           # Diagonals
           if self.board_status[0][0] == self.board_status[1][1] == self.board_status[2][2] == player:
               return True
   
           if self.board_status[0][2] == self.board_status[1][1] == self.board_status[2][0] == player:
               return True
   
           return False
   
       def is_tie(self):
   
           r, c = np.where(self.board_status == 0)
           tie = False
           if len(r) == 0:
               tie = True
   
           return tie
   
       def is_gameover(self):
           # Either someone wins or all grid occupied
           self.X_wins = self.is_winner('X')
           if not self.X_wins:
               self.O_wins = self.is_winner('O')
   
           if not self.O_wins:
               self.tie = self.is_tie()
   
           gameover = self.X_wins or self.O_wins or self.tie
   
           if self.X_wins:
               print('X wins')
           if self.O_wins:
               print('O wins')
           if self.tie:
               print('Its a tie')
   
           return gameover
   
   
   
   
   
       def click(self, event):
           grid_position = [event.x, event.y]
           logical_position = self.convert_grid_to_logical_position(grid_position)
   
           if not self.reset_board:
               if self.player_X_turns:
                   if not self.is_grid_occupied(logical_position):
                       self.draw_X(logical_position)
                       self.board_status[logical_position[0]][logical_position[1]] = -1
                       self.player_X_turns = not self.player_X_turns
               else:
                   if not self.is_grid_occupied(logical_position):
                       self.draw_O(logical_position)
                       self.board_status[logical_position[0]][logical_position[1]] = 1
                       self.player_X_turns = not self.player_X_turns
   
               # Check if game is concluded
               if self.is_gameover():
                   self.display_gameover()
                   # print('Done')
           else:  # Play Again
               self.canvas.delete("all")
               self.play_again()
               self.reset_board = False
   
   
   game_instance = Tic_Tac_Toe()
   game_instance.mainloop()
#################################################################################
def mors1():
                import time
                import os
                import tkinter.messagebox as messagebox
                import winsound
                
                root = Tk() 
                root.title("mourse code") 
                root.iconbitmap(bitmap = 'Icons\\icon.ico')
                root.geometry("350x200") 
                def morss():
                
                        MourseCode = { 'A':'.-',   'B':'-...',   'C':'-.-',   'D':'-..',   'E':'.',   'F':'..-.',
                            'G':'--.',   'H':'....',   'I':'..',   'J':'.--',   'K':'-.-',   'L':'.-..', 
                            'M':'--',   'N':'-.',   'O':'---',   'P':'.--.',   'Q':'--.-',   'R':'.-.', 
                            'S':'...',   'T':'-',   'U':'..-',   'V':'...-',   'W':'.--',   'X':'-..-', 
                            'Y':'-.--',   'Z':'--..', ' ':'   ',
                            '0':'-----',   '1':'.----',   '2':'..---',   '3':'...--',   '4':'....-',   '5':'.....', 
                            '6':'-....',   '7':'--...',   '8':'---..',   '9':'----.',
                            }
                        MyStr=labE.get()
                        MyStr=MyStr.upper()
                        Mourse=''
                        for c in MyStr:
                            Mourse=Mourse+MourseCode[c]+' '
                        f=(Mourse)
                        for m in Mourse:
                            if m=='.':
                                winsound.Beep(800,500)
                            elif m=='-':
                                winsound.Beep(800,1500)
                            elif m==" ":
                                time.sleep(1)
                        ls.configure(text="your moursecode is : "+f)
                import customtkinter  
                lab = customtkinter.CTkLabel(root,text="Enter text :")
                lab.place(relx=0.1, rely=0.1, anchor=CENTER)
                labE = customtkinter.CTkEntry(root)
                labE.place(relx=0.4, rely=0.1, anchor=CENTER)
                bu = customtkinter.CTkButton(root,text="Enter",width=3,command=morss)
                bu.place(relx=0.7, rely=0.1, anchor=CENTER)
                ls = customtkinter.CTkLabel(root,text="")
                ls.place(relx=0.5, rely=0.3, anchor=CENTER)
                
                root.mainloop()
#################################################################################
def weather1():
                 
                 
                import requests
                root = Tk() 
                root.title("weather") 
                root.iconbitmap(bitmap = 'Icons\\icon.ico')
                root.geometry("350x300")
                def weather():
                    try:
                        api_key="1ceaad7d14b2cb4f9403429d6e09a147"
                        base_url="https://api.openweathermap.org/data/2.5/weather?q="
                        cityname=enter.get()
                        url=base_url + cityname + "&appid=" + api_key
                    #auto=(url)
                    
                        result = requests.get(url)
                        data = result.json()
                    #print(data)
                        if data["cod"] == "404":
                            f=("city not found")
                            j=""
                            k=""
                            d=""
                            
                        else:
                             
                            a=data["main"]
                        #print(a)
                        
                            t=a["temp"]
                            tc = t - 273.15
                            f=(str(round(tc,2)) + "°")
                        #print(int(tc) + "°")
                        
                            p=a["pressure"]
                            j=(str(p) + " HPA")
                        
                            h=a["humidity"]
                            k=(str(h) + "%RH")
                        
                            w=data["weather"]
                            d=w[0]["description"]
                            d=(d) 
                        l.configure(text="temperature : "+f) 
                        b.configure(text="wind pressure : "+j) 
                        c.configure(text="humidity : "+k)
                        v.configure(text="shape of air : "+d)
                    except:
                        pass
                        
                        
                welcomL = Label(root,text="weather",font="Times 20 italic bold", width=len("Please register or Login to your account")+10,pady=15)
                welcomL.place(relx=0.5, rely=0.1, anchor=CENTER)
                label = Label(root,text="Enter city or country or continent:")
                label.place(relx=0.18, rely=0.22, anchor=CENTER)
                enter = Entry(root, bg="#AEAEAE")
                enter.place(relx=0.58, rely=0.22, anchor=CENTER)
                butto = Button(root,text="Enter", command=weather)
                butto.place(relx=0.58, rely=0.32, anchor=CENTER)
                l = Label(root)
                b = Label(root)
                c = Label(root)
                v = Label(root)
                k = Label(root)
                l.place(relx=0.5, rely=0.52, anchor=CENTER)
                b.place(relx=0.5, rely=0.62, anchor=CENTER)
                c.place(relx=0.5, rely=0.72, anchor=CENTER)
                v.place(relx=0.5, rely=0.82, anchor=CENTER)
                k.place(relx=0.5, rely=0.92, anchor=CENTER)
             
                
                root.mainloop()
#################################################################################
def rock1():
                
                 
                import tkinter.messagebox as messagebox
                import os
                import random
                from PIL import ImageTk, Image
                root= Tk()
                root.geometry("350x300")
                root.iconbitmap(bitmap = 'Icons\\icon.ico')
                root.title("rock game")
                labalchoes = Label(root,text="Enter Choes :")
                labalchoes.place(relx=0.1, rely=0.50, anchor=CENTER)
                choesE = Entry(root)
                choesE.place(relx=0.40, rely=0.50, anchor=CENTER)
                labalhalat = Label(root,text="rock")
                labalhalat1 = Label(root,text="scissors")
                labalhalat2 = Label(root,text="paper")
                labalhalat.place(relx=0.1, rely=0.1, anchor=CENTER)
                labalhalat1.place(relx=0.1, rely=0.2, anchor=CENTER)
                labalhalat2.place(relx=0.1, rely=0.3, anchor=CENTER)
                image = Image.open("Icons\saman.ico")
                image = image.resize((100, 100), Image.ANTIALIAS)
                userIMG = ImageTk.PhotoImage(file=image)
                img_label = Label(root, image=userIMG)
                img_label.place(relx=0.5, rely=0.13, anchor=CENTER)
                
                pal = 0
                pc = 0   
                possible_actions = ["rock", "paper", "scissors"]
                computer_action = random.choice(possible_actions) 
                def game():
                
                
                        user_action = choesE.get()
                        if user_action == computer_action:
                            f=(f"Both players selected {user_action}. It's a tie!")
                        elif user_action == "rock":
                            if computer_action == "scissors":
                                f=(f"Rock smashes scissors! You win!")
                            else:
                                f=(f"Paper covers rock! You lose.")
                        elif user_action == "paper":
                            if computer_action == "rock":
                                f=(f"Paper covers rock! You win!")
                            else:
                                f=(f"Scissors cuts paper! You lose.")
                        elif user_action == "scissors":
                            if computer_action == "paper":
                                    f=(f"Scissors cuts paper! You win!")
                            else:
                                    f=(f"Rock smashes scissors! You lose.") 
                        else:
                            messagebox.showerror("warning","not choies")
                         
                        user_name.configure(text=f)
                        
                
                        
                         
                
                
                   
                        
                but = Button(root,text="Enter", command=game)
                but.place(relx=0.4, rely=0.6, anchor=CENTER)

                user_name = Label(root)
                user_name.place(relx=0.4, rely=0.8, anchor=CENTER)
                barm = Label(root)
                pacl = Label(root)
                def back():
                    root.destroy()
                    robo()
    
                #button
                
                backB = Button(root,text="back", command=back)
                backB.place(relx=0.8, rely=0.55, anchor=CENTER)
                barm.place(relx=0.9, rely=0.1, anchor=CENTER)
                pacl.place(relx=1, rely=0.1, anchor=CENTER)
                
                
                
                
                root,mainloop()
#################################################################################
def edit_text():
   import tkinter as tk
   from tkinter.filedialog import askopenfilename, asksaveasfilename
   import customtkinter
   def open_file():
       
       """Open a file for editing."""
       filepath = askopenfilename(
           filetypes=[("Text Files", "*.txt"), ("All Files", "*.*")]
       )
       if not filepath:
           return
       txt_edit.delete(1.0, tk.END)
       with open(filepath, "r") as input_file:
           text = input_file.read()
           txt_edit.insert(tk.END, text)
       window.title(f"Thecleverprogrammer - {filepath}")
   
   def save_file():
       """Save the current file as a new file."""
       filepath = asksaveasfilename(
           defaultextension="txt",
           filetypes=[("Text Files", "*.txt"), ("All Files", "*.*")],
       )
       if not filepath:
           return
       with open(filepath, "w") as output_file:
           text = txt_edit.get(1.0, tk.END)
           output_file.write(text)
       window.title(f"Edit Text - {filepath}")
   
   window = tk.Tk()
   window.title("Edit Text")
   window.iconbitmap(bitmap = 'Icons\\icon.ico')
   window.rowconfigure(0, minsize=300, weight=1)
   window.columnconfigure(1, minsize=300, weight=1)
   txt_edit = customtkinter.CTkTextbox(window)
   fr_buttons = tk.Frame(window, relief=tk.RAISED, bd=2)
   btn_open = customtkinter.CTkButton(fr_buttons, text="Open", command=open_file)
   btn_save = customtkinter.CTkButton(fr_buttons, text="Save As...", command=save_file)
   
   btn_open.grid(row=0, column=0, sticky="ew", padx=5, pady=5)
   btn_save.grid(row=1, column=0, sticky="ew", padx=5)
   
   fr_buttons.grid(row=0, column=0, sticky="nsew")
   txt_edit.grid(row=0, column=1, sticky="nsew")
   
   window.mainloop()
   robo()
#################################################################################
def search1():
                import tkinter.messagebox as messagebox
                 
                import webbrowser
                root = Tk()
                root.iconbitmap(bitmap = 'Icons\\icon.ico')
                root.geometry("350x150")
                root.title("search")
                def search():
                    search = searchE.get()
                    if len(search) == 0:
                        messagebox.showwarning("Warning","Fill out every single entery")
                    else:
                        iila = webbrowser.open(search)
                    user_name.configure(text=iila)
             
                        
                lab = Label(root, text="Tipe text").pack()
                searchE = Entry(root)
                searchE.pack()
                butB = Button(root, text="Enter", command=search).pack()

                
                        
                user_name = Label(root)
                user_name.pack()
                        
                root.mainloop()
#################################################################################
def cul1():
                 
                
                
                
                
                #Create a calculator class
                class Calculator:
                
                
                
                
                    def __init__(self, master):
                
                        '''
                        DOCSTRING: Define what to do on initialization
                        '''
                        
                        #Assign reference to the main window of the application
                        self.master = master
                
                        #Add a name to our application
                        master.title("Python Calculator")
                        
                        #Create a line where we display the equation
                        self.equation=Entry(master, width=36, borderwidth=5)
                
                        #Assign a position for the equation line in the grey application window
                        self.equation.grid(row=0, column=0, columnspan=4, padx=10, pady=10)
                
                        #Execute the .creteButton() method
                        self.createButton()
                
                
                
                
                    def createButton(self):
                
                        '''
                        DOCSTRING: Method that creates the buttons
                        INPUT: nothing
                        OUTPUT: creates a button
                        '''
                        
                        #We first create each button one by one with the value we want
                        #Using addButton() method which is described below
                        b0 = self.addButton(0)
                        b1 = self.addButton(1)
                        b2 = self.addButton(2)
                        b3 = self.addButton(3)
                        b4 = self.addButton(4)
                        b5 = self.addButton(5)
                        b6 = self.addButton(6)
                        b7 = self.addButton(7)
                        b8 = self.addButton(8)
                        b9 =  self.addButton(9)
                        b_add = self.addButton('+')
                        b_sub = self.addButton('-')
                        b_mult = self.addButton("×")
                        b_div = self.addButton('÷')
                        b_clear = self.addButton('c')
                        b_equal = self.addButton('=')
                        b_tavan = self.addButton("^")
                        b_baghi = self.addButton("mod")
                        b_edame = self.addButton("div")
                        b_sqrt = self.addButton("r")
                        import math
                        

                        
                        
                
                        #Arrange the buttons into lists which represent calculator rows
                        row1=[b7,b8,b9,b_add]
                        row2=[b4,b5,b6,b_sub]
                        row3=[b1,b2,b3,b_mult]
                        row4=[b_clear,b0,b_equal,b_div]
                        row5=[b_tavan,b_baghi,b_edame,b_sqrt]
                
                        #Assign each button to a particular location on the GUI
                        r=1
                        for row in [row1, row2, row3, row4, row5]:
                            c=0
                            for buttn in row:
                                buttn.grid(row=r, column=c, columnspan=1)
                                c+=1
                            r+=1
                
                
                
                
                    def addButton(self,value):
                
                            '''
                            DOCSTRING: Method to process the creation of a button and make it clickable
                            INPUT: value of the button (1,2,3,4,5,6,7,8,9,0,+,-,*,/,c,=,**)
                            OUTPUT: returns a designed button object
                            '''
                            return Button(self.master, text=value, width=9, command = lambda: self.clickButton(str(value)))
                    
                
                
                
                    def clickButton(self, value):
                        
                        '''
                        DOCSTRING: Method to program the actions that will happen in the calculator after a click of each button
                        INPUT: value of the button (1,2,3,4,5,6,7,8,9,0,+,-,*,/,c,=)
                        OUTPUT: what action will be performed when a particular button is clicked
                        '''
                        
                        #Get the equation that's entered by the user
                        current_equation=str(self.equation.get())
                        
                        #If user clicked "c", then clear the screen
                        if value == 'c':
                            self.equation.delete(-1, END)
                        
                        #If user clicked "=", then compute the answer and display it
                        elif value == '=':
                            answer = str(eval(current_equation))
                            self.equation.delete(-1, END)
                            self.equation.insert(0, answer)
                        elif value == "back":
                            root.destroy()
                        elif value == "^":
                            self.equation.insert((self.clickButton(str("**"))))
                        elif value == "÷":
                            self.equation.insert((self.clickButton(str("/"))))
                        elif value == "×":
                            self.equation.insert((self.clickButton(str("*"))))
                        elif value == "mod":
                            self.equation.insert((self.clickButton(str("%"))))
                        elif value == "div":
                            self.equation.insert((self.clickButton(str("//"))))
                        
                        
                            
                        
                        #If user clicked any other button, then add it to the equation line
                        else:
                            self.equation.delete(0, END)
                            self.equation.insert(-1, current_equation+value)
                        if value == "r":
                            import math
                            mn10 =math.sqrt(int(eval(current_equation)))
                            answer = str(mn10)
                            self.equation.delete(-1, END)
                            self.equation.insert(0, answer)
                
                
                
                #Execution
                if __name__=='__main__':
                    
                    #Create the main window of an application
                
                    root = Tk()
                    
                    #Tell our calculator class to use this window
                    my_gui = Calculator(root)
                    root.iconbitmap(bitmap = 'Icons\\icon.ico')    
                    
                    #Executable loop on the application, waits for user input
                    root.mainloop()                    
#################################################################################
def Profile1():
    
    
    ws  = Tk()
    ws.title('Profil')
    ws.geometry('565x50')
    ws['bg'] = '#AC99F2'
    ws.iconbitmap(bitmap = 'Icons\\icon.ico')
    game_frame = Frame(ws)
    game_frame.pack()
    
    my_game = ttk.Treeview(game_frame)
    
    my_game['columns'] = ('player_id', 'player_name', 'player_Rank', 'player_states', 'player_city', "player_ci","player_cit")
    conn = sqlite3.connect("user1.db")
    c = conn.cursor()
    
    c.execute("SELECT * FROM user_profile")
    results = c.fetchall()
    
    if results:
        for i in results:
            profilename = i[0]
            profikelastname = i[1]
            profileemail = i[2]
            profiletime = i[3]
            profilenumber = i[4]
            profileborn = i[5]
            profilemr = i[6]
    my_game.column("#0", width=0,  stretch=NO)
    my_game.column("player_id",anchor=CENTER, width=80)
    my_game.column("player_name",anchor=CENTER,width=80)
    my_game.column("player_Rank",anchor=CENTER,width=80)
    my_game.column("player_states",anchor=CENTER,width=80)
    my_game.column("player_city",anchor=CENTER,width=80)
    my_game.column("player_cit",anchor=CENTER,width=80)
    my_game.column("player_ci",anchor=CENTER,width=80)
    my_game.heading("#0",text="",anchor=CENTER)
    my_game.heading("player_id",text=("Name"),anchor=CENTER)
    my_game.heading("player_name",text=("Last Name"),anchor=CENTER)
    my_game.heading("player_Rank",text="Email",anchor=CENTER)
    my_game.heading("player_states",text="Time",anchor=CENTER)
    my_game.heading("player_city",text="Number",anchor=CENTER)
    my_game.heading("player_ci",text="Born",anchor=CENTER)
    my_game.heading("player_cit",text="Age",anchor=CENTER)
    import datetime
    now = datetime.datetime.now()
    year_born =  now.year - int(profileborn)
    
    my_game.insert(parent='',index='end',iid=0,text='',
    values=(str(profilemr)+"."+str(profilename),str(profikelastname),str(profileemail),str(profiletime), str(profilenumber),str(profileborn),str(year_born)))
    
    
    my_game.pack()
    
    ws.mainloop()
#################################################################################
def todo():
    import tkinter as tk
    import sqlite3
     
    conn1 = sqlite3.connect("user1.db")
    c = conn1.cursor()
    c.execute("SELECT * FROM user_profile")
    results = c.fetchall()
    if results:
              for i in results:
                  profilename = i[0]
                  profikelastname = i[1]
                  profilemail = i[2]
    class ToDoList:
        def __init__(self, master):
            self.master = master
            self.db = sqlite3.connect('user1.db')
            self.db.execute('''CREATE TABLE IF NOT EXISTS '''+profilename+profilemail+profikelastname+'''
                                (task_id INTEGER PRIMARY KEY AUTOINCREMENT, task_name TEXT,
                                task_date TEXT);
                                ''')
            self.create_widgets()
    
        def create_widgets(self):
            # Create a label for the task name
            customtkinter.CTkLabel(self.master, text="Task: ").grid(row=0)
            # Create a label for the date
            customtkinter.CTkLabel(self.master, text="Task Date (MM/DD/YYYY): ").grid(row=1)
    
            # Create an entry widget for the task name
            self.name_entry = customtkinter.CTkEntry(self.master)
            self.name_entry.grid(row=0, column=1)
    
            # Create an entry widget for the task date
            self.date_entry = customtkinter.CTkEntry(self.master)
            self.date_entry.grid(row=1, column=1)
    
            # Create a button for adding a task
            customtkinter.CTkButton(self.master, text="Add Task", command=self.add_task).grid(row=2, column=0)
    
            # Create a listbox for displaying '''+profilename+profilemail+profikelastname+'''
            self.listbox = tk.Listbox(self.master, height=10, width=50)
            self.listbox.grid(row=3, columnspan=2)
    
            # Create a button for deleting a task
    
            # Call a function to populate the listbox
            self.show_tasks()
    
        def add_task(self):
            name = self.name_entry.get()
            date = self.date_entry.get()
            if name:
                self.db.execute(f"INSERT INTO "+profilename+profilemail+profikelastname+" (task_name, task_date) VALUES (?, ?)",(name,date))
                self.db.commit()
                self.show_tasks()
    
        
        def show_tasks(self):
            self.listbox.delete(0, tk.END)
            for row in self.db.execute("SELECT * FROM "+profilename+profilemail+profikelastname+" ORDER BY task_date"):
                self.listbox.insert(tk.END, f"{row[0]} {row[1]} ({row[2]})")
    
        def __del__(self):
            self.db.close()
    
    if __name__ == '__main__':
        root = tk.Tk()
        my_app = ToDoList(root)
        root.mainloop() 
#################################################################################
def note1():
    import tkinter as tk
    from tkinter import ttk, messagebox
    import json
    from tkinter import ttk
    from tkinter import messagebox
    import sqlite3
    import tkinter as tk
    import tkinter.messagebox as messagebox
    import customtkinter
     
    
    conn = sqlite3.connect('user1.db')
    c = conn.cursor()
        
    c.execute("SELECT * FROM user_profile")
    results = c.fetchall()
    if results:
            for i in results:
                profilename = i[0]
                profikelastname = i[1]
                profileemail = i[2]
    # Create the main window
    root = tk.Tk()
    root.title("Notes App")
    root.geometry("500x500")
    style = ttk.Style()
    root.iconbitmap(bitmap = 'Icons\\icon.ico')
    
    # Configure the tab font to be bold
    style.configure("TNotebook.Tab", font=("TkDefaultFont", 14, "bold"))
    
    # Create the notebook to hold the notes
    notebook = ttk.Notebook(root, style="TNotebook")
    
    # Load saved notes
    notes = {}
    try:
        with open('notebook\\'+profileemail+profikelastname+".json", "r") as f:
            notes = json.load(f)
    except FileNotFoundError:
        pass
    
    # Create the notebook to hold the notes
    notebook = ttk.Notebook(root)
    notebook.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)
    
    # Create a function to add a new note
    def add_note():
        # Create a new tab for the note
        note_frame = ttk.Frame(notebook, padding=10)
        notebook.add(note_frame, text="New Note")
        
        # Create entry widgets for the title and content of the note
        title_label = customtkinter.CTkLabel(note_frame, text="Title:")
        title_label.grid(row=0, column=0, padx=10, pady=10, sticky="W")
        
        title_entry = customtkinter.CTkEntry(note_frame, width=180)
        title_entry.grid(row=0, column=1, padx=10, pady=10)
        
        content_label = customtkinter.CTkLabel(note_frame, text="Content:")
        content_label.grid(row=1, column=0, padx=10, pady=10, sticky="W")
        
        content_entry = tk.Text(note_frame, width=40, height=10)
        content_entry.grid(row=1, column=1, padx=10, pady=10)
        
        # Create a function to save the note
        def save_note():
            # Get the title and content of the note
            title = title_entry.get()
            content = content_entry.get("1.0", tk.END)
            
            # Add the note to the notes dictionary
            notes[title] = content.strip()
            
            # Save the notes dictionary to the file
            with open('notebook\\'+profileemail+profikelastname+".json", "r") as f:
                notes = json.load(f)
                
            
            # Add the note to the notebook
            note_content = tk.Text(notebook, width=40, height=10)
            note_content.insert(tk.END, content)
            notebook.forget(notebook.select())
            notebook.add(note_content, text=title)
            
        # Add a save button to the note frame
        save_button = customtkinter.CTkButton(note_frame, text="Save", 
                                 command=save_note)
        save_button.grid(row=2, column=1, padx=10, pady=10)
    
    def load_notes():
        try:
            with open('notebook\\'+profileemail+profikelastname+".json", "r") as f:
                notes = json.load(f)
    
            for title, content in notes.items():
                # Add the note to the notebook
                note_content = tk.Text(notebook, width=40, height=10)
                note_content.insert(tk.END, content)
                notebook.add(note_content, text=title)
    
        except FileNotFoundError:
            # If the file does not exist, do nothing
            pass
    
    # Call the load_notes function when the app starts
    load_notes()
    
    # Create a function to delete a note
    def delete_note():
        # Get the current tab index
        current_tab = notebook.index(notebook.select())
        
        # Get the title of the note to be deleted
        note_title = notebook.tab(current_tab, "text")
        
        # Show a confirmation dialog
        confirm = messagebox.askyesno("Delete Note", 
                                      f"Are you sure you want to delete {note_title}?")
        
    # Add buttons to the main window
    new_button = customtkinter.CTkButton(root, text="New Note", 
                            command=add_note)
    new_button.pack(side=tk.LEFT, padx=10, pady=10)
    
    delete_button = customtkinter.CTkButton(root, text="Delete", 
                               command=delete_note)
    delete_button.pack(side=tk.LEFT, padx=10, pady=10)
    
    root.mainloop()
#################################################################################
def chatbot():
    import tkinter as tk

    import sqlite3
    conn = sqlite3.connect('user1.db')
    c = conn.cursor()
        
    c.execute("SELECT * FROM user_profile")
    results = c.fetchall()
    
    if results:
            for i in results:
                profilename = i[0]
                profikelastname = i[1]
                profileemail = i[2]
                profiletime = i[3]
    
    # Create the table for storing contact information
    conn.execute('CREATE TABLE IF NOT EXISTS '+profilename+profileemail+profileemail+profileemail+profileemail+' (user text,bot text);')
    from tkinter import scrolledtext
    import openai
    import tkinter as tk
    # insert your API key here
    openai.api_key = "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"
    
    # set up the GPT model
    model_engine = "text-davinci-003"
    prompt = "Hello! I'm a chatbot. How can I help you?"
    # load the conversation data from a JSON file
    
    # create the main window
    window = tk.Tk()
    window.title("Chatbot")
    window.geometry("400x500")
    window.iconbitmap(bitmap = 'Icons\\icon.ico')
    # add a scrolled text widget to display the messages
    chat_log = scrolledtext.ScrolledText(window, state='disabled')
    chat_log.configure(font='TkFixedFont')
    chat_log.pack(side='top', fill='both', expand=True)
    
    # add an input field for the user's messages
    user_input = tk.Entry(window)
    user_input.pack(side='left', fill='x', expand=True)
    
    # add a send button to send messages
    def send_message():
        user_message = user_input.get()
        chat_log.configure(state='normal')
        chat_log.insert('end', "User: " + user_message + "\n")
        chat_log.configure(state='disabled')
        user_input.delete(0,'end')
        generate_response(user_message)
    send_button = tk.Button(window, text='Send', command=send_message)
    send_button.pack(side='right')
    def history():
        for row in c.execute('SELECT user, bot FROM '+profilename+profileemail+profileemail+profileemail+profileemail+''):
            chat_log.configure(state='normal')
            chat_log.insert('end', "User: "+row[0] + "\n")
            chat_log.configure(state='disabled')
            user_input.delete(0,'end')
            chat_log.configure(state='normal')
            chat_log.insert('end', "Chatbot: "+row[1] + "\n")
            chat_log.configure(state='disabled')
    sende_button = tk.Button(window, text='History', command=history)
    sende_button.pack(side='right')
    # create a function to generate a response
    def generate_response(user_message):
        # check if the user's message matches any of the conversation data
    
        # use OpenAI's GPT to generate a response
        response = openai.Completion.create(
          engine=model_engine,
          prompt=prompt + "\nYou: " + user_message,
          temperature=0.7,
          max_tokens=60,
          n=1,
          stop=None
        )
        
        bot_text = response.choices[0].text.strip()
        
        chat_log.configure(state='normal')
        chat_log.insert('end', "Chatbot: " + bot_text + "\n")
        chat_log.configure(state='disabled')
        c.execute('INSERT INTO '+profilename+profileemail+profileemail+profileemail+profileemail+' (user, bot) VALUES (?, ?)', (user_message, bot_text))
        conn.commit()
        
    
    # run the main loop
    window.mainloop()
#################################################################################
def phone():
   import tkinter as tk

   # Initialize the database connection
   conn = sqlite3.connect('user1.db')
   c = conn.cursor()
       
   c.execute("SELECT * FROM user_profile")
   results = c.fetchall()
   
   if results:
           for i in results:
               profilename = i[0]
               profikelastname = i[1]
               profileemail = i[2]
               profiletime = i[3]
   
   # Create the table for storing contact information
   conn.execute('CREATE TABLE IF NOT EXISTS '+profileemail+profilename+' (id INTEGER PRIMARY KEY,name text,phone text);')
   
   # Function to add a new contact to the database
   def add_contact():
       name = name_entry.get()
       phone = phone_entry.get()
       c.execute('INSERT INTO '+profileemail+profilename+' (name, phone) VALUES (?, ?)', (name, phone))
       conn.commit()
       name_entry.delete(0, tk.END)
       phone_entry.delete(0, tk.END)
       display_contacts()
   
   # Function to display contact information in the listbox
   def display_contacts():
       contact_list.delete(0, tk.END)
       for row in c.execute('SELECT name, phone FROM '+profileemail+profilename+''):
           contact_list.insert(tk.END, row[0] + ' - ' + row[1])
   
   # Set up the GUI
   root = tk.Tk()
   root.title('دفتر تلفن')
   
   name_label = customtkinter.CTkLabel(root, text='نام :')
   name_label.grid(row=0, column=0)
   
   name_entry = customtkinter.CTkEntry(root, width=160)
   name_entry.grid(row=0, column=1)
   
   phone_label = customtkinter.CTkLabel(root, text='شماره :')
   phone_label.grid(row=1, column=0)
   
   phone_entry = customtkinter.CTkEntry(root, width=160)
   phone_entry.grid(row=1, column=1)
   
   add_button = customtkinter.CTkButton(root, text='کردن اضافه', command=add_contact)
   add_button.grid(row=2, column=0)
   
   contact_list = tk.Listbox(root, width=50)
   contact_list.grid(row=3, column=0, columnspan=2)
   
   # Populate the list with existing contacts from the database
   display_contacts()
   
   root.mainloop()
   
   # Close the database connection
#################################################################################
def Introduction1():
    root = Tk() 
    root.title("Help") 
    root.iconbitmap(bitmap = 'Icons\\icon.ico')
    root.geometry("330x180") 
    conn = sqlite3.connect('user1.db')
    c = conn.cursor() 
    c.execute("SELECT * FROM Languge")         
    results2 = c.fetchall()
    if results2:
        for i in results2:
            profilename12 = i[0]
    Label(root, text=("سلام بنده ساسان حوتی هستم سازنده روبوسان و ")).pack()
    Label(root, text=("اپلیکیشن های این قبیل این برنامه برای معلمین")).pack()
    Label(root, text=("طراحی شده تا بتوانن ضعف هایی که در حین تدریس ")).pack()
    Label(root, text=("صورت میگیرد تاحد ممکن کم کند و کمک دست معلمین باشد")).pack()
    Label(root, text=("اگر نظری درباره ی برنامه داری که آن را بهبود دهم")).pack()
    Label(root, text=("به این ایدی پیام بدید")).pack()
    Label(root, text=("sasanhooti0@gmail.com")).pack()

    root.mainloop()
#################################################################################
def trans1():
    def sign_in():
        first_name = name.get()
        if male_btn.get() == 1 and female_btn.get() == 0:
            gender = "fa"
            be = "en"
        elif female_btn.get() == 1 and male_btn.get() == 0:
            gender = "en"
            be = "fa"
        elif female_btn.get() == 1 and male_btn.get() == 1:
           messagebox.showwarning("Warning","Choose an option")
        else:
           messagebox.showwarning("Warning","Choose an option") 
        output = GoogleTranslator(source=be,target=gender).translate(first_name)               
        user_name.configure(text=output)
    
    # building a window
    window = Tk()
    
    #adding window title
    window.geometry("350x300")
    window.iconbitmap(bitmap = 'Icons\\icon.ico')
    window.title("trans")
    
    #making a label for first name entry
    Label(window, text="Enter Text").pack()
    
    #making the entry for first name
    name = Entry(window)
    name.pack()
    
    
    
    #making the checkbutton for male
    male_btn = IntVar()
    Checkbutton(window, text="فارسی", variable=male_btn).pack()
    
    #making the checkbutton for female
    female_btn = IntVar()
    Checkbutton(window, text="Engelish", variable=female_btn).pack()
    
    #making an action button
    btn = Button(window,text="Enter",command=sign_in)
    btn.pack()
    
    
    #making a label to be able to say welcome
    user_name = Label(window)
    user_name.pack()
    
    

    
    
    
    #creating the loop for the program
    window.mainloop()
#################################################################################
def help():
    root = Tk() 
    root.title("Help") 
    root.iconbitmap(bitmap = 'Icons\\icon.ico')
    root.geometry("430x110") 
    def pack():
        root.destroy()
        robo()
    labalL = Label(root, text=GoogleTranslator(target="en").translate("""شما مینوانید از منوی بالا صفحه ی اصلی عملیات
                                 را انتخاب کنید تا آن صفحه برای شما باز شود
                             در کادر خالی میتوانید پرسش هایتان را از برنامه بپرسید""")).pack()

    root.mainloop() 
#################################################################################
def helpmain1():
    root = Tk() 
    root.title("Help") 
    root.iconbitmap(bitmap = 'Icons\\icon.ico')
    root.geometry("440x110") 
    labalL = Label(root, text=GoogleTranslator(target="en").translate("1.از دو کلید پایین کلید مورد نظر (ثبت نام)و(وارد شوید) یکی را انتخاب کنید")).pack()
    root.mainloop()    
#################################################################################
def corno():
    
    import tkinter.messagebox as messagebox
    ws = Tk()
    ws.iconbitmap(bitmap = 'Icons\icon.ico')
    ws.geometry("300x200")
    
    ws.title("countdown")
    
    Hour=StringVar()
    Minute=StringVar()
    Second=StringVar()
    
    Hour.set("00")
    Minute.set("00")
    Second.set("00")
    Hour_entry= customtkinter.CTkEntry(ws,width=3,  font=("Arial",30),textvariable=Hour)
    Hour_entry.place(x=80,y=20)
    
    Minute_entry= customtkinter.CTkEntry(ws,  font=("Arial",30),width=3,
    				textvariable=Minute)
    Minute_entry.place(x=130,y=20)
    
    Second_entry= customtkinter.CTkEntry(ws,  font=("Arial",30),width=3,
    				textvariable=Second)
    Second_entry.place(x=180,y=20)
    
    
    def OK():
        times = int(Hour_entry.get()) * 3600 + int(Minute_entry.get()) * 60 + int(Second_entry.get())
        while times > -1:
           minute, second = (times // 60, times % 60)
           hour = 0
           if minute > 60:
              hour, minute = (minute // 60, minute % 60)
           Second.set(second)
           Minute.set(minute)
           Hour.set(hour)
     
           # Update the time
           ws.update()
           import time
           time.sleep(1)
           if (times == 0):
              messagebox.showinfo("Time Countdown", "Time up ")
              Second.set('00')
              Minute.set('00')
              Hour.set('00')
           times -= 1

    button = customtkinter.CTkButton(ws, text=' countdown',command= OK)
    button.place(x = 75,y = 110)
    
    ws.mainloop()
#################################################################################
def Gram():
   window = Tk()
   window.title("Gram Pound Ounce")
   window.iconbitmap(bitmap = 'Icons\icon.ico')
   def from_kg():
    gram = float(e2_value.get())*1000
    pound = float(e2_value.get())*2.20462
    ounce = float(e2_value.get())*35.274
    t1.delete("1.0",END)
    t1.insert(END, gram)
    t2.delete("1.0", END)
    t2.insert(END, pound)
    t3.delete("1.0", END)
    t3.insert(END, ounce)

   e1 = customtkinter.CTkLabel(window, text="Input the weight in KG")
   e2_value = StringVar()
   e2 = customtkinter.CTkEntry(window, textvariable=e2_value)
   e3 = customtkinter.CTkLabel(window, text="Gram")
   e4 = customtkinter.CTkLabel(window, text="Pound")
   e5 = customtkinter.CTkLabel(window, text="Ounce")
   
   t1 = customtkinter.CTkTextbox(window, height=5, width=190)
   t2 = customtkinter.CTkTextbox(window, height=5, width=190)
   t3 = customtkinter.CTkTextbox(window, height=5, width=190)
   
   b1 = customtkinter.CTkButton(window, text="Convert", command=from_kg)
   
   e1.grid(row=0, column=0)
   e2.grid(row=0, column=1)
   e3.grid(row=1, column=0)
   e4.grid(row=1, column=1)
   e5.grid(row=1, column=2)
   t1.grid(row=2, column=0)
   t2.grid(row=2, column=1)
   t3.grid(row=2, column=2)
   b1.grid(row=0, column=2)
   
   window.mainloop()

load()